# app.py - Production-ready for Render
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_, and_, text, func
from sqlalchemy.orm import load_only
from werkzeug.exceptions import NotFound
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
import os
import json
import re
import hashlib
import secrets
import sqlite3
import html
from urllib.parse import quote_plus, urlparse, urljoin, parse_qs, unquote
from functools import wraps
import uuid
import logging
import requests as http_req
from logging.handlers import RotatingFileHandler
import requests
import threading
import time
from bs4 import BeautifulSoup

# Simple TTL cache for expensive count queries
_count_cache = {}
def _cached_count(key, query_fn, ttl=300):
    entry = _count_cache.get(key)
    if entry and time.time() - entry[0] < ttl:
        return entry[1]
    val = query_fn()
    _count_cache[key] = (time.time(), val)
    return val

def _fast_customer_count():
    """MAX(rowid) is O(1) in SQLite — reads B-tree root only, no full scan."""
    from sqlalchemy import text
    result = db.session.execute(text('SELECT MAX(rowid) FROM customer_data')).scalar()
    return result or 0

# 60-second search result cache — repeated queries skip the DB entirely
_search_cache = {}

def _scache_get(search_type, term, limit_key):
    key = (search_type, term, limit_key)
    e = _search_cache.get(key)
    return e[1] if (e and time.time() - e[0] < 60) else None

def _scache_set(search_type, term, limit_key, rows, meta):
    serialized = [{
        'id':              getattr(r, 'id', None),
        'name':            getattr(r, 'name', None),
        'contact_number':  getattr(r, 'contact_number', None),
        'ic_number':       getattr(r, 'ic_number', None),
        'address':         getattr(r, 'address', None),
        'email':           getattr(r, 'email', None),
        'additional_data': getattr(r, 'additional_data', None),
        'data_source':     getattr(r, 'data_source', None),
        'upload_id':       getattr(r, 'upload_id', None),
        'created_at':      getattr(r, 'created_at', None),
        'updated_at':      getattr(r, 'updated_at', None),
    } for r in rows]
    _search_cache[(search_type, term, limit_key)] = (time.time(), (serialized, meta))
    # Prune stale entries when cache grows large
    if len(_search_cache) > 500:
        cutoff = time.time() - 60
        for k in list(_search_cache):
            if _search_cache.get(k, (0,))[0] < cutoff:
                _search_cache.pop(k, None)

# In-memory data_source lookup (5,649 unique values — tiny)
_ds_cache = {'ts': 0, 'values': []}
def _get_data_sources():
    """Return all unique data_source values, cached for 10 minutes."""
    if time.time() - _ds_cache['ts'] > 600:
        from sqlalchemy import text
        rows = db.session.execute(
            text("SELECT DISTINCT data_source FROM customer_data WHERE data_source IS NOT NULL AND data_source != '' ORDER BY data_source")
        ).fetchall()
        _ds_cache['values'] = [r[0] for r in rows]
        _ds_cache['ts'] = time.time()
    return _ds_cache['values']

def _match_data_sources(term):
    """Match cached data_source list. Single word: substring. Multi-word: all tokens must appear."""
    term_lower = term.lower()
    tokens = [t for t in term_lower.split() if len(t) >= 2]
    if not tokens:
        return []
    if len(tokens) == 1:
        return [ds for ds in _get_data_sources() if tokens[0] in ds.lower()]
    return [ds for ds in _get_data_sources() if all(t in ds.lower() for t in tokens)]

try:
    from playwright.sync_api import sync_playwright
except Exception:
    sync_playwright = None

# Initialize Flask app
app = Flask(__name__)

# ==================== BACKGROUND JOB TRACKING (file-based) ====================
_JOB_DIR = '/tmp/agentsystem_jobs'
os.makedirs(_JOB_DIR, exist_ok=True)

def _set_job(job_id, **kwargs):
    path = os.path.join(_JOB_DIR, f'{job_id}.json')
    try:
        try:
            with open(path, 'r') as f:
                data = json.load(f)
        except Exception:
            data = {}
        data.update(kwargs)
        with open(path, 'w') as f:
            json.dump(data, f)
    except Exception:
        pass

def _get_job(job_id):
    path = os.path.join(_JOB_DIR, f'{job_id}.json')
    try:
        with open(path, 'r') as f:
            return json.load(f)
    except Exception:
        return {}

# ==================== PRODUCTION ENVIRONMENT DETECTION ====================
IS_RENDER = os.environ.get('RENDER', False)
IS_PRODUCTION = IS_RENDER or os.environ.get('PRODUCTION', False)

# ==================== CUSTOM JINJA2 FILTERS ====================
def from_json(value):
    """Parse JSON string to Python object"""
    try:
        return json.loads(value) if value else {}
    except:
        return {}

# Register custom filters
app.jinja_env.filters['from_json'] = from_json

# ==================== CONFIGURATION ====================
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# Database configuration - use PostgreSQL on Render, SQLite locally
if IS_RENDER:
    # Use PostgreSQL if DATABASE_URL is provided
    database_url = os.environ.get('DATABASE_URL')
    if database_url:
        # Fix for Render PostgreSQL URLs (sometimes start with postgres://)
        if database_url.startswith('postgres://'):
            database_url = database_url.replace('postgres://', 'postgresql://', 1)
        app.config['SQLALCHEMY_DATABASE_URI'] = database_url
        app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
            'pool_pre_ping': True,
            'pool_recycle': 300,
        }
        print(f"[CONFIG] Using PostgreSQL database")
    else:
        # Fallback to SQLite in /tmp (temporary - data will be lost on restart)
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:////tmp/agent_system.db'
        print(f"[CONFIG] Using SQLite in /tmp (temporary storage)")
else:
    # Local development
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///agent_system.db'
    print(f"[CONFIG] Using local SQLite database")

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SESSION_TIMEOUT_MINUTES'] = int(os.environ.get('SESSION_TIMEOUT_MINUTES', '30'))
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=int(os.environ.get('SESSION_TIMEOUT_MINUTES', '30')))

# Upload configuration - use /tmp on Render for temporary storage
if IS_RENDER:
    app.config['UPLOAD_FOLDER'] = '/tmp/uploads'
    logs_dir = '/tmp/logs'
else:
    app.config['UPLOAD_FOLDER'] = 'uploads'
    logs_dir = 'logs'

app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
app.config['ENRICHMENT_CRAWL_MODE'] = os.environ.get('ENRICHMENT_CRAWL_MODE', 'auto').strip().lower()
app.config['ENRICHMENT_PLAYWRIGHT_HEADLESS'] = os.environ.get(
    'ENRICHMENT_PLAYWRIGHT_HEADLESS',
    '1' if IS_PRODUCTION else '0'
) == '1'
app.config['ENRICHMENT_PLAYWRIGHT_CHANNEL'] = os.environ.get('ENRICHMENT_PLAYWRIGHT_CHANNEL', 'chrome').strip().lower()

# Ensure directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(logs_dir, exist_ok=True)
SUSPECT_IC_FLAGS_FILE = os.path.join(app.config['UPLOAD_FOLDER'], 'suspect_ic_review_cases.jsonl')

# Initialize extensions
db = SQLAlchemy(app)

# SQLite performance: WAL mode allows concurrent reads during writes,
# NORMAL sync is safe with WAL, 8 MB page cache reduces disk I/O
from sqlalchemy import event as _sa_event
from sqlalchemy.engine import Engine as _Engine
import sqlite3 as _sqlite3_mod

@_sa_event.listens_for(_Engine, "connect")
def _set_sqlite_pragmas(dbapi_conn, _rec):
    if isinstance(dbapi_conn, _sqlite3_mod.Connection):
        cur = dbapi_conn.cursor()
        cur.execute("PRAGMA journal_mode=WAL")
        cur.execute("PRAGMA synchronous=NORMAL")
        cur.execute("PRAGMA cache_size=-131072")   # 128 MB page cache (was 8 MB)
        cur.execute("PRAGMA mmap_size=2147483648") # 2 GB memory-mapped I/O
        cur.execute("PRAGMA temp_store=MEMORY")
        cur.execute("PRAGMA wal_autocheckpoint=1000")
        cur.close()

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Setup logging
file_handler = RotatingFileHandler(os.path.join(logs_dir, 'agent_system.log'), maxBytes=10240, backupCount=10)
file_handler.setFormatter(logging.Formatter(
    '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
))
file_handler.setLevel(logging.INFO)
app.logger.addHandler(file_handler)
app.logger.setLevel(logging.INFO)

# ==================== DATABASE MODELS ====================

class User(UserMixin, db.Model):
    __tablename__ = 'users'
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    full_name = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    
    login_logs = db.relationship('LoginLog', backref='user', lazy=True)
    search_logs = db.relationship('SearchLog', backref='user', lazy=True)
    
    password_changed_at = db.Column(db.DateTime, default=datetime.utcnow)
    must_change_password = db.Column(db.Boolean, default=False)
    credit_balance = db.Column(db.Integer, default=0)
    failed_login_attempts = db.Column(db.Integer, default=0)
    locked_until = db.Column(db.DateTime, nullable=True)
    telegram_chat_id = db.Column(db.String(50), nullable=True)
    tg_link_token = db.Column(db.String(20), nullable=True)
    max_devices = db.Column(db.Integer, nullable=True)
    is_resigned = db.Column(db.Boolean, default=False)
    resigned_at = db.Column(db.DateTime, nullable=True)
    pw_reset_count = db.Column(db.Integer, default=0)
    pw_reset_window_start = db.Column(db.DateTime, nullable=True)

    def is_locked_out(self):
        if self.locked_until and datetime.utcnow() < self.locked_until:
            return True
        return False

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
        self.password_changed_at = datetime.utcnow()
        self.must_change_password = False

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def is_password_expired(self):
        policy = SystemSettings.get('password_expiry_days')
        if not policy or int(policy) == 0:
            return False
        expiry_days = int(policy)
        if not self.password_changed_at:
            return True
        return (datetime.utcnow() - self.password_changed_at).days >= expiry_days

class CustomerData(db.Model):
    __tablename__ = 'customer_data'
    
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200))
    contact_number = db.Column(db.String(50), index=True)
    ic_number = db.Column(db.String(50), index=True)
    address = db.Column(db.Text)
    email = db.Column(db.String(120))
    additional_data = db.Column(db.Text)
    data_source = db.Column(db.String(200), index=True)
    upload_id = db.Column(db.Integer, db.ForeignKey('uploads.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    upload = db.relationship('Upload', backref='customers')

class CustomerEnrichment(db.Model):
    __tablename__ = 'customer_enrichments'

    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer_data.id'), nullable=False, index=True)
    platform = db.Column(db.String(50))
    source_url = db.Column(db.String(500))
    source_type = db.Column(db.String(50), default='public_web')
    matched_name = db.Column(db.String(200))
    matched_location = db.Column(db.String(200))
    matched_company = db.Column(db.String(200))
    matched_title = db.Column(db.String(200))
    matched_phone = db.Column(db.String(50))
    matched_email = db.Column(db.String(120))
    confidence_score = db.Column(db.Integer, default=50)
    review_status = db.Column(db.String(20), default='pending')
    notes = db.Column(db.Text)
    raw_data = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))

    customer = db.relationship('CustomerData', backref='enrichments')
    reviewer = db.relationship('User', foreign_keys=[created_by])

class CustomerEnrichmentJob(db.Model):
    __tablename__ = 'customer_enrichment_jobs'

    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer_data.id'), nullable=False, index=True)
    requested_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    mode = db.Column(db.String(20), default='auto')
    status = db.Column(db.String(20), default='queued', index=True)
    message = db.Column(db.String(300))
    created_count = db.Column(db.Integer, default=0)
    checked_count = db.Column(db.Integer, default=0)
    providers_used = db.Column(db.Text, default='[]')
    error_log = db.Column(db.Text, default='[]')
    debug_samples = db.Column(db.Text, default='[]')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    started_at = db.Column(db.DateTime)
    completed_at = db.Column(db.DateTime)

    customer = db.relationship('CustomerData', backref='enrichment_jobs')
    requester = db.relationship('User', foreign_keys=[requested_by])

class Upload(db.Model):
    __tablename__ = 'uploads'
    
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(200), nullable=False)
    file_path = db.Column(db.String(500))
    upload_date = db.Column(db.DateTime, default=datetime.utcnow)
    admin_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    row_count = db.Column(db.Integer)
    column_count = db.Column(db.Integer)
    columns_found = db.Column(db.Text)
    column_mapping = db.Column(db.Text, default='{}')
    year = db.Column(db.Integer)
    is_current = db.Column(db.Boolean, default=True)
    status = db.Column(db.String(20), default='pending')
    
    admin = db.relationship('User', backref='uploads')

class LoginLog(db.Model):
    __tablename__ = 'login_logs'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    login_time = db.Column(db.DateTime, default=datetime.utcnow)
    logout_time = db.Column(db.DateTime)
    ip_address = db.Column(db.String(45))
    location = db.Column(db.String(200))
    device_info = db.Column(db.String(500))
    fingerprint = db.Column(db.String(64))
    device_category = db.Column(db.String(50))
    connection_type = db.Column(db.String(50))
    session_duration = db.Column(db.Integer)
    session_id = db.Column(db.String(100), unique=True)

class SecurityEventLog(db.Model):
    __tablename__ = 'security_event_logs'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    username = db.Column(db.String(80))
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.String(500))
    event_type = db.Column(db.String(50), nullable=False)
    status = db.Column(db.String(20), nullable=False)
    details = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship('User', foreign_keys=[user_id])

class SearchLog(db.Model):
    __tablename__ = 'search_logs'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    search_term = db.Column(db.String(200))
    search_type = db.Column(db.String(20))
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    results_count = db.Column(db.Integer)
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.String(500))
    screenshot_taken = db.Column(db.Boolean, default=False)
    data_downloaded = db.Column(db.Boolean, default=False)

class ScreenshotLog(db.Model):
    __tablename__ = 'screenshot_logs'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    search_log_id = db.Column(db.Integer, db.ForeignKey('search_logs.id'))
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    screenshot_path = db.Column(db.String(500))
    ip_address = db.Column(db.String(45))

class DataDownloadLog(db.Model):
    __tablename__ = 'data_download_logs'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    search_log_id = db.Column(db.Integer, db.ForeignKey('search_logs.id'))
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    download_format = db.Column(db.String(20))
    ip_address = db.Column(db.String(45))

class MysDownloadLog(db.Model):
    __tablename__ = 'mys_download_logs'

    id            = db.Column(db.Integer, primary_key=True)
    user_id       = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, index=True)
    download_type = db.Column(db.String(10))    # 'single' or 'bulk'
    search_type   = db.Column(db.String(20))
    search_term   = db.Column(db.String(200))
    record_count  = db.Column(db.Integer, default=1)
    ip_address    = db.Column(db.String(45))
    source        = db.Column(db.String(20), default='mys')  # 'mys' or 'jpph'
    timestamp     = db.Column(db.DateTime, default=datetime.utcnow, index=True)

    user = db.relationship('User', foreign_keys=[user_id])

class CreditLog(db.Model):
    __tablename__ = 'credit_logs'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    amount = db.Column(db.Integer, nullable=False)
    balance_after = db.Column(db.Integer, nullable=False)
    reason = db.Column(db.String(200))
    admin_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship('User', foreign_keys=[user_id])
    admin = db.relationship('User', foreign_keys=[admin_id])

class DuplicateRecordLog(db.Model):
    __tablename__ = 'duplicate_records_log'
    
    id = db.Column(db.Integer, primary_key=True)
    upload_id = db.Column(db.Integer, db.ForeignKey('uploads.id'))
    existing_customer_id = db.Column(db.Integer, db.ForeignKey('customer_data.id'))
    duplicate_data = db.Column(db.Text)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    action_taken = db.Column(db.String(50))
    resolution_time = db.Column(db.DateTime)

class SystemSettings(db.Model):
    __tablename__ = 'system_settings'

    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.String(200))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    @staticmethod
    def get(key, default=None):
        s = SystemSettings.query.filter_by(key=key).first()
        return s.value if s else default

    @staticmethod
    def set(key, value):
        s = SystemSettings.query.filter_by(key=key).first()
        if s:
            s.value = str(value)
        else:
            db.session.add(SystemSettings(key=key, value=str(value)))
        db.session.commit()

from data_protection import DataProtection
dp = DataProtection(app, SystemSettings)

class WorkLog(db.Model):
    __tablename__ = 'work_log'
    id          = db.Column(db.Integer, primary_key=True)
    date        = db.Column(db.Date, nullable=False)
    title       = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    category    = db.Column(db.String(50), default='Feature')
    hours       = db.Column(db.Float, default=0.0)
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)

def _seed_work_log():
    from datetime import date as _date
    entries = [
        WorkLog(date=_date(2026,5,31), category='Security',    hours=2.0,
                title='Data Masking Module',
                description='Built data_protection.py with server-side masking for phone, email, IC, name and address. '
                            'Agents see scrambled data on screen and in Excel downloads. '
                            'Admin always sees full data. Toggle persists across restarts via SystemSettings.'),
        WorkLog(date=_date(2026,5,31), category='Security',    hours=3.0,
                title='Panic Mode — Decoy Database',
                description='When activated, all customer searches return realistic fake Malaysian records — real DB never touched. '
                            'Secret URL bookmarks for boss phone (ON/OFF, no login needed). '
                            'Security and Management nav tabs hidden from attacker during panic. '
                            'Panic card hidden from admin UI when active. Telegram /panic /unpanic command support.'),
        WorkLog(date=_date(2026,5,31), category='Feature',     hours=0.5,
                title='Admin Logs — Security Reference Card',
                description='Added a 3-column reference card at the bottom of /admin/logs explaining '
                            'data masking, panic mode behaviour, and how to read logs during a security incident.'),
        WorkLog(date=_date(2026,5,31), category='Security',    hours=1.0,
                title='Panic URL Copy Button & Nav Hide Fix',
                description='Fixed clipboard copy to use textarea fallback (works on plain HTTP). '
                            'Security and Management nav sections now hidden via context processor when panic is active. '
                            'Panic card and status banner hidden from security page while panic is on.'),
        WorkLog(date=_date(2026,5,31), category='Performance', hours=1.0,
                title='SQLite WAL Mode + 60s Search Cache',
                description='Enabled WAL journal mode, NORMAL sync, 8 MB page cache and MEMORY temp store — '
                            'allows concurrent reads during writes, reducing lock contention under agent load. '
                            'Added 60-second in-memory search result cache: repeated identical queries skip the DB entirely. '
                            'Downloads always bypass cache to ensure accurate credit billing.'),
        WorkLog(date=_date(2026,5,31), category='Feature',     hours=1.0,
                title='Work Log & Invoice Page',
                description='Built /admin/work-log: full development work log with category badges, hours tracking, '
                            'add-entry form, and print invoice view. Pre-populated with all work done today. '
                            'Accessible from Management nav.'),
    ]
    db.session.add_all(entries)
    db.session.commit()

class AdminDevice(db.Model):
    __tablename__ = 'admin_devices'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    fingerprint = db.Column(db.String(64), nullable=False)
    ip_address = db.Column(db.String(45))
    label = db.Column(db.String(100))
    status = db.Column(db.String(20), default='pending')
    first_seen = db.Column(db.DateTime, default=datetime.utcnow)
    last_seen = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship('User', foreign_keys=[user_id], backref='admin_devices')

class AdminLoginApproval(db.Model):
    __tablename__ = 'admin_login_approvals'

    id = db.Column(db.Integer, primary_key=True)
    approval_token = db.Column(db.String(64), unique=True, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    fingerprint = db.Column(db.String(64))
    ip_address = db.Column(db.String(45))
    status = db.Column(db.String(20), default='pending')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    expires_at = db.Column(db.DateTime)
    telegram_message_id = db.Column(db.Integer)

    user = db.relationship('User', foreign_keys=[user_id])

class CustomerDeleteApproval(db.Model):
    __tablename__ = 'customer_delete_approvals'

    id = db.Column(db.Integer, primary_key=True)
    approval_token = db.Column(db.String(64), unique=True, nullable=False)
    customer_id = db.Column(db.Integer, nullable=False, index=True)
    requested_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    status = db.Column(db.String(20), default='pending', index=True)
    reason = db.Column(db.String(200))
    snapshot_name = db.Column(db.String(200))
    snapshot_contact_number = db.Column(db.String(50))
    snapshot_ic_number = db.Column(db.String(50))
    snapshot_address = db.Column(db.Text)
    duplicate_name_count = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    expires_at = db.Column(db.DateTime)
    processed_at = db.Column(db.DateTime)
    telegram_message_id = db.Column(db.Integer)

    requester = db.relationship('User', foreign_keys=[requested_by])

class AgentDevice(db.Model):
    __tablename__ = 'agent_devices'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    fingerprint = db.Column(db.String(64), nullable=False)
    ip_address = db.Column(db.String(45))
    pending_ip = db.Column(db.String(45))  # new IP awaiting admin approval
    label = db.Column(db.String(100))
    status = db.Column(db.String(20), default='pending')
    first_seen = db.Column(db.DateTime, default=datetime.utcnow)
    last_seen = db.Column(db.DateTime, default=datetime.utcnow)
    approved_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    approved_at = db.Column(db.DateTime)
    approval_token = db.Column(db.String(64), nullable=True)

    user = db.relationship('User', foreign_keys=[user_id], backref='devices')

class SearchFeedback(db.Model):
    __tablename__ = 'search_feedback'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    query = db.Column(db.String(500))
    result_counts = db.Column(db.String(200))
    rating = db.Column(db.String(20))
    comment = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user = db.relationship('User', foreign_keys=[user_id])

class PendingLoginApproval(db.Model):
    __tablename__ = 'pending_login_approvals'
    
    id = db.Column(db.Integer, primary_key=True)
    admin_user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    device_fingerprint = db.Column(db.String(64), nullable=False)
    ip_address = db.Column(db.String(45))
    location = db.Column(db.String(200))
    user_agent = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    status = db.Column(db.String(20), default='pending', nullable=False)
    approval_token = db.Column(db.String(64), unique=True, nullable=False)
    approved_by_user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    approved_at = db.Column(db.DateTime)
    notes = db.Column(db.Text)
    
    admin_user = db.relationship('User', foreign_keys=[admin_user_id], backref='pending_login_approvals')
    approved_by_user = db.relationship('User', foreign_keys=[approved_by_user_id])
    
    def __repr__(self):
        return f'<PendingLoginApproval {self.id} {self.status}>'

class BruteForceAlert(db.Model):
    __tablename__ = 'brute_force_alerts'

    id = db.Column(db.Integer, primary_key=True)
    alert_token = db.Column(db.String(64), unique=True, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    ip_address = db.Column(db.String(45))
    attempt_count = db.Column(db.Integer, default=0)
    status = db.Column(db.String(20), default='pending')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    telegram_message_id = db.Column(db.Integer)

    user = db.relationship('User', foreign_keys=[user_id])

# ==================== HELPER FUNCTIONS ====================

def calc_download_cost(row_count):
    """Return credit cost based on number of rows: 1–10."""
    if row_count <= 10:   return 1
    if row_count <= 50:   return 2
    if row_count <= 150:  return 3
    if row_count <= 300:  return 4
    if row_count <= 500:  return 5
    if row_count <= 750:  return 6
    if row_count <= 1000: return 7
    if row_count <= 1500: return 8
    if row_count <= 2000: return 9
    return 10

SEARCH_RESULT_LIMIT = int(os.environ.get('SEARCH_RESULT_LIMIT', '500'))
SEARCH_DOWNLOAD_LIMIT = int(os.environ.get('SEARCH_DOWNLOAD_LIMIT', '5000'))
SEARCH_MIN_LENGTHS = {
    'keyword': 3,
    'name': 2,
    'phone': 6,
    'ic': 6,
    'address': 4,
    'data_source': 3,
    'director': 2,
}


def _base_customer_search_query():
    return CustomerData.query.options(load_only(
        CustomerData.id,
        CustomerData.name,
        CustomerData.contact_number,
        CustomerData.ic_number,
        CustomerData.address,
        CustomerData.email,
        CustomerData.additional_data,
        CustomerData.data_source
    ))


def _normalize_search_term(term):
    return ' '.join((term or '').strip().split())


def _normalized_phone_digits(value):
    return re.sub(r'[^0-9]', '', value or '')


def _normalized_phone_query_expression():
    expr = CustomerData.contact_number
    for token in [' ', '-', '+', '(', ')']:
        expr = func.replace(expr, token, '')
    return expr


def _search_limit_for_type(search_type, requested_limit=None):
    is_download = requested_limit == 'download'
    limits = {
        'ic':          (300,  5000),
        'phone':       (300,  5000),
        'name':        (500,  5000),
        'address':     (300,  5000),
        'data_source': (500,  5000),
        'keyword':     (200,  2000),
    }
    screen, download = limits.get(search_type, (SEARCH_RESULT_LIMIT, SEARCH_DOWNLOAD_LIMIT))
    return download if is_download else screen


def _search_minimum_for_type(search_type):
    return SEARCH_MIN_LENGTHS.get(search_type or '', 2)


def _search_message_for_minimum(search_type, minimum):
    label = {
        'keyword': 'keyword',
        'name': 'name',
        'phone': 'phone number',
        'ic': 'IC number',
        'address': 'address'
    }.get(search_type, 'search term')
    return f'Please enter at least {minimum} characters for {label} search.'

def search_by_ic(term):
    """Search IC number tolerantly — with or without dashes."""
    clean = term.replace('-', '').replace(' ', '')
    results = CustomerData.query.filter(CustomerData.ic_number.ilike(f'%{term}%')).all()
    if results:
        return results
    all_customers = CustomerData.query.filter(CustomerData.ic_number != '').all()
    return [c for c in all_customers if c.ic_number.replace('-', '').replace(' ', '') == clean]

def _prefix_range(prefix):
    """Convert prefix to >= / < range so SQLite uses the B-tree index."""
    if not prefix:
        return None, None
    # Increment last character to get upper bound
    upper = prefix[:-1] + chr(ord(prefix[-1]) + 1)
    return prefix, upper

def search_by_ic_safe(term):
    """Search IC using range query on indexed column — avoids full table scan."""
    search_term = _normalize_search_term(term)
    clean = _normalize_identity_ic(search_term)
    if len(clean) < _search_minimum_for_type('ic'):
        return []

    query = (_base_customer_search_query()
             .filter(CustomerData.ic_number.isnot(None), CustomerData.ic_number != ''))

    if len(clean) >= 10:
        query = query.filter(CustomerData.ic_number == clean)
    else:
        lo, hi = _prefix_range(clean)
        query = query.filter(CustomerData.ic_number >= lo, CustomerData.ic_number < hi)

    return (query
            .order_by(CustomerData.id.desc())
            .limit(_search_limit_for_type('ic'))
            .all())


def perform_customer_search(search_type, term, requested_limit=None):
    search_term = _normalize_search_term(term)
    minimum = _search_minimum_for_type(search_type)
    compact_term = re.sub(r'\s+', '', search_term)
    limit = _search_limit_for_type(search_type, requested_limit)

    if not search_term:
        return [], {'ok': False, 'message': 'Please enter a search term.', 'minimum': minimum, 'limit': limit}
    if len(compact_term) < minimum:
        return [], {'ok': False, 'message': _search_message_for_minimum(search_type, minimum), 'minimum': minimum, 'limit': limit}

    # Panic mode: never touch the real DB — serve convincing decoy data instead
    if dp.is_panic():
        return dp.panic_results(search_type, search_term)

    # Cache check — downloads always hit DB (credit billing needs accurate count)
    _lkey = str(requested_limit)
    if requested_limit != 'download':
        _hit = _scache_get(search_type, search_term, _lkey)
        if _hit is not None:
            from data_protection import _MRow
            _rows, _meta = _hit
            return [_MRow(**r) for r in _rows], _meta

    query = _base_customer_search_query()
    results = []

    if search_type == 'ic':
        results = search_by_ic_safe(search_term)
    elif search_type == 'phone':
        # Data is already normalized at import — query indexed column directly
        digits = _normalized_phone_digits(search_term)
        # Ensure leading 0 for Malaysian numbers
        if digits and not digits.startswith('0') and not search_term.startswith('+'):
            digits = '0' + digits
        if len(digits) >= 10:
            # Full number — exact match (instant via index)
            results = (query
                       .filter(CustomerData.contact_number.isnot(None))
                       .filter(CustomerData.contact_number == digits)
                       .order_by(CustomerData.id.desc())
                       .limit(limit)
                       .all())
        elif len(digits) >= 6:
            # Partial — range query forces B-tree index usage
            lo, hi = _prefix_range(digits)
            results = (query
                       .filter(CustomerData.contact_number.isnot(None))
                       .filter(CustomerData.contact_number >= lo,
                               CustomerData.contact_number < hi)
                       .order_by(CustomerData.id.desc())
                       .limit(limit)
                       .all())
    elif search_type == 'name':
        # Use FTS5 for fast prefix search — falls back to LIKE if FTS table missing
        try:
            from sqlalchemy import text
            fts_pattern = f'"{search_term}"*'
            id_rows = db.session.execute(
                text('SELECT rowid FROM fts_names WHERE name MATCH :q ORDER BY rowid DESC LIMIT :lim'),
                {'q': fts_pattern, 'lim': limit}
            ).fetchall()
            ids = [r[0] for r in id_rows]
            if ids:
                results = (query
                           .filter(CustomerData.id.in_(ids))
                           .order_by(CustomerData.id.desc())
                           .all())
            else:
                results = []
        except Exception:
            results = (query
                       .filter(CustomerData.name.isnot(None), CustomerData.name != '')
                       .filter(CustomerData.name.ilike(f'{search_term}%'))
                       .order_by(CustomerData.id.desc())
                       .limit(limit)
                       .all())
    elif search_type == 'address':
        try:
            from sqlalchemy import text
            id_rows = db.session.execute(
                text('SELECT rowid FROM fts_address WHERE address MATCH :q ORDER BY rowid DESC LIMIT :lim'),
                {'q': f'"{search_term}"*', 'lim': limit}
            ).fetchall()
            ids = [r[0] for r in id_rows]
            results = (query.filter(CustomerData.id.in_(ids)).order_by(CustomerData.id.desc()).all()) if ids else []
        except Exception:
            results = (query
                       .filter(CustomerData.address.isnot(None), CustomerData.address != '')
                       .filter(CustomerData.address.ilike(f'%{search_term}%'))
                       .order_by(CustomerData.id.desc())
                       .limit(limit)
                       .all())
    elif search_type == 'data_source':
        matched_sources = _match_data_sources(search_term)
        if not matched_sources:
            return [], {'ok': False, 'message': f'No property/source matching "{search_term}" found.', 'minimum': minimum, 'limit': limit}
        results = (query
                   .filter(CustomerData.data_source.in_(matched_sources))
                   .order_by(CustomerData.id.desc())
                   .limit(limit)
                   .all())
    elif search_type == 'director':
        results = (query
                   .filter(CustomerData.data_source == 'Director')
                   .filter(or_(
                       CustomerData.name.ilike(f'%{search_term}%'),
                       CustomerData.address.ilike(f'%{search_term}%'),
                       CustomerData.contact_number.ilike(f'%{search_term}%'),
                   ))
                   .order_by(CustomerData.id.desc())
                   .limit(limit)
                   .all())
    elif search_type == 'keyword':
        if search_term.isdigit():
            if len(search_term) >= 12:
                return perform_customer_search('ic', search_term, requested_limit=requested_limit)
            if len(search_term) >= 9:
                return perform_customer_search('phone', search_term, requested_limit=requested_limit)
            return perform_customer_search('ic', search_term, requested_limit=requested_limit)

        # Try data_source match first (fast: exact IN lookup via index)
        matched_sources = _match_data_sources(search_term)
        if matched_sources:
            results = (query
                       .filter(CustomerData.data_source.in_(matched_sources))
                       .order_by(CustomerData.id.desc())
                       .limit(limit)
                       .all())
        if not results:
            # Use FTS5 for fast name search — avoids full table scan
            try:
                from sqlalchemy import text
                fts_pattern = f'"{search_term}"*'
                id_rows = db.session.execute(
                    text('SELECT rowid FROM fts_names WHERE name MATCH :q ORDER BY rowid DESC LIMIT :lim'),
                    {'q': fts_pattern, 'lim': limit}
                ).fetchall()
                ids = [r[0] for r in id_rows]
                if ids:
                    results = (query.filter(CustomerData.id.in_(ids))
                               .order_by(CustomerData.id.desc()).all())
            except Exception:
                pass
        if not results:
            # Last resort: prefix name match only (no address LIKE — too slow on large DB)
            first_token = re.split(r'\s+', search_term)[0]
            if len(first_token) >= 3:
                results = (query
                           .filter(CustomerData.name.ilike(f'{first_token}%'))
                           .order_by(CustomerData.id.desc())
                           .limit(limit)
                           .all())
    else:
        return [], {'ok': False, 'message': 'Please choose a valid search type.', 'minimum': minimum, 'limit': limit}

    truncated = len(results) >= limit
    message = f'Found {len(results):,} result(s)'
    if truncated:
        # Fast total count so user knows how many are available
        try:
            total = _base_customer_search_query().filter(
                *[f for f in query.whereclause.clauses]
                if hasattr(query, 'whereclause') and query.whereclause is not None else []
            ).count() if False else None  # placeholder — use simple note instead
        except Exception:
            total = None
        message += f' — showing first {limit:,}. Use Download Excel to get all.'

    meta = {'ok': True, 'message': message, 'minimum': minimum, 'limit': limit, 'truncated': truncated}
    if requested_limit != 'download':
        _scache_set(search_type, search_term, _lkey, results, meta)
    return results, meta


def build_customer_search_queries(customer):
    identity_parts = [clean_data_value(customer.name), clean_data_value(customer.contact_number), clean_data_value(customer.email)]
    identity_query = ' '.join(part for part in identity_parts if part)
    broad_query = ' '.join(part for part in [
        clean_data_value(customer.name),
        clean_data_value(customer.contact_number),
        clean_data_value(customer.ic_number),
        clean_data_value(customer.address)
    ] if part)

    links = [
        {
            'label': 'Google Search',
            'platform': 'Google',
            'url': f'https://www.google.com/search?q={quote_plus(broad_query or identity_query or str(customer.id))}'
        },
        {
            'label': 'Google News',
            'platform': 'Google News',
            'url': f'https://www.google.com/search?tbm=nws&q={quote_plus(identity_query or broad_query or str(customer.id))}'
        },
        {
            'label': 'LinkedIn Search',
            'platform': 'LinkedIn',
            'url': f'https://www.google.com/search?q={quote_plus("site:linkedin.com/in " + (identity_query or broad_query or str(customer.id)))}'
        },
        {
            'label': 'Facebook Search',
            'platform': 'Facebook',
            'url': f'https://www.google.com/search?q={quote_plus("site:facebook.com " + (identity_query or broad_query or str(customer.id)))}'
        },
        {
            'label': 'Instagram Search',
            'platform': 'Instagram',
            'url': f'https://www.google.com/search?q={quote_plus("site:instagram.com " + (identity_query or broad_query or str(customer.id)))}'
        }
    ]

    if customer.contact_number:
        links.append({
            'label': 'Search Phone Number',
            'platform': 'Phone Lookup',
            'url': f'https://www.google.com/search?q={quote_plus(customer.contact_number)}'
        })
    if customer.email:
        links.append({
            'label': 'Search Email',
            'platform': 'Email Lookup',
            'url': f'https://www.google.com/search?q={quote_plus(customer.email)}'
        })

    return links

def infer_platform_from_url(url):
    url_lower = (url or '').lower()
    if 'linkedin.com' in url_lower:
        return 'LinkedIn'
    if 'facebook.com' in url_lower:
        return 'Facebook'
    if 'instagram.com' in url_lower:
        return 'Instagram'
    if 'tiktok.com' in url_lower:
        return 'TikTok'
    if 'twitter.com' in url_lower or 'x.com' in url_lower:
        return 'X'
    return 'Public Web'

def is_google_internal_url(url):
    lowered = (url or '').lower()
    return (
        'google.com/search?' in lowered or
        'google.com/sorry/' in lowered or
        'google.com/sorry/index' in lowered or
        '/sorry/' in lowered
    )

def normalize_result_url(url, base_url=''):
    raw_url = clean_data_value(url)
    if not raw_url:
        return ''

    resolved = urljoin(base_url, raw_url) if base_url else raw_url
    parsed = urlparse(resolved)
    query = parse_qs(parsed.query)

    redirect_keys = ['uddg', 'u', 'url', 'q']
    for key in redirect_keys:
        if key in query and query[key]:
            candidate = clean_data_value(unquote(query[key][0]))
            if candidate.startswith('http://') or candidate.startswith('https://'):
                return candidate

    return resolved

def is_blocked_search_page(title, snippet, url):
    combined = ' '.join([title or '', snippet or '', url or '']).casefold()
    blocked_phrases = [
        'unusual traffic',
        'our systems have detected unusual traffic',
        'why did this happen',
        'your computer network',
        'google sorry',
        '/sorry/index'
    ]
    return any(phrase in combined for phrase in blocked_phrases) or is_google_internal_url(url)

def strip_html_tags(value):
    return re.sub(r'<[^>]+>', '', value or '')

def parse_duckduckgo_results(html_text, base_url='https://html.duckduckgo.com/'):
    results = []
    soup = BeautifulSoup(html_text or '', 'html.parser')
    for anchor in soup.select('a.result__a, a[href][class*="result__a"]'):
        href = normalize_result_url(html.unescape(anchor.get('href', '')).strip(), base_url)
        title = clean_data_value(anchor.get_text(' ', strip=True))
        snippet = ''

        container = anchor.find_parent(['div', 'article'])
        if container:
            snippet_node = container.select_one('.result__snippet')
            if snippet_node:
                snippet = clean_data_value(snippet_node.get_text(' ', strip=True))
            elif container.get_text(' ', strip=True):
                snippet = clean_data_value(container.get_text(' ', strip=True).replace(title, '', 1))[:700]

        if href and title:
            results.append({
                'url': href,
                'title': title,
                'snippet': snippet
            })
    return results

def parse_bing_results_html(html_text, base_url='https://www.bing.com/'):
    results = []
    soup = BeautifulSoup(html_text or '', 'html.parser')

    for item in soup.select('li.b_algo'):
        anchor = item.select_one('h2 a[href], a[href]')
        title_node = item.select_one('h2')
        snippet_node = item.select_one('.b_caption p, .b_snippet, p')

        href = normalize_result_url(clean_data_value(anchor.get('href') if anchor else ''), base_url)
        title = clean_data_value(title_node.get_text(' ', strip=True) if title_node else (anchor.get_text(' ', strip=True) if anchor else ''))
        snippet = clean_data_value(snippet_node.get_text(' ', strip=True) if snippet_node else '')

        if href and title:
            results.append({
                'url': href,
                'title': title,
                'snippet': snippet
            })
    return results

def query_signals_for_matching(query):
    cleaned = clean_data_value(query)
    digits = re.sub(r'[^0-9]+', '', cleaned)
    email = extract_email_from_text(cleaned)
    tokens = [token.casefold() for token in re.split(r'[\s,]+', cleaned) if len(token) >= 4]
    return {
        'cleaned': cleaned,
        'digits': digits,
        'email': email.casefold() if email else '',
        'tokens': tokens[:6]
    }

def text_matches_query_signals(text_value, signals):
    haystack = clean_data_value(text_value).casefold()
    haystack_digits = re.sub(r'[^0-9]+', '', haystack)

    if signals['email'] and signals['email'] in haystack:
        return True
    if signals['digits'] and signals['digits'] in haystack_digits:
        return True
    if signals['tokens']:
        token_hits = sum(1 for token in signals['tokens'] if token in haystack)
        if token_hits >= min(2, len(signals['tokens'])):
            return True
    return False

def parse_generic_search_results_html(html_text, query, base_url=''):
    results = []
    soup = BeautifulSoup(html_text or '', 'html.parser')
    signals = query_signals_for_matching(query)
    seen = set()

    for anchor in soup.select('a[href]'):
        href = normalize_result_url(clean_data_value(anchor.get('href')), base_url)
        if not href or href in seen:
            continue
        if not href.startswith('http'):
            continue
        if not is_allowed_destination_url(href):
            continue

        title = clean_data_value(anchor.get_text(' ', strip=True))
        container = anchor.find_parent(['div', 'li', 'article', 'section'])
        snippet = ''
        if container:
            snippet = clean_data_value(container.get_text(' ', strip=True).replace(title, '', 1))[:800]

        combined_text = ' '.join(part for part in [title, snippet] if part)
        if not combined_text or not text_matches_query_signals(combined_text, signals):
            continue

        results.append({
            'url': href,
            'title': title or snippet[:120] or 'Search Result',
            'snippet': snippet or title
        })
        seen.add(href)
        if len(results) >= 10:
            break

    return results

def parse_google_results_html(html_text):
    results = []
    soup = BeautifulSoup(html_text or '', 'html.parser')

    for block in soup.select('div.g'):
        anchor = block.select_one('a[href]')
        title_node = block.select_one('h3')
        snippet_node = block.select_one('.VwiC3b, .yXK7lf, .MUxGbd')

        href = anchor.get('href', '').strip() if anchor else ''
        title = clean_data_value(title_node.get_text(' ', strip=True) if title_node else '')
        snippet = clean_data_value(snippet_node.get_text(' ', strip=True) if snippet_node else '')

        if href and title:
            results.append({
                'url': href,
                'title': title,
                'snippet': snippet
            })

    return results

def get_enrichment_http_session():
    session_http = requests.Session()
    session_http.trust_env = False
    return session_http

def fetch_search_results_requests(query, page_start=0):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36'
    }

    session_http = get_enrichment_http_session()
    try:
        try:
            response = session_http.get(
                'https://html.duckduckgo.com/html/',
                params={'q': query, 's': page_start},
                headers=headers,
                timeout=12
            )
            response.raise_for_status()
            results = [
                item for item in parse_duckduckgo_results(response.text, response.url)
                if not is_blocked_search_page(item.get('title'), item.get('snippet'), item.get('url'))
            ]
            if not results:
                results = [
                    item for item in parse_generic_search_results_html(response.text, query, response.url)
                    if not is_blocked_search_page(item.get('title'), item.get('snippet'), item.get('url'))
                ]
        except Exception as exc:
            return {
                'provider': 'duckduckgo_requests',
                'status_code': None,
                'results': [],
                'error': str(exc)
            }
        return {
            'provider': 'duckduckgo_requests',
            'status_code': response.status_code,
            'results': results,
            'error': '',
            'debug_samples': results[:5]
        }
    finally:
        session_http.close()

def fetch_search_results_bing(query, page_start=0):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36'
    }

    session_http = get_enrichment_http_session()
    try:
        try:
            response = session_http.get(
                'https://www.bing.com/search',
                params={'q': query, 'first': page_start + 1},
                headers=headers,
                timeout=12
            )
            response.raise_for_status()
            results = [
                item for item in parse_bing_results_html(response.text, response.url)
                if not is_blocked_search_page(item.get('title'), item.get('snippet'), item.get('url'))
            ]
            if not results:
                results = [
                    item for item in parse_generic_search_results_html(response.text, query, response.url)
                    if not is_blocked_search_page(item.get('title'), item.get('snippet'), item.get('url'))
                ]
        except Exception as exc:
            return {
                'provider': 'bing_requests',
                'status_code': None,
                'results': [],
                'error': str(exc)
            }
        return {
            'provider': 'bing_requests',
            'status_code': response.status_code,
            'results': results,
            'error': '',
            'debug_samples': results[:5]
        }
    finally:
        session_http.close()

def fetch_search_results_playwright(query, page_start=0):
    if sync_playwright is None:
        return {
            'provider': 'google_playwright',
            'status_code': None,
            'results': [],
            'error': 'Playwright is not available in this environment.'
        }

    offset = page_start if page_start > 0 else 0
    search_url = f'https://www.google.com/search?q={quote_plus(query)}&start={offset}'

    browser = None
    playwright_manager = None
    try:
        playwright_manager = sync_playwright().start()
        browser = playwright_manager.chromium.launch(
            headless=app.config['ENRICHMENT_PLAYWRIGHT_HEADLESS'],
            channel=app.config['ENRICHMENT_PLAYWRIGHT_CHANNEL'] or None
        )
        page = browser.new_page()
        page.goto(search_url, wait_until='domcontentloaded', timeout=25000)
        page.wait_for_timeout(1500)
        try:
            page.locator('h3').first.wait_for(timeout=8000)
        except Exception:
            pass

        query_digits = re.sub(r'[^0-9]+', '', query or '')
        results = page.evaluate(
            """
            (queryDigits) => {
              const seen = new Set();
              const items = [];
              const headingNodes = Array.from(document.querySelectorAll('h3'));
              const containerSelectors = [
                'div.g',
                'div.MjjYud',
                'div.Gx5Zad',
                'div.N54PNb',
                'div[data-snc]'
              ];

              for (const h3 of headingNodes) {
                const anchor = h3.closest('a[href]');
                if (!anchor) continue;

                const url = (anchor.href || '').trim();
                const title = (h3.innerText || '').trim();
                if (!url || !title) continue;
                if (url.startsWith('javascript:')) continue;
                if (seen.has(url)) continue;

                let container = anchor.closest('div.g') || anchor.closest('[data-snc]') || anchor.parentElement;
                let snippet = '';
                if (container) {
                  const snippetNode = container.querySelector('.VwiC3b, .yXK7lf, .MUxGbd, .GI74Re, .hgKElc');
                  if (snippetNode) snippet = (snippetNode.innerText || '').trim();
                }

                if (!snippet) {
                  const parentText = ((container && container.innerText) || '').trim();
                  snippet = parentText.replace(title, '').slice(0, 600).trim();
                }

                items.push({ url, title, snippet });
                seen.add(url);
                if (items.length >= 10) break;
              }

              if (items.length === 0) {
                const anchors = Array.from(document.querySelectorAll('a[href]'));
                for (const anchor of anchors) {
                  const url = (anchor.href || '').trim();
                  if (!url || seen.has(url)) continue;
                  if (!/^https?:/i.test(url)) continue;
                  if (url.includes('/search?')) continue;

                  const title = (anchor.innerText || '').trim();
                  const digits = (title + ' ' + ((anchor.parentElement && anchor.parentElement.innerText) || '')).replace(/[^0-9]+/g, '');
                  if (!title && !digits.includes(queryDigits)) continue;

                  let container = null;
                  for (const selector of containerSelectors) {
                    container = anchor.closest(selector);
                    if (container) break;
                  }
                  if (!container) container = anchor.parentElement;

                  const snippet = (((container && container.innerText) || '').trim() || title).slice(0, 700);
                  if (!title && !snippet) continue;
                  if (queryDigits && !snippet.replace(/[^0-9]+/g, '').includes(queryDigits)) continue;

                  items.push({
                    url,
                    title: title || snippet.slice(0, 120),
                    snippet
                  });
                  seen.add(url);
                  if (items.length >= 10) break;
                }
              }

              if (items.length === 0) {
                const bodyText = (document.body && document.body.innerText) || '';
                const bodyDigits = bodyText.replace(/[^0-9]+/g, '');
                if (queryDigits && bodyDigits.includes(queryDigits)) {
                  items.push({
                    url: window.location.href,
                    title: 'Google Search Results',
                    snippet: bodyText.slice(0, 1000)
                  });
                }
              }

              return items;
            }
            """,
            query_digits
        )

        if not results:
            html_text = page.content()
            results = parse_google_results_html(html_text)

        body_text = ''
        try:
            body_text = page.locator('body').inner_text(timeout=3000)
        except Exception:
            body_text = ''

        if any(phrase in body_text.casefold() for phrase in [
            'our systems have detected unusual traffic',
            'why did this happen',
            'your computer network'
        ]):
            return {
                'provider': 'google_playwright',
                'status_code': 429,
                'results': [],
                'error': 'Google blocked the crawler with an unusual traffic challenge.',
                'debug_samples': []
            }

        results = [
            item for item in results
            if not is_blocked_search_page(item.get('title'), item.get('snippet'), item.get('url'))
        ]

        return {
            'provider': 'google_playwright',
            'status_code': 200,
            'results': results,
            'error': '',
            'debug_samples': results[:5]
        }
    except Exception as exc:
        return {
            'provider': 'google_playwright',
            'status_code': None,
            'results': [],
            'error': str(exc),
            'debug_samples': []
        }
    finally:
        if browser is not None:
            browser.close()
        if playwright_manager is not None:
            playwright_manager.stop()

def get_enrichment_strategy_order():
    mode = app.config.get('ENRICHMENT_CRAWL_MODE', 'auto')
    if mode == 'requests':
        return ['requests', 'bing']
    if mode == 'playwright':
        return ['playwright', 'requests', 'bing']
    return ['requests', 'bing', 'playwright']

def extract_email_from_text(text_value):
    match = re.search(r'([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})', text_value or '')
    return match.group(1) if match else ''

def extract_phone_from_text(text_value):
    match = re.search(r'(\+?\d[\d\-\s]{7,}\d)', text_value or '')
    return clean_data_value(match.group(1)) if match else ''

def extract_company_and_title(title, snippet):
    title_text = clean_data_value(title)
    snippet_text = clean_data_value(snippet)
    combined = ' | '.join(part for part in [title_text, snippet_text] if part)

    company = ''
    role = ''

    title_parts = [part.strip() for part in re.split(r'[-|]', title_text) if clean_data_value(part)]
    if len(title_parts) >= 2:
        role = title_parts[0][:200]
        company = title_parts[1][:200]
    else:
        role_match = re.search(r'\b(manager|director|founder|owner|agent|consultant|executive|specialist|advisor|ceo|cto|coo|marketing|sales)\b', combined, re.IGNORECASE)
        if role_match:
            role = role_match.group(1).title()

        company_match = re.search(r'\b(?:at|with|from)\s+([A-Z][A-Za-z0-9&.,\-\s]{2,80})', snippet_text)
        if company_match:
            company = clean_data_value(company_match.group(1))[:200]

    return company, role

def extract_location_from_text(text_value):
    if not text_value:
        return ''

    separators = [' in ', ' at ', ' based in ', ' located in ', ' from ']
    lowered = text_value.casefold()
    for separator in separators:
        idx = lowered.find(separator)
        if idx != -1:
            location = clean_data_value(text_value[idx + len(separator):])
            location = re.split(r'[|.;]', location)[0]
            return location[:200]
    return ''

def extract_candidate_name_from_text(text_value):
    cleaned = clean_data_value(text_value)
    if not cleaned:
        return ''

    match = re.search(r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3})\b', cleaned)
    return clean_data_value(match.group(1))[:200] if match else ''

def is_allowed_destination_url(url):
    if not url:
        return False
    parsed = urlparse(url)
    if parsed.scheme not in ['http', 'https']:
        return False
    host = (parsed.netloc or '').casefold()
    blocked_hosts = [
        'google.com', 'www.google.com', 'bing.com', 'www.bing.com',
        'html.duckduckgo.com', 'duckduckgo.com'
    ]
    return not any(host == blocked or host.endswith('.' + blocked) for blocked in blocked_hosts)

def fetch_destination_page_signals(url, customer):
    if not is_allowed_destination_url(url):
        return None

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36'
    }
    session_http = get_enrichment_http_session()
    try:
        response = session_http.get(url, headers=headers, timeout=12, allow_redirects=True)
        response.raise_for_status()
        html_text = response.text
    except Exception:
        session_http.close()
        return None
    finally:
        try:
            session_http.close()
        except Exception:
            pass

    soup = BeautifulSoup(html_text or '', 'html.parser')
    page_title = clean_data_value(soup.title.get_text(' ', strip=True) if soup.title else '')
    body_text = clean_data_value(soup.get_text(' ', strip=True))[:5000]
    combined_text = ' '.join(part for part in [page_title, body_text] if part)

    found_phone = extract_phone_from_text(combined_text)
    found_email = extract_email_from_text(combined_text)
    found_location = extract_location_from_text(combined_text)
    found_company, found_role = extract_company_and_title(page_title, body_text[:800])
    found_name = ''

    if customer.name and customer.name.casefold() in combined_text.casefold():
        found_name = customer.name
    elif customer.contact_number and re.sub(r'[^0-9]+', '', customer.contact_number) in re.sub(r'[^0-9]+', '', combined_text):
        found_name = extract_candidate_name_from_text(page_title) or extract_candidate_name_from_text(body_text[:300])

    if not any([found_phone, found_email, found_company, found_role, found_location, found_name]):
        return None

    return {
        'url': url,
        'title': page_title[:200] or 'Destination Page',
        'snippet': body_text[:1000],
        'matched_name': found_name,
        'matched_phone': found_phone,
        'matched_email': found_email,
        'matched_location': found_location,
        'matched_company': found_company,
        'matched_title': found_role or page_title[:200]
    }

def calculate_enrichment_confidence(customer, title, snippet, url):
    haystack = ' '.join([title or '', snippet or '', url or '']).casefold()
    score = 20

    if clean_data_value(customer.name) and clean_data_value(customer.name).casefold() in haystack:
        score += 35
    if clean_data_value(customer.contact_number) and re.sub(r'[^0-9]+', '', customer.contact_number) in re.sub(r'[^0-9]+', '', haystack):
        score += 25
    if clean_data_value(customer.email) and clean_data_value(customer.email).casefold() in haystack:
        score += 25
    if clean_data_value(customer.address):
        address_tokens = [token for token in re.split(r'[\s,]+', customer.address.casefold()) if len(token) > 4][:4]
        if any(token in haystack for token in address_tokens):
            score += 10

    return max(0, min(score, 100))

def auto_crawl_customer_enrichment(customer, admin_id, max_results=6):
    """
    Crawl public search results and save candidate enrichment rows.
    Uses DuckDuckGo's HTML results as a lightweight public-web discovery source.
    """
    queries = []
    if customer.name:
        queries.append(' '.join(part for part in [customer.name, customer.contact_number, customer.email] if part))
        queries.append(f'{customer.name} {customer.ic_number}'.strip())
    if customer.contact_number:
        queries.append(customer.contact_number)
    if customer.email:
        queries.append(customer.email)

    queries = [clean_data_value(query) for query in queries if clean_data_value(query)]
    seen_urls = {
        enrichment.source_url
        for enrichment in CustomerEnrichment.query.filter_by(customer_id=customer.id).all()
        if enrichment.source_url
    }

    created = 0
    checked = 0
    errors = []
    strategy_order = get_enrichment_strategy_order()
    providers_used = []
    debug_samples = []

    for query in queries[:3]:
        try:
            for page_start in [0]:
                provider_result = None
                for strategy in strategy_order:
                    if strategy == 'requests':
                        provider_result = fetch_search_results_requests(query, page_start)
                    elif strategy == 'bing':
                        provider_result = fetch_search_results_bing(query, page_start)
                    elif strategy == 'playwright':
                        provider_result = fetch_search_results_playwright(query, page_start)
                    else:
                        continue

                    providers_used.append(provider_result['provider'])
                    for sample in provider_result.get('debug_samples', [])[:3]:
                        if len(debug_samples) < 12:
                            debug_samples.append({
                                'provider': provider_result['provider'],
                                'query': query,
                                'url': sample.get('url', ''),
                                'title': sample.get('title', ''),
                                'snippet': sample.get('snippet', '')[:300]
                            })
                    if provider_result['results']:
                        break

                if provider_result is None:
                    continue

                if provider_result['error']:
                    errors.append(f"{query} [{provider_result['provider']}]: {provider_result['error']}")
                    if provider_result.get('status_code') == 429:
                        break

                parsed_results = provider_result['results']
                checked += len(parsed_results[:max_results])

                for item in parsed_results[:max_results]:
                    source_url = item['url']
                    if source_url in seen_urls:
                        continue
                    if not is_allowed_destination_url(source_url):
                        continue

                    destination_signals = fetch_destination_page_signals(source_url, customer)
                    effective_item = destination_signals or item

                    matched_company, matched_role = extract_company_and_title(effective_item['title'], effective_item['snippet'])
                    matched_email = destination_signals.get('matched_email') if destination_signals else extract_email_from_text(effective_item['snippet'])
                    matched_phone = destination_signals.get('matched_phone') if destination_signals else extract_phone_from_text(effective_item['snippet'])
                    matched_location = destination_signals.get('matched_location') if destination_signals else extract_location_from_text(effective_item['snippet'])
                    matched_name = destination_signals.get('matched_name') if destination_signals else (
                        customer.name if customer.name and customer.name.casefold() in (effective_item['title'] + ' ' + effective_item['snippet']).casefold() else ''
                    )
                    if destination_signals:
                        if destination_signals.get('matched_company'):
                            matched_company = destination_signals.get('matched_company')
                        if destination_signals.get('matched_title'):
                            matched_role = destination_signals.get('matched_title')

                    enrichment = CustomerEnrichment(
                        customer_id=customer.id,
                        platform=infer_platform_from_url(source_url),
                        source_url=source_url,
                        source_type='public_web_auto',
                        matched_name=matched_name,
                        matched_location=matched_location,
                        matched_company=matched_company,
                        matched_title=matched_role or effective_item['title'][:200],
                        matched_phone=matched_phone,
                        matched_email=matched_email,
                        confidence_score=calculate_enrichment_confidence(customer, effective_item['title'], effective_item['snippet'], source_url),
                        review_status='pending',
                        notes=effective_item['snippet'][:1000],
                        raw_data=json.dumps({
                            'query': query,
                            'title': effective_item['title'],
                            'snippet': effective_item['snippet'],
                            'url': source_url,
                            'page_start': page_start,
                            'provider': provider_result['provider'],
                            'destination_fetch': bool(destination_signals)
                        }),
                        created_by=admin_id
                    )
                    db.session.add(enrichment)
                    seen_urls.add(source_url)
                    created += 1
        except Exception as exc:
            errors.append(f'{query}: {str(exc)}')

    if created:
        db.session.commit()
    elif errors:
        db.session.rollback()

    return {
        'created': created,
        'checked': checked,
        'errors': errors,
        'providers_used': providers_used,
        'mode': app.config.get('ENRICHMENT_CRAWL_MODE', 'auto'),
        'debug_samples': debug_samples
    }

def get_latest_enrichment_job(customer_id):
    return CustomerEnrichmentJob.query.filter_by(customer_id=customer_id).order_by(CustomerEnrichmentJob.created_at.desc()).first()

def enqueue_enrichment_job(customer_id, requested_by):
    existing = CustomerEnrichmentJob.query.filter(
        CustomerEnrichmentJob.customer_id == customer_id,
        CustomerEnrichmentJob.status.in_(['queued', 'running'])
    ).order_by(CustomerEnrichmentJob.created_at.desc()).first()
    if existing:
        return existing, False

    job = CustomerEnrichmentJob(
        customer_id=customer_id,
        requested_by=requested_by,
        mode=app.config.get('ENRICHMENT_CRAWL_MODE', 'auto'),
        status='queued',
        message='Queued for background crawl.'
    )
    db.session.add(job)
    db.session.commit()
    return job, True

def run_enrichment_connectivity_test():
    """
    Quick health check for crawler networking and parser readiness.
    """
    result = {
        'mode': app.config.get('ENRICHMENT_CRAWL_MODE', 'auto'),
        'checks': []
    }

    for strategy in get_enrichment_strategy_order():
        if strategy == 'requests':
            provider_result = fetch_search_results_requests('test', 0)
            result['checks'].append({
                'provider': provider_result['provider'],
                'internet_ok': provider_result['status_code'] is not None,
                'search_source_ok': provider_result['status_code'] in [200, 202],
                'parser_ok': len(provider_result['results']) > 0,
                'search_status_code': provider_result['status_code'],
                'parsed_count': len(provider_result['results']),
                'error': provider_result['error'],
                'debug_samples': provider_result.get('debug_samples', [])
            })
        elif strategy == 'bing':
            provider_result = fetch_search_results_bing('test', 0)
            result['checks'].append({
                'provider': provider_result['provider'],
                'internet_ok': provider_result['status_code'] is not None,
                'search_source_ok': provider_result['status_code'] in [200, 202],
                'parser_ok': len(provider_result['results']) > 0,
                'search_status_code': provider_result['status_code'],
                'parsed_count': len(provider_result['results']),
                'error': provider_result['error'],
                'debug_samples': provider_result.get('debug_samples', [])
            })
        elif strategy == 'playwright':
            provider_result = fetch_search_results_playwright('test', 0)
            result['checks'].append({
                'provider': provider_result['provider'],
                'internet_ok': provider_result['status_code'] is not None or not provider_result['error'],
                'search_source_ok': provider_result['status_code'] in [200, 202],
                'parser_ok': len(provider_result['results']) > 0,
                'search_status_code': provider_result['status_code'],
                'parsed_count': len(provider_result['results']),
                'error': provider_result['error'],
                'debug_samples': provider_result.get('debug_samples', [])
            })

    return result

def get_client_ip():
    if request.headers.get('X-Forwarded-For'):
        ip = request.headers.get('X-Forwarded-For').split(',')[0]
    else:
        ip = request.remote_addr
    return ip

def log_security_event(event_type, status, ip_address, user=None, username=None, details=''):
    event = SecurityEventLog(
        user_id=user.id if user else None,
        username=username or (user.username if user else None),
        ip_address=ip_address,
        user_agent=request.headers.get('User-Agent', ''),
        event_type=event_type,
        status=status,
        details=details
    )
    db.session.add(event)
    db.session.commit()
    return event

def get_blocked_ips():
    raw = SystemSettings.get('blocked_ips', '[]')
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return parsed
    except Exception:
        pass
    return []

def is_ip_blocked(ip_address):
    return ip_address in set(get_blocked_ips())

def block_ip_address(ip_address):
    blocked_ips = get_blocked_ips()
    if ip_address not in blocked_ips:
        blocked_ips.append(ip_address)
        SystemSettings.set('blocked_ips', json.dumps(blocked_ips))
        return True
    return False

def check_admin_device(user_id, fingerprint, ip_address):
    """Returns 'trusted', 'new', 'pending', or 'blocked'."""
    device = AdminDevice.query.filter_by(user_id=user_id, fingerprint=fingerprint).first()
    if device:
        device.last_seen = datetime.utcnow()
        device.ip_address = ip_address
        db.session.commit()
        return device.status
    else:
        new_device = AdminDevice(
            user_id=user_id,
            fingerprint=fingerprint,
            ip_address=ip_address,
            label=f'Device from {ip_address}',
            status='pending'
        )
        db.session.add(new_device)
        db.session.commit()
        return 'new'

def is_admin_security_enabled():
    return SystemSettings.get('telegram_approval_enabled', '0') == '1'

def is_agent_ip_approval_enabled():
    return SystemSettings.get('agent_ip_approval_enabled', '0') == '1'

def handle_failed_login(user, ip_address):
    """Track failed attempts, alert boss at 3, lock account at 5."""
    user.failed_login_attempts = (user.failed_login_attempts or 0) + 1
    attempts = user.failed_login_attempts

    if attempts >= 5:
        user.locked_until = datetime.utcnow() + timedelta(minutes=30)
        db.session.commit()
        log_security_event('login_denied', 'locked', ip_address, user=user,
                           details=f'Account locked after {attempts} failed login attempts.')
        tg_send_brute_force_alert(user, ip_address, attempts, locked=True)
        return

    db.session.commit()

    log_security_event('login_denied', 'failed', ip_address, user=user,
                       details=f'Invalid password attempt #{attempts}.')

    if attempts == 3:
        log_security_event('login_denied', 'warning', ip_address, user=user,
                           details='Repeated failed login attempts reached alert threshold.')
        tg_send_brute_force_alert(user, ip_address, attempts, locked=False)

def tg_send_brute_force_alert(user, ip_address, attempts, locked=False):
    """Send Telegram alert to boss about suspicious login attempts."""
    token = SystemSettings.get('telegram_bot_token')
    chat_id = SystemSettings.get('telegram_boss_chat_id')
    if not token or not chat_id:
        return
    try:
        alert_token = secrets.token_hex(16)
        alert = BruteForceAlert(
            alert_token=alert_token,
            user_id=user.id,
            ip_address=ip_address,
            attempt_count=attempts,
            status='pending'
        )
        db.session.add(alert)
        db.session.commit()

        if locked:
            text = (
                f"🔴 *Account Auto-Locked*\n\n"
                f"*{user.full_name}* (`{user.username}`) account has been *automatically locked* "
                f"after *{attempts} failed login attempts*.\n\n"
                f"🌐 IP Address: `{ip_address}`\n"
                f"⏱ Locked for *30 minutes*.\n\n"
                f"Do you want to lock the entire system?"
            )
        else:
            text = (
                f"⚠️ *Suspicious Login Activity*\n\n"
                f"*{attempts} failed login attempts* detected for *{user.full_name}* (`{user.username}`).\n\n"
                f"🌐 IP Address: `{ip_address}`\n\n"
                f"Do you want to lock the system to prevent further attempts?"
            )

        keyboard = {"inline_keyboard": [[
            {"text": "🔒 Lock System", "callback_data": f"lockbf_{alert_token}"},
            {"text": "✅ Ignore", "callback_data": f"ignorebf_{alert_token}"}
        ]]}
        resp = http_req.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": chat_id, "text": text,
                  "parse_mode": "Markdown", "reply_markup": keyboard},
            timeout=10
        )
        if resp.ok:
            alert.telegram_message_id = resp.json()['result']['message_id']
            db.session.commit()
    except Exception as e:
        app.logger.error(f"Brute force alert error: {e}")

def create_pending_login_approval(user_id, fingerprint, ip_address, user_agent):
    approval = AdminLoginApproval(
        approval_token=secrets.token_hex(14),
        user_id=user_id,
        fingerprint=fingerprint,
        ip_address=ip_address,
        status='pending',
        expires_at=datetime.utcnow() + timedelta(minutes=30)
    )
    db.session.add(approval)
    db.session.commit()
    return approval

def send_telegram_notification(approval):
    return tg_send_approval_request(approval)

def create_customer_delete_approval(customer, admin_user, reason='Duplicate customer cleanup'):
    duplicate_name_count = 0
    if customer.name:
        duplicate_name_count = CustomerData.query.filter(
            CustomerData.id != customer.id,
            CustomerData.name == customer.name
        ).count()

    approval = CustomerDeleteApproval(
        approval_token=secrets.token_hex(16),
        customer_id=customer.id,
        requested_by=admin_user.id,
        status='pending',
        reason=reason,
        snapshot_name=customer.name,
        snapshot_contact_number=customer.contact_number,
        snapshot_ic_number=customer.ic_number,
        snapshot_address=customer.address,
        duplicate_name_count=duplicate_name_count,
        expires_at=datetime.utcnow() + timedelta(hours=24)
    )
    db.session.add(approval)
    db.session.commit()
    return approval

def tg_send_approval_request(approval):
    """Send Telegram message to boss with Approve/Decline buttons."""
    token = SystemSettings.get('telegram_bot_token')
    chat_id = SystemSettings.get('telegram_boss_chat_id')
    print(f"[TELEGRAM] token={'SET' if token else 'MISSING'} chat_id={'SET' if chat_id else 'MISSING'} enabled={SystemSettings.get('telegram_approval_enabled','0')}")
    if not token or not chat_id:
        print("[TELEGRAM] Skipping — bot token or chat ID not configured")
        return False
    try:
        text = (
            f"🚨 *Admin Login Alert*\n\n"
            f"*{approval.user.full_name}* is attempting to login from an *unrecognised device*.\n\n"
            f"📍 IP Address: `{approval.ip_address}`\n"
            f"🕐 Request expires in *30 minutes*.\n\n"
            f"Please approve or decline this login request."
        )
        keyboard = {"inline_keyboard": [[
            {"text": "✅ Approve", "callback_data": f"approve_{approval.approval_token}"},
            {"text": "❌ Decline", "callback_data": f"decline_{approval.approval_token}"}
        ]]}
        resp = http_req.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": chat_id, "text": text,
                  "parse_mode": "Markdown", "reply_markup": keyboard},
            timeout=10
        )
        print(f"[TELEGRAM] API response: {resp.status_code} {resp.text[:200]}")
        if resp.ok:
            approval.telegram_message_id = resp.json()['result']['message_id']
            db.session.commit()
        return resp.ok
    except Exception as e:
        print(f"[TELEGRAM] Exception: {e}")
        app.logger.error(f"Telegram send error: {e}")
        return False

def send_customer_delete_approval_request(approval):
    token = SystemSettings.get('telegram_bot_token')
    chat_id = SystemSettings.get('telegram_boss_chat_id')
    if not token or not chat_id:
        print("[TELEGRAM] Skipping customer delete approval - bot token or chat ID not configured")
        return False

    requester_name = approval.requester.full_name if approval.requester and approval.requester.full_name else (
        approval.requester.username if approval.requester else f'Admin #{approval.requested_by}'
    )
    customer_name = approval.snapshot_name or '-'
    customer_phone = approval.snapshot_contact_number or '-'
    customer_ic = approval.snapshot_ic_number or '-'
    customer_address = approval.snapshot_address or '-'
    if len(customer_address) > 160:
        customer_address = customer_address[:157] + '...'

    text = (
        "Customer delete approval request\n\n"
        f"Requested by: {requester_name}\n"
        f"Customer ID: {approval.customer_id}\n"
        f"Name: {customer_name}\n"
        f"Phone: {customer_phone}\n"
        f"IC: {customer_ic}\n"
        f"Address: {customer_address}\n"
        f"Other records with same name: {approval.duplicate_name_count}\n"
        f"Reason: {approval.reason or 'Duplicate customer cleanup'}\n"
        "Approve or decline this delete request."
    )
    keyboard = {"inline_keyboard": [[
        {"text": "Approve Delete", "callback_data": f"custdelapprove_{approval.approval_token}"},
        {"text": "Decline Delete", "callback_data": f"custdeldecline_{approval.approval_token}"}
    ]]}

    try:
        resp = http_req.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": chat_id, "text": text, "reply_markup": keyboard},
            timeout=10
        )
        print(f"[TELEGRAM] Customer delete approval response: {resp.status_code} {resp.text[:200]}")
        if resp.ok:
            approval.telegram_message_id = resp.json()['result']['message_id']
            db.session.commit()
        return resp.ok
    except Exception as e:
        print(f"[TELEGRAM] Customer delete approval exception: {e}")
        app.logger.error(f"Customer delete approval send error: {e}")
        return False

def tg_edit_message(chat_id, message_id, text, parse_mode='Markdown'):
    """Edit a Telegram message after action taken."""
    token = SystemSettings.get('telegram_bot_token')
    if not token:
        return
    try:
        payload = {"chat_id": chat_id, "message_id": message_id, "text": text}
        if parse_mode:
            payload["parse_mode"] = parse_mode
        http_req.post(
            f"https://api.telegram.org/bot{token}/editMessageText",
            json=payload,
            timeout=10
        )
    except Exception as e:
        app.logger.error(f"Telegram edit error: {e}")

def tg_answer_callback(callback_id, text):
    token = SystemSettings.get('telegram_bot_token')
    if not token:
        return
    try:
        http_req.post(
            f"https://api.telegram.org/bot{token}/answerCallbackQuery",
            json={"callback_query_id": callback_id, "text": text},
            timeout=10
        )
    except Exception:
        pass

def send_agent_device_confirmation(user, device):
    """Send Telegram Confirm/Deny message to agent for new device login."""
    token = SystemSettings.get('telegram_bot_token')
    chat_id = user.telegram_chat_id
    app.logger.info(f"[TG-DEVICE] Sending to chat_id={chat_id!r} token_len={len(device.approval_token) if device.approval_token else 0} device_id={device.id}")
    if not token or not chat_id:
        app.logger.warning(f"[TG-DEVICE] Aborted — bot_token set={bool(token)} chat_id set={bool(chat_id)}")
        return False
    try:
        text = (
            f"🔐 *New Device Login Request*\n\n"
            f"Hello *{user.full_name}*,\n\n"
            f"A login attempt was made to your account from a *new device*.\n\n"
            f"📍 IP: `{device.ip_address}`\n\n"
            f"If this was you, tap *Allow*. If not, tap *Deny* to block it."
        )
        keyboard = {"inline_keyboard": [[
            {"text": "✅ Allow — it's me", "callback_data": f"agentconfirm_{device.approval_token}"},
            {"text": "❌ Deny — not me", "callback_data": f"agentdeny_{device.approval_token}"}
        ]]}
        resp = http_req.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "Markdown", "reply_markup": keyboard},
            timeout=10
        )
        app.logger.info(f"[TG-DEVICE] Response {resp.status_code}: {resp.text[:200]}")
        return resp.ok
    except Exception as e:
        app.logger.error(f"[TG-DEVICE] Exception: {e}")
        return False

def _notify_admin_new_agent_device(user, device, ip_address):
    """Notify boss via Telegram when an agent logs in from a new unlinked device."""
    token = SystemSettings.get('telegram_bot_token')
    chat_id = SystemSettings.get('telegram_boss_chat_id')
    if not token or not chat_id:
        return
    try:
        text = (
            f"📱 *New Agent Device*\n\n"
            f"Agent *{user.full_name}* (`{user.username}`) logged in from a *new device*.\n\n"
            f"📍 IP: `{ip_address}`\n\n"
            f"Approve or block this device from the admin panel, or click below."
        )
        keyboard = {"inline_keyboard": [[
            {"text": "✅ Approve", "callback_data": f"agentconfirm_{device.approval_token}"},
            {"text": "❌ Block", "callback_data": f"agentdeny_{device.approval_token}"}
        ]]}
        http_req.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "Markdown", "reply_markup": keyboard},
            timeout=10
        )
    except Exception as e:
        app.logger.error(f"[TELEGRAM] Admin agent device notify error: {e}")

def send_agent_ip_change_confirmation(user, device, new_ip):
    """Send Telegram IP-change confirmation to the agent themselves."""
    token = SystemSettings.get('telegram_bot_token')
    chat_id = user.telegram_chat_id
    if not token or not chat_id:
        return False
    try:
        text = (
            f"🌐 *New Location Login Detected*\n\n"
            f"Hello *{user.full_name}*,\n\n"
            f"Your account was accessed from a *new IP address*.\n\n"
            f"📍 Previous IP: `{device.ip_address}`\n"
            f"📍 New IP: `{new_ip}`\n\n"
            f"If this was you, tap *Allow*. If not, tap *Deny* to block access."
        )
        keyboard = {"inline_keyboard": [[
            {"text": "✅ Allow — it's me", "callback_data": f"agentconfirm_{device.approval_token}"},
            {"text": "❌ Deny — not me", "callback_data": f"agentdeny_{device.approval_token}"}
        ]]}
        resp = http_req.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "Markdown", "reply_markup": keyboard},
            timeout=10
        )
        return resp.ok
    except Exception as e:
        app.logger.error(f"[TELEGRAM] Agent IP change confirmation error: {e}")
        return False

def _notify_admin_agent_ip_change(user, device, new_ip):
    """Notify admin via Telegram when an agent's approved device logs in from a new IP."""
    token = SystemSettings.get('telegram_bot_token')
    chat_id = SystemSettings.get('telegram_boss_chat_id')
    if not token or not chat_id:
        return
    try:
        text = (
            f"🌐 *Agent IP Change Detected*\n\n"
            f"Agent *{user.full_name}* (`{user.username}`) logged in from a *new IP address*.\n\n"
            f"📍 Previous IP: `{device.ip_address}`\n"
            f"📍 New IP: `{new_ip}`\n\n"
            f"Approve to allow login from new IP, or block to deny."
        )
        keyboard = {"inline_keyboard": [[
            {"text": "✅ Approve", "callback_data": f"agentconfirm_{device.approval_token}"},
            {"text": "❌ Block", "callback_data": f"agentdeny_{device.approval_token}"}
        ]]}
        http_req.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "Markdown", "reply_markup": keyboard},
            timeout=10
        )
    except Exception as e:
        app.logger.error(f"[TELEGRAM] Agent IP change notify error: {e}")

def delete_customer_with_related_records(customer):
    CustomerEnrichment.query.filter_by(customer_id=customer.id).delete(synchronize_session=False)
    CustomerEnrichmentJob.query.filter_by(customer_id=customer.id).delete(synchronize_session=False)
    DuplicateRecordLog.query.filter_by(existing_customer_id=customer.id).delete(synchronize_session=False)
    db.session.delete(customer)

def process_customer_delete_callback(action, token, callback_id=None, chat_id=None, message_id=None):
    approval = CustomerDeleteApproval.query.filter_by(approval_token=token).first()
    if not approval:
        if callback_id:
            tg_answer_callback(callback_id, 'Delete request not found.')
        return True

    if approval.status != 'pending':
        if callback_id:
            tg_answer_callback(callback_id, 'This delete request was already handled.')
        return True

    if approval.expires_at and datetime.utcnow() > approval.expires_at:
        approval.status = 'expired'
        approval.processed_at = datetime.utcnow()
        db.session.commit()
        if callback_id:
            tg_answer_callback(callback_id, 'This delete request has expired.')
        if chat_id and message_id:
            tg_edit_message(chat_id, message_id, 'Customer delete request expired.', parse_mode=None)
        return True

    customer = CustomerData.query.get(approval.customer_id)
    customer_name = approval.snapshot_name or f'Customer #{approval.customer_id}'

    if action == 'custdelapprove':
        if customer:
            delete_customer_with_related_records(customer)
        approval.status = 'approved'
        approval.processed_at = datetime.utcnow()
        db.session.commit()
        if callback_id:
            tg_answer_callback(callback_id, f'Delete approved for {customer_name}')
        if chat_id and message_id:
            result_text = 'Record deleted.' if customer else 'Record was already missing.'
            tg_edit_message(
                chat_id,
                message_id,
                f'Customer delete approved.\n\nName: {customer_name}\nCustomer ID: {approval.customer_id}\nResult: {result_text}',
                parse_mode=None
            )
        return True

    if action == 'custdeldecline':
        approval.status = 'declined'
        approval.processed_at = datetime.utcnow()
        db.session.commit()
        if callback_id:
            tg_answer_callback(callback_id, f'Delete declined for {customer_name}')
        if chat_id and message_id:
            tg_edit_message(
                chat_id,
                message_id,
                f'Customer delete declined.\n\nName: {customer_name}\nCustomer ID: {approval.customer_id}',
                parse_mode=None
            )
        return True

    return False

def _max_agent_devices():
    try:
        return int(SystemSettings.get('max_agent_devices', '3'))
    except (TypeError, ValueError):
        return 3


def check_device_trusted(user_id, fingerprint, ip_address):
    """
    Returns 'approved', 'pending', 'blocked', 'new', 'ip_changed', or 'device_limit'.
    'device_limit' means the agent already has max approved devices and cannot add more.
    """
    device = AgentDevice.query.filter_by(user_id=user_id, fingerprint=fingerprint).first()
    if device:
        device.last_seen = datetime.utcnow()
        if (device.status == 'approved'
                and is_agent_ip_approval_enabled()
                and device.ip_address
                and device.ip_address != ip_address):
            device.pending_ip = ip_address
            device.status = 'pending'
            device.approval_token = secrets.token_hex(25)
            db.session.commit()
            return 'ip_changed'
        device.ip_address = ip_address
        db.session.commit()
        return device.status
    else:
        # Check device limit before registering a new device
        approved_count = AgentDevice.query.filter_by(
            user_id=user_id, status='approved'
        ).count()
        agent_user = User.query.get(user_id)
        limit = agent_user.max_devices if (agent_user and agent_user.max_devices) else _max_agent_devices()
        if approved_count >= limit:
            return 'device_limit'

        new_device = AgentDevice(
            user_id=user_id,
            fingerprint=fingerprint,
            ip_address=ip_address,
            label=f'Device ({ip_address})',
            status='pending',
            approval_token=secrets.token_hex(25)
        )
        db.session.add(new_device)
        db.session.commit()
        return 'new'

def get_location_from_ip(ip):
    try:
        return "Location lookup not implemented"
    except:
        return "Unknown"

def detect_device_category(user_agent):
    user_agent_lower = (user_agent or '').lower()
    if any(token in user_agent_lower for token in ['iphone', 'android', 'mobile']):
        return 'phone'
    if any(token in user_agent_lower for token in ['ipad', 'tablet']):
        return 'tablet'
    if any(token in user_agent_lower for token in ['windows', 'macintosh', 'mac os', 'linux', 'x11', 'cros']):
        return 'laptop'
    return 'unknown'

def detect_browser(user_agent):
    user_agent_lower = (user_agent or '').lower()
    if 'edg/' in user_agent_lower:
        return 'Microsoft Edge'
    if 'chrome/' in user_agent_lower and 'edg/' not in user_agent_lower:
        return 'Google Chrome'
    if 'firefox/' in user_agent_lower:
        return 'Mozilla Firefox'
    if 'safari/' in user_agent_lower and 'chrome/' not in user_agent_lower:
        return 'Safari'
    if 'opr/' in user_agent_lower or 'opera/' in user_agent_lower:
        return 'Opera'
    return 'Unknown Browser'

def detect_operating_system(user_agent):
    user_agent_lower = (user_agent or '').lower()
    if 'windows' in user_agent_lower:
        return 'Windows'
    if 'iphone' in user_agent_lower or 'ipad' in user_agent_lower or 'ios' in user_agent_lower:
        return 'iOS'
    if 'android' in user_agent_lower:
        return 'Android'
    if 'mac os' in user_agent_lower or 'macintosh' in user_agent_lower:
        return 'macOS'
    if 'linux' in user_agent_lower or 'x11' in user_agent_lower:
        return 'Linux'
    return 'Unknown OS'

def detect_connection_type(user_agent):
    user_agent_lower = user_agent.lower()
    if 'mobile' in user_agent_lower or 'android' in user_agent_lower or 'iphone' in user_agent_lower:
        return 'mobile'
    return 'wifi'

def analyze_login_behavior(user, login_log):
    """Create descriptive security events for multi-device login patterns."""
    previous_logs = LoginLog.query.filter(
        LoginLog.user_id == user.id,
        LoginLog.id != login_log.id
    ).order_by(LoginLog.login_time.desc()).limit(20).all()

    previous_device_seen = any(
        login_log.fingerprint and log.fingerprint == login_log.fingerprint
        for log in previous_logs
    )

    if previous_logs:
        previous = previous_logs[0]
        time_gap = login_log.login_time - previous.login_time if previous.login_time and login_log.login_time else None
        minutes_gap = int(time_gap.total_seconds() // 60) if time_gap else None

        if login_log.fingerprint and previous.fingerprint and login_log.fingerprint != previous.fingerprint and minutes_gap is not None and minutes_gap <= 120:
            log_security_event(
                'login_behavior',
                'warning',
                login_log.ip_address,
                user=user,
                details=f'Rapid device switch detected: {previous.device_category or "unknown device"} to {login_log.device_category or "unknown device"} within {minutes_gap} minute(s).'
            )
        elif login_log.ip_address and previous.ip_address and login_log.ip_address != previous.ip_address and previous_device_seen:
            log_security_event(
                'login_behavior',
                'info',
                login_log.ip_address,
                user=user,
                details=f'Known {login_log.device_category or "device"} login observed from a new IP address.'
            )

    if not previous_device_seen and login_log.fingerprint:
        log_security_event(
            'login_behavior',
            'warning',
            login_log.ip_address,
            user=user,
            details=f'New {login_log.device_category or "device"} detected for this account.'
        )
    elif previous_device_seen:
        log_security_event(
            'login_behavior',
            'info',
            login_log.ip_address,
            user=user,
            details=f'Known {login_log.device_category or "device"} login recorded.'
        )

    window_start = datetime.utcnow() - timedelta(hours=5)
    recent_logs = LoginLog.query.filter(
        LoginLog.user_id == user.id,
        LoginLog.login_time >= window_start
    ).order_by(LoginLog.login_time.asc()).all()

    device_keys = []
    for log in recent_logs:
        key = log.fingerprint or f'{log.device_category}:{log.ip_address}'
        if key and key not in device_keys:
            device_keys.append(key)

    if len(device_keys) >= 3:
        elapsed_hours = max(1, int((recent_logs[-1].login_time - recent_logs[0].login_time).total_seconds() // 3600))
        log_security_event(
            'login_behavior',
            'warning',
            login_log.ip_address,
            user=user,
            details=f'Multiple device anomaly: {len(device_keys)} devices used within {elapsed_hours} hour(s). Latest device: {login_log.device_category or "unknown device"}.'
        )

def clean_data_value(value):
    if isinstance(value, (pd.Series, np.ndarray, list, tuple)):
        # Duplicate/misaligned column selection can hand back an array-like
        # instead of a scalar. Take the first real value rather than
        # silently stringifying the whole thing into the DB.
        value = next((v for v in value if not pd.isna(v)), '')

    if pd.isna(value):
        return ''

    value_str = str(value).strip()
    value_str = ' '.join(value_str.split())

    if value_str.lower() in ['null', 'none', 'nan', 'n/a', '']:
        return ''

    return value_str


MERGE_STANDARD_FIELDS = ['name', 'contact_number', 'ic_number', 'address', 'email']
MERGE_FACT_BUCKETS = {
    'names': 'name',
    'phones': 'contact_number',
    'emails': 'email',
    'addresses': 'address',
}
MERGE_RESERVED_EXTRA_KEYS = {
    'identity_profile',
    'identity_facts',
    'merge_history',
    'field_conflicts',
    'household_keys',
}
MERGE_IGNORED_PAYLOAD_KEYS = {'upload_id', 'merged_from_id'}


def _json_object(value):
    if isinstance(value, dict):
        return dict(value)
    if not value:
        return {}
    try:
        data = json.loads(value)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _normalize_identity_ic(value):
    return re.sub(r'[^0-9]', '', clean_data_value(value))


def _normalize_identity_phone(value):
    raw = re.sub(r'[^0-9]', '', clean_data_value(value))
    if raw.startswith('60') and len(raw) > 9:
        raw = raw[2:]
    return raw


def _normalize_identity_email(value):
    value = clean_data_value(value)
    return value.casefold() if value else ''


def _normalize_identity_name(value):
    value = clean_data_value(value)
    return re.sub(r'\s+', ' ', value).strip().casefold() if value else ''


def _normalize_identity_address(value):
    value = clean_data_value(value)
    if not value:
        return ''
    value = value.upper()
    replacements = {
        r'\bJLN\b': 'JALAN',
        r'\bJAL\b': 'JALAN',
        r'\bTMN\b': 'TAMAN',
        r'\bKG\b': 'KAMPUNG',
        r'\bAPT\b': 'APARTMENT',
    }
    for pat, repl in replacements.items():
        value = re.sub(pat, repl, value)
    value = re.sub(r'[^A-Z0-9]+', ' ', value)
    return re.sub(r'\s+', ' ', value).strip()


def _normalize_fact_value(bucket, value):
    if bucket == 'names':
        return _normalize_identity_name(value)
    if bucket == 'phones':
        return _normalize_identity_phone(value)
    if bucket == 'emails':
        return _normalize_identity_email(value)
    if bucket == 'addresses':
        return _normalize_identity_address(value)
    return clean_data_value(value)


def _append_unique(items, value):
    if value in (None, '', [], {}):
        return
    if value not in items:
        items.append(value)


def _canonical_json(value):
    try:
        return json.dumps(value, sort_keys=True, ensure_ascii=False)
    except Exception:
        return str(value)


def _merge_field_conflict(extra, key, value):
    if value in (None, '', [], {}):
        return
    conflicts = extra.setdefault('field_conflicts', {})
    bucket = conflicts.setdefault(key, [])
    sigs = {_canonical_json(item) for item in bucket}
    candidate_sig = _canonical_json(value)
    if candidate_sig not in sigs:
        bucket.append(value)


def _ensure_identity_store(extra):
    extra = dict(extra or {})
    profile = extra.setdefault('identity_profile', {})
    profile.setdefault('ic_norm', '')
    profile.setdefault('ic_aliases', [])
    facts = extra.setdefault('identity_facts', {})
    for bucket in MERGE_FACT_BUCKETS:
        facts.setdefault(bucket, [])
    extra.setdefault('merge_history', [])
    extra.setdefault('field_conflicts', {})
    extra.setdefault('household_keys', [])
    return extra


def _value_richness_score(field_name, value, seen_count=1, current_norm=''):
    value = clean_data_value(value)
    if not value:
        return -1
    score = len(value) + (max(seen_count, 1) * 12)
    if field_name == 'name':
        score += min(len(value.split()), 5) * 8
        if len(value) >= 8:
            score += 10
    elif field_name == 'contact_number':
        digits = _normalize_identity_phone(value)
        score += 25 if 9 <= len(digits) <= 11 else 0
        score += 8 if value.startswith('0') else 0
    elif field_name == 'email':
        score += 35 if '@' in value and '.' in value else 0
    elif field_name == 'address':
        words = len(value.split())
        score += min(words, 12) * 5
        score += 20 if any(ch.isdigit() for ch in value) else 0
    elif field_name == 'ic_number':
        digits = _normalize_identity_ic(value)
        score += 30 if len(digits) >= 12 else len(digits)
        score += 6 if '-' in value else 0

    norm = ''
    if field_name == 'name':
        norm = _normalize_identity_name(value)
    elif field_name == 'contact_number':
        norm = _normalize_identity_phone(value)
    elif field_name == 'email':
        norm = _normalize_identity_email(value)
    elif field_name == 'address':
        norm = _normalize_identity_address(value)
    elif field_name == 'ic_number':
        norm = _normalize_identity_ic(value)

    if current_norm and norm and norm == current_norm:
        score += 14
    return score


def _upsert_fact(extra, bucket, value, source_label=None, customer_id=None):
    clean = clean_data_value(value)
    norm = _normalize_fact_value(bucket, clean)
    if not clean or not norm:
        return

    extra = _ensure_identity_store(extra)
    facts = extra['identity_facts'][bucket]
    key_name = 'address_hash' if bucket == 'addresses' else 'norm'
    key_value = hashlib.sha1(norm.encode('utf-8')).hexdigest() if bucket == 'addresses' else norm

    for item in facts:
        item_key = item.get('address_hash') if bucket == 'addresses' else item.get('norm')
        if item_key == key_value:
            item['seen_count'] = int(item.get('seen_count') or 1) + 1
            if source_label:
                _append_unique(item.setdefault('sources', []), source_label)
            if customer_id:
                _append_unique(item.setdefault('customer_ids', []), int(customer_id))
            if _value_richness_score(MERGE_FACT_BUCKETS[bucket], clean, item['seen_count']) > _value_richness_score(MERGE_FACT_BUCKETS[bucket], item.get('value', ''), item['seen_count']):
                item['value'] = clean
            return

    entry = {
        'value': clean,
        'norm': norm,
        'seen_count': 1,
        'sources': [source_label] if source_label else [],
        'customer_ids': [int(customer_id)] if customer_id else [],
        'last_seen_at': datetime.utcnow().isoformat()
    }
    if bucket == 'addresses':
        entry['address_hash'] = key_value
    facts.append(entry)


def _upsert_ic_alias(extra, ic_value, source_label=None, customer_id=None):
    clean = clean_data_value(ic_value)
    norm = _normalize_identity_ic(clean)
    if not norm:
        return
    extra = _ensure_identity_store(extra)
    profile = extra['identity_profile']
    profile['ic_norm'] = norm
    aliases = profile.setdefault('ic_aliases', [])
    for alias in aliases:
        if alias.get('norm') == norm and alias.get('value') == clean:
            if source_label:
                _append_unique(alias.setdefault('sources', []), source_label)
            if customer_id:
                _append_unique(alias.setdefault('customer_ids', []), int(customer_id))
            alias['seen_count'] = int(alias.get('seen_count') or 1) + 1
            return
    aliases.append({
        'value': clean,
        'norm': norm,
        'seen_count': 1,
        'sources': [source_label] if source_label else [],
        'customer_ids': [int(customer_id)] if customer_id else [],
    })


def _merge_existing_fact_store(target_extra, source_extra, source_label=None, customer_id=None):
    target_extra = _ensure_identity_store(target_extra)
    source_extra = _ensure_identity_store(source_extra)

    for alias in source_extra.get('identity_profile', {}).get('ic_aliases', []):
        _upsert_ic_alias(target_extra, alias.get('value'), source_label or 'merged_record', customer_id)

    for bucket in MERGE_FACT_BUCKETS:
        for entry in source_extra.get('identity_facts', {}).get(bucket, []):
            _upsert_fact(target_extra, bucket, entry.get('value'), source_label or 'merged_record', customer_id)

    for value in source_extra.get('household_keys', []):
        _append_unique(target_extra['household_keys'], value)


def _merge_generic_additional_data(target_extra, source_extra):
    for key, value in (source_extra or {}).items():
        if key in MERGE_RESERVED_EXTRA_KEYS or value in (None, '', [], {}):
            continue
        if key not in target_extra or target_extra.get(key) in (None, '', [], {}):
            target_extra[key] = value
        elif _canonical_json(target_extra.get(key)) != _canonical_json(value):
            _merge_field_conflict(target_extra, key, value)


def _sync_customer_identity_facts(customer, extra, source_label=None):
    extra = _ensure_identity_store(extra)
    source_label = source_label or f'customer:{customer.id}'
    _upsert_ic_alias(extra, customer.ic_number, source_label, customer.id)
    _upsert_fact(extra, 'names', customer.name, source_label, customer.id)
    _upsert_fact(extra, 'phones', customer.contact_number, source_label, customer.id)
    _upsert_fact(extra, 'emails', customer.email, source_label, customer.id)
    _upsert_fact(extra, 'addresses', customer.address, source_label, customer.id)
    return extra


def _select_best_fact_value(customer, extra, bucket, current_value=''):
    field_name = MERGE_FACT_BUCKETS[bucket]
    current_norm = _normalize_fact_value(bucket, current_value)
    candidates = extra.get('identity_facts', {}).get(bucket, [])
    if not candidates:
        return clean_data_value(current_value)
    ranked = sorted(
        candidates,
        key=lambda item: (
            _value_richness_score(field_name, item.get('value', ''), item.get('seen_count', 1), current_norm),
            item.get('seen_count', 1),
            len(clean_data_value(item.get('value', '')))
        ),
        reverse=True
    )
    return clean_data_value(ranked[0].get('value', '')) or clean_data_value(current_value)


def _select_best_ic_alias(extra, current_value=''):
    aliases = extra.get('identity_profile', {}).get('ic_aliases', [])
    if not aliases:
        return clean_data_value(current_value)
    current_norm = _normalize_identity_ic(current_value)
    ranked = sorted(
        aliases,
        key=lambda alias: (
            _value_richness_score('ic_number', alias.get('value', ''), alias.get('seen_count', 1), current_norm),
            alias.get('seen_count', 1),
            len(clean_data_value(alias.get('value', '')))
        ),
        reverse=True
    )
    return clean_data_value(ranked[0].get('value', '')) or clean_data_value(current_value)


def _refresh_household_keys(extra):
    facts = extra.get('identity_facts', {}).get('addresses', [])
    hashes = []
    for entry in facts:
        addr_hash = entry.get('address_hash')
        if addr_hash:
            _append_unique(hashes, addr_hash)
    extra['household_keys'] = hashes


def _apply_primary_identity_fields(customer, extra):
    extra = _ensure_identity_store(extra)
    customer.name = _select_best_fact_value(customer, extra, 'names', customer.name)
    customer.contact_number = _select_best_fact_value(customer, extra, 'phones', customer.contact_number)
    customer.email = _select_best_fact_value(customer, extra, 'emails', customer.email)
    customer.address = _select_best_fact_value(customer, extra, 'addresses', customer.address)
    best_ic = _select_best_ic_alias(extra, customer.ic_number)
    if best_ic:
        customer.ic_number = best_ic
    profile = extra.setdefault('identity_profile', {})
    profile['ic_norm'] = _normalize_identity_ic(customer.ic_number)
    _refresh_household_keys(extra)
    customer.additional_data = json.dumps(extra, ensure_ascii=False)
    customer.updated_at = datetime.utcnow()
    return customer


def _merge_payload_into_customer(customer, payload, source_label=None, customer_id=None, merge_reason=None):
    extra = _ensure_identity_store(_json_object(customer.additional_data))
    _sync_customer_identity_facts(customer, extra, source_label=f'customer:{customer.id}:before')

    source_label = source_label or 'payload'
    _upsert_ic_alias(extra, payload.get('ic_number'), source_label, customer_id)
    _upsert_fact(extra, 'names', payload.get('name'), source_label, customer_id)
    _upsert_fact(extra, 'phones', payload.get('contact_number'), source_label, customer_id)
    _upsert_fact(extra, 'emails', payload.get('email'), source_label, customer_id)
    _upsert_fact(extra, 'addresses', payload.get('address'), source_label, customer_id)

    source_extra = _json_object(payload.get('additional_data'))
    if source_extra:
        _merge_existing_fact_store(extra, source_extra, source_label=source_label, customer_id=customer_id)
        _merge_generic_additional_data(extra, source_extra)

    for key, value in payload.items():
        if key in MERGE_STANDARD_FIELDS or key == 'additional_data' or key in MERGE_IGNORED_PAYLOAD_KEYS:
            continue
        if value not in (None, '', [], {}):
            if key not in extra or extra.get(key) in (None, '', [], {}):
                extra[key] = value
            elif _canonical_json(extra.get(key)) != _canonical_json(value):
                _merge_field_conflict(extra, key, value)

    if merge_reason:
        history = extra.setdefault('merge_history', [])
        event = {
            'merged_at': datetime.utcnow().isoformat(),
            'reason': merge_reason,
            'source_label': source_label,
        }
        if customer_id:
            event['source_customer_id'] = int(customer_id)
        history.append(event)

    return _apply_primary_identity_fields(customer, extra)


def _customer_payload(customer):
    return {
        'name': customer.name,
        'contact_number': customer.contact_number,
        'ic_number': customer.ic_number,
        'address': customer.address,
        'email': customer.email,
        'additional_data': customer.additional_data,
    }


def _customer_merge_score(customer):
    score = 0
    for field in MERGE_STANDARD_FIELDS:
        score += 1 if clean_data_value(getattr(customer, field, '')) else 0
    extra = _json_object(customer.additional_data)
    score += len(extra.keys()) * 0.05
    score += max(0, 1000000 - int(customer.id)) * 0.000001
    return score


def _choose_primary_customer(customers):
    return max(customers, key=_customer_merge_score)


def merge_customer_records(primary, others, merge_reason='merge'):
    if not primary:
        return None
    others = [c for c in others if c and c.id != primary.id]
    extra = _ensure_identity_store(_json_object(primary.additional_data))
    _sync_customer_identity_facts(primary, extra, source_label=f'customer:{primary.id}:primary')
    primary.additional_data = json.dumps(extra, ensure_ascii=False)

    for other in others:
        merge_payload = _customer_payload(other)
        merge_payload['merged_from_id'] = other.id
        _merge_payload_into_customer(
            primary,
            merge_payload,
            source_label=f'customer:{other.id}:merged',
            customer_id=other.id,
            merge_reason=merge_reason
        )

    return primary


def _normalized_ic_query_expression():
    if app.config['SQLALCHEMY_DATABASE_URI'].startswith('postgresql'):
        return func.regexp_replace(CustomerData.ic_number, '[^0-9]', '', 'g')
    return func.replace(func.replace(CustomerData.ic_number, '-', ''), ' ', '')


def find_customer_by_normalized_ic(ic_value):
    ic_norm = _normalize_identity_ic(ic_value)
    if not ic_norm:
        return None
    expr = _normalized_ic_query_expression()
    return (CustomerData.query
            .filter(CustomerData.ic_number.isnot(None), CustomerData.ic_number != '')
            .filter(expr == ic_norm)
            .order_by(CustomerData.id.asc())
            .first())


def fetch_customers_by_normalized_ics(ic_values):
    ic_norms = sorted({_normalize_identity_ic(v) for v in ic_values if _normalize_identity_ic(v)})
    if not ic_norms:
        return {}
    expr = _normalized_ic_query_expression()
    rows = (CustomerData.query
            .filter(CustomerData.ic_number.isnot(None), CustomerData.ic_number != '')
            .filter(expr.in_(ic_norms))
            .all())
    return {_normalize_identity_ic(row.ic_number): row for row in rows if _normalize_identity_ic(row.ic_number)}


def _displayable_additional_data(additional):
    cleaned = {}
    for key, value in (additional or {}).items():
        if key in MERGE_RESERVED_EXTRA_KEYS:
            continue
        if value in ('', 'nan', 'None', None, [], {}):
            continue
        cleaned[key] = value
    return cleaned


def _identity_summary_from_customer(customer):
    extra = _ensure_identity_store(_json_object(customer.additional_data))
    _sync_customer_identity_facts(customer, extra, source_label=f'customer:{customer.id}:detail_view')
    _refresh_household_keys(extra)
    profile = extra.get('identity_profile', {})
    facts = extra.get('identity_facts', {})

    def _fact_values(bucket):
        seen = []
        for item in facts.get(bucket, []):
            value = clean_data_value(item.get('value'))
            if value and value not in seen:
                seen.append(value)
        return seen

    ic_aliases = []
    for alias in profile.get('ic_aliases', []):
        value = clean_data_value(alias.get('value'))
        if value and value not in ic_aliases:
            ic_aliases.append(value)

    return {
        'ic_norm': profile.get('ic_norm') or _normalize_identity_ic(customer.ic_number),
        'ic_aliases': ic_aliases,
        'names': _fact_values('names'),
        'phones': _fact_values('phones'),
        'emails': _fact_values('emails'),
        'addresses': _fact_values('addresses'),
        'household_keys': list(extra.get('household_keys', [])),
        'merge_history_count': len(extra.get('merge_history', [])),
    }, extra


def _address_search_tokens(address_norm):
    stop_words = {'NO', 'JALAN', 'TAMAN', 'LORONG', 'BLOK', 'BLOCK', 'TINGKAT', 'LEVEL', 'UNIT', 'LOT', 'KAMPUNG'}
    parts = [part for part in address_norm.split() if part]
    alpha = [part for part in parts if part.isalpha() and len(part) >= 4 and part not in stop_words]
    numeric = [part for part in parts if part.isdigit() and len(part) >= 2]
    alpha = sorted(alpha, key=len, reverse=True)[:2]
    numeric = numeric[:1]
    return alpha + numeric


def _household_lookup_enabled():
    if app.config['SQLALCHEMY_DATABASE_URI'].startswith('sqlite'):
        return os.environ.get('ENABLE_SQLITE_HOUSEHOLD_LOOKUP', '').strip().lower() in ('1', 'true', 'yes', 'on')
    return True


def _same_household_candidates(customer, identity_summary, limit=8):
    if not _household_lookup_enabled():
        return []
    target_norms = {_normalize_identity_address(addr) for addr in identity_summary.get('addresses', []) if _normalize_identity_address(addr)}
    if not target_norms and customer.address:
        addr_norm = _normalize_identity_address(customer.address)
        if addr_norm:
            target_norms.add(addr_norm)
    if not target_norms:
        return []

    candidates = []
    seen_ids = set()
    address_hashes = [hashlib.sha1(norm.encode('utf-8')).hexdigest() for norm in target_norms]

    hash_filters = [CustomerData.additional_data.ilike(f'%\"address_hash\": \"{addr_hash}\"%') for addr_hash in address_hashes]
    if hash_filters:
        hashed_matches = (CustomerData.query
                          .filter(CustomerData.id != customer.id)
                          .filter(or_(*hash_filters))
                          .limit(max(limit * 3, 15))
                          .all())
    else:
        hashed_matches = []

    def _add_candidate(row, match_type):
        if row.id in seen_ids:
            return
        row_summary, _ = _identity_summary_from_customer(row)
        row_norms = {_normalize_identity_address(addr) for addr in row_summary.get('addresses', []) if _normalize_identity_address(addr)}
        if not row_norms and row.address:
            norm = _normalize_identity_address(row.address)
            if norm:
                row_norms.add(norm)
        shared = sorted(target_norms & row_norms)
        if not shared:
            return
        candidates.append({
            'id': row.id,
            'name': row.name,
            'contact_number': row.contact_number,
            'ic_number': row.ic_number,
            'address': row.address,
            'email': row.email,
            'shared_address': shared[0],
            'match_type': match_type,
        })
        seen_ids.add(row.id)

    for row in hashed_matches:
        _add_candidate(row, 'shared_address_hash')
        if len(candidates) >= limit:
            return candidates[:limit]

    search_tokens = _address_search_tokens(next(iter(target_norms)))
    if search_tokens:
        token_filters = [CustomerData.address.ilike(f'%{token}%') for token in search_tokens]
        broad_matches = (CustomerData.query
                         .filter(CustomerData.id != customer.id)
                         .filter(CustomerData.address.isnot(None), CustomerData.address != '')
                         .filter(and_(*token_filters))
                         .limit(max(limit * 10, 40))
                         .all())
        for row in broad_matches:
            _add_candidate(row, 'normalized_address_match')
            if len(candidates) >= limit:
                break

    return candidates[:limit]

def detect_columns_by_content(df):
    """
    Detect columns by analyzing the actual data content
    """
    found_columns = {}
    
    for idx, col in enumerate(df.columns):
        sample_data = df[col].dropna().head(20).astype(str)
        
        if len(sample_data) == 0:
            continue
        
        ic_pattern = re.compile(r'^\d{6}-\d{2}-\d{4}$|^\d{12}$|^\d{6}\d{2}\d{4}$')
        ic_count = sum(1 for val in sample_data if ic_pattern.match(re.sub(r'[^0-9]', '', val)))
        
        phone_pattern = re.compile(r'^\+?\d{1,3}[-.\s]?\d{3,4}[-.\s]?\d{3,4}$|^\d{9,11}$')
        phone_count = sum(1 for val in sample_data if phone_pattern.match(re.sub(r'[^0-9+]', '', val)))
        
        email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
        email_count = sum(1 for val in sample_data if email_pattern.match(val))
        
        address_count = sum(1 for val in sample_data if len(val) > 20 and ' ' in val and any(c.isdigit() for c in val))
        
        name_count = sum(1 for val in sample_data if re.match(r'^[A-Za-z\s]+$', val) and len(val.split()) >= 2 and len(val) < 50)
        
        if ic_count > len(sample_data) * 0.3 and 'ic_number' not in found_columns:
            found_columns['ic_number'] = col
        elif phone_count > len(sample_data) * 0.3 and 'contact_number' not in found_columns:
            found_columns['contact_number'] = col
        elif email_count > len(sample_data) * 0.3 and 'email' not in found_columns:
            found_columns['email'] = col
        elif address_count > len(sample_data) * 0.3 and 'address' not in found_columns:
            found_columns['address'] = col
        elif name_count > len(sample_data) * 0.3 and 'name' not in found_columns:
            found_columns['name'] = col
    
    return found_columns

def find_best_header_row(df_raw, max_rows=10):
    """
    Find the most likely header row from the first few rows of a worksheet.
    This helps with Excel files that contain title rows before the real table.
    """
    header_keywords = {
        'name', 'full name', 'customer name', 'nama',
        'contact', 'contact no', 'contact number', 'phone', 'mobile', 'tel', 'telefon',
        'ic', 'ic number', 'nric', 'id', 'id number', 'no ic',
        'address', 'alamat', 'email', 'e-mail', 'ren no', 'comm %', 'team leader', 'nick name'
    }

    best_index = 0
    best_score = -1
    rows_to_check = min(len(df_raw), max_rows)

    for idx in range(rows_to_check):
        row = df_raw.iloc[idx].fillna('').astype(str)
        normalized = [re.sub(r'\s+', ' ', value).strip().lower() for value in row]

        non_empty = [value for value in normalized if value]
        if not non_empty:
            continue

        keyword_score = sum(1 for value in non_empty if value in header_keywords)
        text_score = sum(1 for value in non_empty if len(value) <= 40 and not any(ch.isdigit() for ch in value))
        unique_score = len(set(non_empty))
        total_score = (keyword_score * 10) + (text_score * 2) + unique_score

        if total_score > best_score:
            best_score = total_score
            best_index = idx

    return best_index

def _stream_excel_rows(file_path, max_rows=None):
    """Stream rows using calamine (Rust) — fast even for large files."""
    try:
        from python_calamine import CalamineWorkbook
        wb = CalamineWorkbook.from_path(file_path)
        ws = wb.get_sheet_by_index(0)
        rows = []
        for i, row in enumerate(ws.iter_rows()):
            if max_rows is not None and i >= max_rows:
                break
            rows.append(['' if v is None else v for v in row])
        return rows
    except Exception:
        # Fallback to openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if max_rows is not None and i >= max_rows:
                break
            rows.append(['' if v is None else v for v in row])
        wb.close()
        return rows


def read_excel_file(file_path, nrows=None):
    """
    Read Excel file and recover the most likely header row.
    Uses calamine (Rust) for xlsx — fast even for very large files.
    If nrows is set, only that many data rows are returned (used for previews).
    """
    fetch_rows = (nrows + 15) if nrows is not None else None

    try:
        if file_path.lower().endswith('.csv'):
            df_raw = pd.read_csv(file_path, header=None, nrows=fetch_rows, dtype=str)
        else:
            raw = _stream_excel_rows(file_path, max_rows=fetch_rows)
            df_raw = pd.DataFrame(raw)
    except Exception:
        read_kwargs = {'header': None}
        if fetch_rows:
            read_kwargs['nrows'] = fetch_rows
        try:
            df_raw = pd.read_excel(file_path, engine='openpyxl', **read_kwargs)
        except Exception:
            df_raw = pd.read_excel(file_path, **read_kwargs)

    df_raw = df_raw.fillna('')

    header_row_idx = find_best_header_row(df_raw)
    header_row = df_raw.iloc[header_row_idx].astype(str)
    has_header = any(clean_data_value(value) for value in header_row)

    col_names = []
    used_names = set()
    for i, val in enumerate(header_row):
        val_str = clean_data_value(val)
        if not val_str:
            val_str = f'Column_{i}'
        base_name = val_str
        suffix = 1
        while val_str in used_names:
            suffix += 1
            val_str = f'{base_name}_{suffix}'
        used_names.add(val_str)
        col_names.append(val_str)

    df = df_raw.iloc[header_row_idx + 1:].reset_index(drop=True).copy()
    df.columns = col_names
    if nrows is not None:
        df = df.head(nrows)

    return df, has_header

def normalize_duplicate_value(value, field_name):
    value = clean_data_value(value)
    if not value:
        return ''

    if field_name in ['contact_number', 'ic_number']:
        return re.sub(r'[^0-9A-Za-z]+', '', value).upper()

    return value.casefold()

def build_aligned_dataframe(df, manual_mapping, extra_columns=None):
    """
    Build a database-friendly dataframe with standard columns first,
    then any selected extra columns using their chosen labels.
    """
    standard_columns = ['name', 'contact_number', 'ic_number', 'address', 'email']
    aligned = pd.DataFrame(index=df.index)

    for field in standard_columns:
        source_col = manual_mapping.get(field)
        if source_col and source_col in df.columns:
            aligned[field] = df[source_col].apply(clean_data_value)
        else:
            aligned[field] = ''

    if extra_columns:
        for source_col, label in extra_columns.items():
            if source_col in df.columns:
                target_label = clean_data_value(label) or source_col
                aligned[target_label] = df[source_col].apply(clean_data_value)

    aligned.insert(0, 'source_row', range(2, len(aligned) + 2))
    aligned = aligned.fillna('')
    return aligned

def detect_internal_duplicates(df, duplicate_fields):
    """
    Detect duplicates within the uploaded file using one or more chosen keys.
    """
    duplicate_fields = [field for field in duplicate_fields if field in df.columns]
    if not duplicate_fields:
        return pd.DataFrame(), pd.DataFrame()

    working = df.copy()
    duplicate_masks = []

    for field in duplicate_fields:
        normalized_col = f'__dup_{field}'
        working[normalized_col] = working[field].apply(lambda value: normalize_duplicate_value(value, field))
        duplicate_masks.append((working[normalized_col] != '') & working.duplicated(normalized_col, keep=False))

    combined_mask = duplicate_masks[0]
    for mask in duplicate_masks[1:]:
        combined_mask = combined_mask | mask

    duplicates = working.loc[combined_mask].copy()
    if duplicates.empty:
        return pd.DataFrame(), pd.DataFrame()

    group_parts = []
    for field in duplicate_fields:
        normalized_col = f'__dup_{field}'
        group_parts.append(
            duplicates[normalized_col].where(duplicates[normalized_col] != '', '').map(
                lambda value: f'{field}:{value}' if value else ''
            )
        )

    duplicate_keys = []
    for idx in duplicates.index:
        keys = [part.loc[idx] for part in group_parts if part.loc[idx]]
        duplicate_keys.append(' | '.join(keys))

    duplicates.insert(1, 'duplicate_key', duplicate_keys)
    duplicates.insert(
        2,
        'duplicate_by',
        duplicates['duplicate_key'].map(
            lambda value: ', '.join(sorted({part.split(':', 1)[0] for part in value.split(' | ') if ':' in part}))
        )
    )

    normalized_columns = [col for col in duplicates.columns if col.startswith('__dup_')]
    duplicates = duplicates.drop(columns=normalized_columns)
    deduped = df.drop(index=duplicates.index).copy()

    return duplicates, deduped

def apply_duplicate_strategy(df, duplicate_fields, strategy):
    """
    Return cleaned rows plus a duplicate report according to the chosen strategy.
    """
    duplicate_fields = [field for field in duplicate_fields if field in df.columns]
    if not duplicate_fields:
        return df.copy(), pd.DataFrame(), 0

    working = df.copy()
    normalized_columns = []
    report_masks = []

    for field in duplicate_fields:
        normalized_col = f'__norm_{field}'
        working[normalized_col] = working[field].apply(lambda value: normalize_duplicate_value(value, field))
        normalized_columns.append(normalized_col)
        report_masks.append((working[normalized_col] != '') & working.duplicated(normalized_col, keep=False))

    duplicate_report_mask = report_masks[0]
    for mask in report_masks[1:]:
        duplicate_report_mask = duplicate_report_mask | mask

    duplicate_report = working.loc[duplicate_report_mask].copy()
    if not duplicate_report.empty:
        duplicate_report.insert(
            1,
            'duplicate_by',
            duplicate_report.apply(
                lambda row: ', '.join([
                    field for field, normalized_col in zip(duplicate_fields, normalized_columns)
                    if row.get(normalized_col)
                    and (working[normalized_col] == row[normalized_col]).sum() > 1
                ]),
                axis=1
            )
        )
        duplicate_report = duplicate_report.drop(columns=normalized_columns)

    if strategy == 'keep_all':
        cleaned = df.copy()
    else:
        keep = 'first' if strategy == 'keep_first' else 'last'
        row_mask = pd.Series(False, index=working.index)
        for normalized_col in normalized_columns:
            non_empty = working[normalized_col] != ''
            row_mask = row_mask | working[non_empty].duplicated(normalized_col, keep=keep).reindex(working.index, fill_value=False)
        cleaned = working.loc[~row_mask].drop(columns=normalized_columns).copy()

    duplicates_removed = max(len(df) - len(cleaned), 0)
    return cleaned, duplicate_report, duplicates_removed

def create_sql_insert_file(df, table_name, output_path):
    safe_table_name = re.sub(r'[^A-Za-z0-9_]', '_', table_name) or 'imported_data'

    def sql_value(value):
        cleaned = clean_data_value(value)
        if cleaned == '':
            return 'NULL'
        return "'" + cleaned.replace("'", "''") + "'"

    column_names = [re.sub(r'[^A-Za-z0-9_]', '_', str(col)).strip('_') or f'column_{idx + 1}' for idx, col in enumerate(df.columns)]

    with open(output_path, 'w', encoding='utf-8') as handle:
        handle.write(f'CREATE TABLE IF NOT EXISTS {safe_table_name} (\n')
        handle.write(',\n'.join([f'    {col} TEXT' for col in column_names]))
        handle.write('\n);\n\n')

        for _, row in df.iterrows():
            values = ', '.join(sql_value(row[col]) for col in df.columns)
            handle.write(f'INSERT INTO {safe_table_name} ({", ".join(column_names)}) VALUES ({values});\n')

def create_sqlite_database(df, table_name, output_path):
    safe_table_name = re.sub(r'[^A-Za-z0-9_]', '_', table_name) or 'imported_data'
    connection = sqlite3.connect(output_path)
    try:
        export_df = df.copy()
        export_df.columns = [
            re.sub(r'[^A-Za-z0-9_]', '_', str(col)).strip('_') or f'column_{idx + 1}'
            for idx, col in enumerate(export_df.columns)
        ]
        export_df.to_sql(safe_table_name, connection, if_exists='replace', index=False)
    finally:
        connection.close()

def process_excel_alignment_tool(file_path, manual_mapping, extra_columns=None, duplicate_fields=None, duplicate_strategy='keep_first'):
    """
    Clean, align, deduplicate, and export a user-uploaded Excel file.
    """
    try:
        df, _ = read_excel_file(file_path)

        required_fields = ['name', 'contact_number']
        missing_fields = [field for field in required_fields if field not in manual_mapping]
        if missing_fields:
            return {
                'success': False,
                'error': f"Missing required mapping: {', '.join(missing_fields)}"
            }

        aligned_df = build_aligned_dataframe(df, manual_mapping, extra_columns)
        data_columns = [col for col in aligned_df.columns if col != 'source_row']
        aligned_df = aligned_df[
            aligned_df[data_columns].apply(lambda row: any(clean_data_value(value) for value in row), axis=1)
        ].copy()

        duplicate_fields = duplicate_fields or ['ic_number', 'contact_number']
        cleaned_df, duplicate_report, duplicates_removed = apply_duplicate_strategy(
            aligned_df,
            duplicate_fields,
            duplicate_strategy
        )

        base_name = os.path.splitext(os.path.basename(file_path))[0]
        token = uuid.uuid4().hex[:12]
        safe_base_name = secure_filename(base_name) or 'aligned_data'
        output_prefix = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_base_name}_{token}"

        aligned_excel_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{output_prefix}_aligned.xlsx')
        duplicates_excel_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{output_prefix}_duplicates.xlsx')
        sql_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{output_prefix}.sql')
        sqlite_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{output_prefix}.db')

        cleaned_df.to_excel(aligned_excel_path, index=False)
        duplicate_report.to_excel(duplicates_excel_path, index=False)
        create_sql_insert_file(cleaned_df, 'aligned_customer_data', sql_path)
        create_sqlite_database(cleaned_df, 'aligned_customer_data', sqlite_path)

        return {
            'success': True,
            'input_rows': int(len(aligned_df)),
            'output_rows': int(len(cleaned_df)),
            'duplicate_rows': int(len(duplicate_report)),
            'duplicates_removed': int(duplicates_removed),
            'preview_rows': cleaned_df.head(8).to_dict('records'),
            'duplicate_preview': duplicate_report.head(8).to_dict('records'),
            'columns': list(cleaned_df.columns),
            'duplicate_columns': list(duplicate_report.columns),
            'downloads': {
                'aligned_excel': os.path.basename(aligned_excel_path),
                'duplicates_excel': os.path.basename(duplicates_excel_path),
                'sql': os.path.basename(sql_path),
                'sqlite': os.path.basename(sqlite_path)
            }
        }
    except Exception as e:
        app.logger.error(f"Excel alignment tool error: {str(e)}")
        return {'success': False, 'error': str(e)}

def process_uploaded_file_with_mapping(file_path, admin_id, manual_mapping=None, extra_columns=None, job_id=None):
    """
    Process file with optional manual column mapping.
    """
    try:
        # Use streaming preview for column detection — avoids loading full file
        df_preview, has_header = read_excel_file(file_path, nrows=50)

        if manual_mapping:
            columns_mapping = manual_mapping
        else:
            columns_mapping = detect_columns_by_content(df_preview)

        # Use preview df for the needs_manual_mapping check
        df = df_preview

        required_fields = ['name', 'contact_number']
        missing_fields = [field for field in required_fields if field not in columns_mapping]

        if missing_fields:
            return {
                'success': False,
                'needs_manual_mapping': True,
                'available_columns': list(df.columns),
                'file_path': file_path,
                'filename': os.path.basename(file_path),
                'detected_mapping': columns_mapping,
                'preview_data': df.head(5).to_dict('records')
            }

        # --- Resolve address/field columns from preview ---
        ADDRESS_PART_PATTERNS = {
            'add 1','add1','add_1','address 1','address1','address_1','alamat 1','alamat1',
            'add 2','add2','add_2','address 2','address2','address_2','alamat 2','alamat2',
            'add 3','add3','add_3','address 3','address3','address_3','alamat 3','alamat3',
            'add 4','add4','add_4',
            'state','negeri','city','bandar','town',
            'poskod','postcode','post code','zip','zip code','postal','postal code',
        }
        col_names = list(df.columns)
        mapped_cols = set(columns_mapping.values())
        addr_part_cols = [c for c in col_names if c not in mapped_cols and c.lower().strip() in ADDRESS_PART_PATTERNS]
        addr_mapped_col = columns_mapping.get('address')
        use_merged_address = bool(addr_part_cols)
        all_mapped = set(columns_mapping.values()) | set(addr_part_cols) | {'_merged_address'}

        ic_col    = columns_mapping.get('ic_number')
        phone_col = columns_mapping.get('contact_number')
        name_col  = columns_mapping.get('name')
        email_col = columns_mapping.get('email')

        import csv as _csv

        def _upd(msg):
            if job_id:
                _set_job(job_id, message=msg)

        # If already CSV (pre-converted by background thread), use directly
        # Otherwise do a quick in-place conversion
        if file_path.lower().endswith('.csv'):
            csv_path = file_path
            # Find header skip by peeking first 15 lines
            _skip = 1
            try:
                with open(csv_path, newline='', encoding='utf-8', errors='replace') as _pf:
                    _pr = _csv.reader(_pf)
                    _peek_rows = [next(_pr) for _ in range(15)]
                _skip = find_best_header_row(pd.DataFrame(_peek_rows)) + 1
            except Exception:
                pass
            csv_written = True
        else:
            # xlsx fallback (should not normally happen now)
            from python_calamine import CalamineWorkbook
            _skip = 1
            try:
                _wb_p = CalamineWorkbook.from_path(file_path)
                _ws_p = _wb_p.get_sheet_by_index(0)
                _pr = []
                for _i, _r in enumerate(_ws_p.iter_rows()):
                    _pr.append(['' if v is None else str(v) for v in _r])
                    if _i >= 15: break
                _skip = find_best_header_row(pd.DataFrame(_pr)) + 1
            except Exception:
                pass
            csv_path = file_path + '_tmp.csv'
            _rows_written = 0
            _cal_wb = CalamineWorkbook.from_path(file_path)
            _cal_ws = _cal_wb.get_sheet_by_index(0)
            with open(csv_path, 'w', newline='', encoding='utf-8') as _cf:
                _writer = _csv.writer(_cf, quoting=_csv.QUOTE_ALL)
                for _i, _r in enumerate(_cal_ws.iter_rows()):
                    if _i < _skip: continue
                    _writer.writerow(['' if v is None else str(v).strip() for v in _r])
                    _rows_written += 1
                    if _rows_written % 100000 == 0:
                        _upd(f'Converting... {_rows_written:,} rows done')
            csv_written = _rows_written > 0

        upload = Upload(
            filename=os.path.basename(file_path),
            file_path=file_path,
            admin_id=admin_id,
            row_count=0,  # updated after processing
            column_count=len(col_names),
            columns_found=json.dumps(columns_mapping),
            column_mapping=json.dumps(columns_mapping),
            year=datetime.now().year,
            status='pending'
        )
        db.session.add(upload)
        db.session.flush()

        # --- Single-pass: stream CSV, dedup within-file, check DB dupes inline ---
        seen_keys   = set()
        insert_batch = []
        dup_batch    = []
        BATCH_SIZE   = 1000
        DB_CHECK_SIZE = 2000   # accumulate this many unique ICs/phones before checking DB
        pending_ic_map   = {}  # ic -> row_data
        pending_ph_map   = {}  # phone -> row_data (only rows not already in ic_map)
        customers_inserted = 0
        customers_updated  = 0
        infile_dupes = 0
        row_count = 0
        errors = []
        row_num = 1

        def _flush_pending_db_check():
            nonlocal customers_inserted, customers_updated
            if not pending_ic_map and not pending_ph_map:
                return
            existing_by_ic = {}
            existing_by_ph = {}
            ic_list = list(pending_ic_map.keys())
            ph_list = list(pending_ph_map.keys())
            if ic_list:
                existing_by_ic = fetch_customers_by_normalized_ics(ic_list)
            if ph_list:
                for rec in CustomerData.query.filter(CustomerData.contact_number.in_(ph_list)).with_entities(CustomerData.id, CustomerData.contact_number).all():
                    existing_by_ph[rec.contact_number] = rec.id

            # IC match = same person. Auto-update address/phone silently. No duplicate flag.
            for ic_norm, rdata in pending_ic_map.items():
                existing = existing_by_ic.get(ic_norm)
                if existing:
                    _merge_payload_into_customer(
                        existing,
                        rdata,
                        source_label=f'upload:{upload.id}:same_ic',
                        merge_reason='upload_same_ic'
                    )
                    customers_updated += 1
                else:
                    insert_batch.append(rdata)
                    customers_inserted += 1

            # Phone-only match (no IC) = flag for review — could be different person
            for ph, rdata in pending_ph_map.items():
                existing_id = existing_by_ph.get(ph)
                if existing_id:
                    dup_batch.append({'upload_id': upload.id, 'existing_customer_id': existing_id,
                                      'duplicate_data': json.dumps(rdata), 'timestamp': datetime.utcnow(), 'action_taken': 'pending'})
                else:
                    insert_batch.append(rdata)
                    customers_inserted += 1
            pending_ic_map.clear()
            pending_ph_map.clear()

            if len(insert_batch) >= BATCH_SIZE:
                db.session.bulk_insert_mappings(CustomerData, insert_batch)
                db.session.flush()
                insert_batch.clear()
            if len(dup_batch) >= BATCH_SIZE:
                db.session.bulk_insert_mappings(DuplicateRecordLog, dup_batch)
                db.session.flush()
                dup_batch.clear()

        _upd('Importing records into database...')
        with open(csv_path, newline='', encoding='utf-8', errors='replace') as _cf:
            _reader = _csv.reader(_cf)
            for csv_row in _reader:
                row_dict = dict(zip(col_names, csv_row + [''] * max(0, len(col_names) - len(csv_row))))

                ic = clean_data_value(row_dict.get(ic_col, ''))    if ic_col    else ''
                ph = clean_data_value(row_dict.get(phone_col, '')) if phone_col else ''
                ic_norm = _normalize_identity_ic(ic)
                ph_norm = _normalize_identity_phone(ph)
                key = ic_norm or ph_norm
                if not key:
                    row_num += 1
                    continue
                if key in seen_keys:
                    infile_dupes += 1
                    row_num += 1
                    continue
                seen_keys.add(key)

                name  = clean_data_value(row_dict.get(name_col, ''))  if name_col  else ''
                email = clean_data_value(row_dict.get(email_col, '')) if email_col else ''

                if use_merged_address:
                    parts = []
                    if addr_mapped_col:
                        parts.append(clean_data_value(row_dict.get(addr_mapped_col, '')))
                    for ac in addr_part_cols:
                        if ac != addr_mapped_col:
                            parts.append(clean_data_value(row_dict.get(ac, '')))
                    address = ', '.join(p for p in parts if p)
                else:
                    address = clean_data_value(row_dict.get(addr_mapped_col, '')) if addr_mapped_col else ''

                if extra_columns is not None:
                    additional = {extra_columns[c]: clean_data_value(row_dict.get(c, '')) for c in col_names if c in extra_columns and c not in all_mapped}
                else:
                    additional = {c: clean_data_value(row_dict.get(c, '')) for c in col_names if c not in all_mapped}

                rdata = {'name': name, 'contact_number': ph, 'ic_number': ic,
                         'address': address, 'email': email,
                         'additional_data': json.dumps(additional), 'upload_id': upload.id}

                if ic_norm:
                    pending_ic_map[ic_norm] = rdata
                else:
                    pending_ph_map[ph_norm] = rdata

                row_count += 1
                row_num += 1

                if (len(pending_ic_map) + len(pending_ph_map)) >= DB_CHECK_SIZE:
                    _flush_pending_db_check()
                    _upd(f'Importing... {row_count:,} rows processed, {customers_inserted:,} inserted')

        # Final flush
        _flush_pending_db_check()
        if insert_batch:
            db.session.bulk_insert_mappings(CustomerData, insert_batch)
        if dup_batch:
            db.session.bulk_insert_mappings(DuplicateRecordLog, dup_batch)

        # Cleanup temp CSV
        try:
            os.remove(csv_path)
        except Exception:
            pass

        upload.row_count = row_count
        dup_count = DuplicateRecordLog.query.filter_by(upload_id=upload.id, action_taken='pending').count()
        upload.status = 'pending' if dup_count > 0 else 'processed'
        db.session.commit()

        if dup_count > 0:
            return {
                'success': False,
                'needs_review': True,
                'duplicate_count': dup_count,
                'upload_id': upload.id,
                'message': f"Found {dup_count} duplicate records. Please review."
            }

        msg = f"Successfully inserted {customers_inserted} new records."
        if customers_updated:
            msg += f" {customers_updated} existing records updated (same IC, new address/phone)."
        if infile_dupes:
            msg += f" {infile_dupes} within-file duplicates skipped."
        return {
            'success': True,
            'records_inserted': customers_inserted,
            'records_updated': customers_updated,
            'infile_duplicates_skipped': infile_dupes,
            'errors': errors,
            'upload_id': upload.id,
            'message': msg
        }

    except Exception as e:
        app.logger.error(f"Error processing file: {str(e)}")
        db.session.rollback()
        try:
            if 'csv_path' in dir() and os.path.exists(csv_path):
                os.remove(csv_path)
        except Exception:
            pass
        return {'success': False, 'error': str(e)}


def sync_file_with_mapping(file_path, admin_id, manual_mapping, extra_columns=None):
    """
    Sync a new file against existing DB records.
    """
    try:
        df, _ = read_excel_file(file_path)

        updated = 0
        inserted = 0
        unchanged = 0
        errors = []

        upload = Upload(
            filename=os.path.basename(file_path),
            file_path=file_path,
            admin_id=admin_id,
            row_count=len(df),
            column_count=len(df.columns),
            columns_found=json.dumps(manual_mapping),
            column_mapping=json.dumps(manual_mapping),
            year=datetime.now().year,
            status='synced'
        )
        db.session.add(upload)
        db.session.flush()

        for idx, row in df.iterrows():
            try:
                new_name    = clean_data_value(row[manual_mapping['name']]) if 'name' in manual_mapping else ''
                new_contact = clean_data_value(row[manual_mapping['contact_number']]) if 'contact_number' in manual_mapping else ''
                new_ic      = clean_data_value(row[manual_mapping['ic_number']]) if 'ic_number' in manual_mapping else ''
                new_address = clean_data_value(row[manual_mapping['address']]) if 'address' in manual_mapping else ''
                new_email   = clean_data_value(row[manual_mapping['email']]) if 'email' in manual_mapping else ''

                if not new_name and not new_ic and not new_contact:
                    continue

                if extra_columns is not None:
                    new_extra = {label: clean_data_value(row[col]) for col, label in extra_columns.items() if col in df.columns}
                else:
                    new_extra = {col: clean_data_value(row[col]) for col in df.columns if col not in manual_mapping.values()}

                existing = None
                if new_ic:
                    existing = find_customer_by_normalized_ic(new_ic)
                if not existing and new_contact:
                    existing = CustomerData.query.filter_by(contact_number=new_contact).first()

                if existing:
                    before = {
                        'name': existing.name,
                        'contact_number': existing.contact_number,
                        'ic_number': existing.ic_number,
                        'address': existing.address,
                        'email': existing.email,
                        'additional_data': existing.additional_data,
                    }
                    payload = {
                        'name': new_name,
                        'contact_number': new_contact,
                        'ic_number': new_ic,
                        'address': new_address,
                        'email': new_email,
                        'additional_data': json.dumps(new_extra) if new_extra else '',
                    }
                    _merge_payload_into_customer(
                        existing,
                        payload,
                        source_label=f'manual_merge_upload:{upload.id}',
                        merge_reason='manual_merge_existing'
                    )
                    changed = any([
                        before['name'] != existing.name,
                        before['contact_number'] != existing.contact_number,
                        before['ic_number'] != existing.ic_number,
                        before['address'] != existing.address,
                        before['email'] != existing.email,
                        before['additional_data'] != existing.additional_data,
                    ])
                    if changed:
                        existing.updated_at = datetime.utcnow()
                        updated += 1
                    else:
                        unchanged += 1
                else:
                    customer = CustomerData(
                        name=new_name,
                        contact_number=new_contact,
                        ic_number=new_ic,
                        address=new_address,
                        email=new_email,
                        additional_data=json.dumps(new_extra),
                        upload_id=upload.id
                    )
                    db.session.add(customer)
                    inserted += 1

            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

        db.session.commit()
        return {
            'success': True,
            'updated': updated,
            'inserted': inserted,
            'unchanged': unchanged,
            'errors': errors,
            'upload_id': upload.id
        }

    except Exception as e:
        app.logger.error(f"Sync error: {str(e)}")
        db.session.rollback()
        return {'success': False, 'error': str(e)}

# ==================== DECORATORS ====================

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Admin access required', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ==================== SECURITY HELPERS ====================

def is_system_locked():
    """Return True if system lockout is active."""
    return SystemSettings.get('system_locked', '0') == '1'

# ==================== ROUTES ====================

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

@app.route('/favicon.ico')
def favicon():
    raise NotFound()

@app.after_request
def add_no_cache_headers(response):
    if 'text/html' in response.content_type:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
    return response

@app.before_request
def enforce_password_policy():
    open_routes = {'login', 'logout', 'forgot_password', 'force_change_password', 'static',
                   'waiting_approval', 'check_approval', 'telegram_webhook', 'system_locked_page', 'health'}

    # Session inactivity timeout
    if current_user.is_authenticated and request.endpoint not in open_routes:
        timeout_minutes = app.config.get('SESSION_TIMEOUT_MINUTES', 30)
        last_activity = session.get('last_activity')
        if last_activity:
            try:
                idle_seconds = (datetime.utcnow() - datetime.fromisoformat(last_activity)).total_seconds()
                if idle_seconds > timeout_minutes * 60:
                    logout_user()
                    session.clear()
                    flash(f'Your session expired after {timeout_minutes} minutes of inactivity. Please log in again.', 'warning')
                    return redirect(url_for('login'))
            except Exception:
                pass
        session['last_activity'] = datetime.utcnow().isoformat()
        session.permanent = True   # sliding window: every request pushes expiry forward
        session.modified = True

    if request.endpoint not in open_routes:
        if SystemSettings.get('system_locked', '0') == '1':
            if not current_user.is_authenticated:
                return redirect(url_for('system_locked_page'))
            if request.endpoint != 'unlock_system' and current_user.role != 'admin':
                return redirect(url_for('system_locked_page'))

    if current_user.is_authenticated and request.endpoint not in open_routes:
        if session.get('force_pw_change') or current_user.is_password_expired() or current_user.must_change_password:
            session['force_pw_change'] = True
            return redirect(url_for('force_change_password'))

@app.context_processor
def inject_globals():
    try:
        from flask_login import current_user
        if current_user.is_authenticated and current_user.role == 'admin':
            count = AgentDevice.query.filter_by(status='pending').count()
            phase2 = SystemSettings.get('phase2_enabled', '0') == '1'
            panic = dp.is_panic()
            return {'pending_devices_count': count, 'phase2_enabled': phase2, 'panic_enabled': panic}
    except:
        pass
    return {'pending_devices_count': 0, 'phase2_enabled': False, 'panic_enabled': False}

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if is_system_locked():
        session['lock_page_reason'] = 'System locked due to security review.'
        flash('System is temporarily locked due to security review. Please contact the administrator.', 'danger')
        return redirect(url_for('system_locked_page'))
    
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('agent_dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        ip_address = get_client_ip()
        user_agent = request.headers.get('User-Agent', '')

        if is_ip_blocked(ip_address):
            session['lock_page_reason'] = 'Access denied: this IP address has been blocked by an administrator.'
            log_security_event('login_denied', 'blocked', ip_address, username=username,
                               details='Blocked IP attempted to access login page.')
            flash('Access denied: your IP address has been blocked.', 'danger')
            return redirect(url_for('system_locked_page'))

        user = User.query.filter_by(username=username).first()

        if user and user.locked_until:
            if datetime.utcnow() < user.locked_until:
                remaining = int((user.locked_until - datetime.utcnow()).total_seconds() / 60) + 1
                log_security_event('login_denied', 'locked', ip_address, user=user,
                                   details=f'Login blocked while account lockout remained active for ~{remaining} minute(s).')
                flash(f'Account temporarily locked due to too many failed attempts. Try again in {remaining} minute(s).', 'danger')
                return redirect(url_for('login'))
            else:
                user.failed_login_attempts = 0
                user.locked_until = None
                db.session.commit()

        if user and user.check_password(password) and user.is_active:
            user.failed_login_attempts = 0
            user.locked_until = None
            connection_type = detect_connection_type(user_agent)
            session_id = str(uuid.uuid4())
            fingerprint = request.form.get('fp', '')
            device_category = detect_device_category(user_agent)

            if user.role == 'admin' and fingerprint and is_admin_security_enabled():
                admin_status = check_admin_device(user.id, fingerprint, ip_address)
                session['admin_device_status'] = admin_status
                
                if admin_status == 'blocked':
                    log_security_event('login_denied', 'blocked', ip_address, user=user,
                                       details='Admin login denied because this device is blocked.')
                    flash('Access denied: this device is blocked.', 'danger')
                    return redirect(url_for('login'))
                
                if admin_status in ('new', 'pending'):
                    pending = create_pending_login_approval(
                        user.id, fingerprint, ip_address, user_agent
                    )
                    send_telegram_notification(pending)
                    session['pending_approval_token'] = pending.approval_token
                    return redirect(url_for('waiting_approval'))
            
            login_log = LoginLog(
                user_id=user.id,
                ip_address=ip_address,
                location=get_location_from_ip(ip_address),
                device_info=user_agent,
                fingerprint=fingerprint,
                device_category=device_category,
                connection_type=connection_type,
                session_id=session_id
            )
            db.session.add(login_log)
            user.last_login = datetime.utcnow()
            db.session.commit()

            analyze_login_behavior(user, login_log)

            login_user(user)
            session.permanent = True
            session['login_log_id'] = login_log.id
            session['session_id'] = session_id

            if user.role == 'agent' and fingerprint:
                agent_device_status = check_device_trusted(user.id, fingerprint, ip_address)
                session['agent_device_status'] = agent_device_status
                app.logger.info(f"[LOGIN] user={user.username} fp={fingerprint[:8]} status={agent_device_status} tg={user.telegram_chat_id!r}")
                if agent_device_status == 'device_limit':
                    logout_user()
                    max_d = _max_agent_devices()
                    log_security_event('login_denied', 'device_limit', ip_address, user=user,
                                       details=f'Device limit of {max_d} reached.')
                    flash(f'Login denied: you have reached the maximum of {max_d} registered devices. '
                          f'Please contact your admin to remove an old device.', 'danger')
                    return redirect(url_for('login'))
                elif agent_device_status == 'ip_changed':
                    device = AgentDevice.query.filter_by(user_id=user.id, fingerprint=fingerprint).first()
                    if device:
                        if user.telegram_chat_id:
                            send_agent_ip_change_confirmation(user, device, ip_address)
                            session['pending_device_fp'] = fingerprint
                            return redirect(url_for('agent_device_waiting'))
                        else:
                            _notify_admin_agent_ip_change(user, device, ip_address)
                            session['pending_device_fp'] = fingerprint
                            return redirect(url_for('agent_device_pending'))
                elif agent_device_status == 'new':
                    device = AgentDevice.query.filter_by(user_id=user.id, fingerprint=fingerprint).first()
                    if device:
                        if user.telegram_chat_id:
                            send_agent_device_confirmation(user, device)
                            session['pending_device_fp'] = fingerprint
                            return redirect(url_for('agent_device_waiting'))
                        else:
                            _notify_admin_new_agent_device(user, device, ip_address)
                            session['pending_device_fp'] = fingerprint
                            return redirect(url_for('agent_device_pending'))
                elif agent_device_status in ('pending', 'blocked'):
                    return redirect(url_for('agent_device_pending'))

            if user.is_password_expired() or user.must_change_password:
                session['force_pw_change'] = True
                flash('Your password has expired. Please set a new password.', 'warning')
                return redirect(url_for('force_change_password'))

            if user.role == 'admin' and fingerprint and not is_admin_security_enabled():
                admin_status = check_admin_device(user.id, fingerprint, ip_address)
                session['admin_device_status'] = admin_status
                if admin_status in ('new', 'pending'):
                    flash(f'Warning: Unrecognised device detected from {ip_address}. This login has been flagged.', 'warning')
                elif admin_status == 'blocked':
                    logout_user()
                    log_security_event('login_denied', 'blocked', ip_address, user=user,
                                       details='Admin session rejected because the device is blocked.')
                    flash('Access denied: this device is blocked.', 'danger')
                    return redirect(url_for('login'))

            flash(f'Welcome back, {user.full_name}!', 'success')

            if user.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('agent_dashboard'))
        else:
            flash('Invalid username or password', 'danger')
            if user:
                handle_failed_login(user, ip_address)
            else:
                log_security_event('login_denied', 'failed', ip_address, username=username,
                                   details='Unknown username or invalid password submitted.')

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    if 'login_log_id' in session:
        login_log = LoginLog.query.get(session['login_log_id'])
        if login_log:
            login_log.logout_time = datetime.utcnow()
            if login_log.login_time:
                duration = (login_log.logout_time - login_log.login_time).total_seconds()
                login_log.session_duration = int(duration)
            db.session.commit()
    
    logout_user()
    session.clear()
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        user = User.query.filter_by(username=username, role='agent').first()

        success_msg = 'If this account has a linked Telegram, a temporary password has been sent.'

        if user and user.telegram_chat_id:
            token = SystemSettings.get('telegram_bot_token')
            if token:
                max_resets = int(SystemSettings.get('pw_reset_max', '3'))
                window_days = int(SystemSettings.get('pw_reset_window_days', '7'))
                now = datetime.utcnow()

                # Reset counter if window has expired
                if user.pw_reset_window_start:
                    elapsed = (now - user.pw_reset_window_start).days
                    if elapsed >= window_days:
                        user.pw_reset_count = 0
                        user.pw_reset_window_start = None

                # Check limit
                if (user.pw_reset_count or 0) >= max_resets:
                    flash(f'Password reset limit reached ({max_resets} resets per {window_days} days). Please contact your admin.', 'danger')
                    return redirect(url_for('forgot_password'))

                import random, string
                temp_pw = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
                user.set_password(temp_pw)
                user.must_change_password = True

                # Track reset count
                if not user.pw_reset_window_start:
                    user.pw_reset_window_start = now
                user.pw_reset_count = (user.pw_reset_count or 0) + 1
                db.session.commit()

                resets_left = max_resets - user.pw_reset_count
                try:
                    http_req.post(
                        f"https://api.telegram.org/bot{token}/sendMessage",
                        json={
                            "chat_id": user.telegram_chat_id,
                            "text": (
                                f"🔐 *Password Reset*\n\n"
                                f"Hello {user.full_name}, your temporary password is:\n\n"
                                f"`{temp_pw}`\n\n"
                                f"Please log in and change your password immediately. "
                                f"This password is single-use.\n\n"
                                f"_You have {resets_left} reset(s) remaining in this {window_days}-day window._"
                            ),
                            "parse_mode": "Markdown"
                        },
                        timeout=10
                    )
                except Exception as e:
                    app.logger.error(f"[FORGOT-PW] Telegram send error: {e}")

        flash(success_msg, 'info')
        return redirect(url_for('login'))

    return render_template('forgot_password.html')

# ==================== ADMIN ROUTES ====================

@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    total_agents = User.query.filter_by(role='agent').count()
    total_customers = _cached_count('total_customers', _fast_customer_count)
    total_uploads = Upload.query.count()
    recent_uploads = Upload.query.order_by(Upload.upload_date.desc()).limit(10).all()
    recent_logins = LoginLog.query.order_by(LoginLog.login_time.desc()).limit(20).all()
    pending_duplicates = DuplicateRecordLog.query.filter_by(action_taken='pending').count()

    # Uploads that still have unresolved duplicates
    from sqlalchemy import func
    pending_dup_uploads = (
        db.session.query(Upload, func.count(DuplicateRecordLog.id).label('dup_count'))
        .join(DuplicateRecordLog, DuplicateRecordLog.upload_id == Upload.id)
        .filter(DuplicateRecordLog.action_taken == 'pending')
        .group_by(Upload.id)
        .order_by(Upload.upload_date.desc())
        .all()
    )

    recent_mys_downloads = (MysDownloadLog.query
                            .order_by(MysDownloadLog.timestamp.desc())
                            .limit(10).all())

    pending_ip_devices = (AgentDevice.query
                          .filter_by(status='pending')
                          .order_by(AgentDevice.last_seen.desc())
                          .all())

    return render_template('admin_dashboard.html',
                         total_agents=total_agents,
                         total_customers=total_customers,
                         total_uploads=total_uploads,
                         recent_uploads=recent_uploads,
                         recent_logins=recent_logins,
                         pending_duplicates=pending_duplicates,
                         pending_dup_uploads=pending_dup_uploads,
                         recent_mys_downloads=recent_mys_downloads,
                         pending_ip_devices=pending_ip_devices)

@app.route('/admin/upload', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_upload():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(request.url)
        
        if file and file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{filename}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)

            try:
                df, _ = read_excel_file(file_path, nrows=30)  # preview only — fast
                detected = detect_columns_by_content(df)
                preview = []
                for _, row in df.head(5).iterrows():
                    preview.append({col: str(row[col]) if not pd.isna(row[col]) else '' for col in df.columns})
                session['pending_file'] = {
                    'file_path': file_path,
                    'filename': filename,
                    'available_columns': list(df.columns),
                    'detected_mapping': detected,
                    'preview_data': preview
                }
                return redirect(url_for('manual_column_mapping'))
            except Exception as e:
                flash(f"Error reading file: {str(e)}", 'danger')
                return redirect(url_for('admin_upload'))
        else:
            flash('Please upload an Excel (.xlsx/.xls) or CSV (.csv) file', 'danger')
            return redirect(request.url)
    
    return render_template('admin_upload.html')


@app.route('/admin/process-server-file', methods=['GET', 'POST'])
@login_required
@admin_required
def process_server_file():
    """Process a large file already SCP'd to the server — bypasses HTTP upload limit."""
    ALLOWED_DIR = os.path.abspath(app.config['UPLOAD_FOLDER'])

    if request.method == 'POST':
        raw_path = request.form.get('file_path', '').strip()
        if not raw_path:
            flash('Please enter a file path.', 'danger')
            return redirect(request.url)

        # Resolve to absolute path, restrict to uploads folder or /root
        import pathlib
        p = pathlib.Path(raw_path)
        if not p.is_absolute():
            p = pathlib.Path('/root') / p

        file_path = str(p.resolve())

        # Security: only allow files in /root or the uploads folder
        allowed_roots = ['/root', str(pathlib.Path(ALLOWED_DIR).resolve())]
        if not any(file_path.startswith(r) for r in allowed_roots):
            flash('Access denied: file must be in /root or the uploads folder.', 'danger')
            return redirect(request.url)

        if not os.path.isfile(file_path):
            flash(f'File not found: {file_path}', 'danger')
            return redirect(request.url)

        if not file_path.lower().endswith(('.xlsx', '.xls', '.csv')):
            flash('Only .xlsx, .xls, or .csv files are supported.', 'danger')
            return redirect(request.url)

        try:
            df, _ = read_excel_file(file_path, nrows=30)
            detected = detect_columns_by_content(df)
            preview = []
            for _, row in df.head(5).iterrows():
                preview.append({col: str(row[col]) if not pd.isna(row[col]) else '' for col in df.columns})
            session['pending_file'] = {
                'file_path': file_path,
                'filename': os.path.basename(file_path),
                'available_columns': list(df.columns),
                'detected_mapping': detected,
                'preview_data': preview
            }
            return redirect(url_for('manual_column_mapping'))
        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'danger')
            return redirect(request.url)

    # List files already in the uploads folder — newest first
    try:
        existing_files = []
        for f in os.listdir(ALLOWED_DIR):
            if f.lower().endswith(('.xlsx', '.xls', '.csv')):
                full = os.path.join(ALLOWED_DIR, f)
                existing_files.append({
                    'name': f,
                    'path': full,
                    'size_mb': round(os.path.getsize(full) / 1024 / 1024, 1),
                    'mtime': os.path.getmtime(full)
                })
        existing_files.sort(key=lambda x: x['mtime'], reverse=True)  # newest first
    except Exception:
        existing_files = []

    # Also check /root for any files dropped there via SCP
    try:
        for f in os.listdir('/root'):
            if f.lower().endswith(('.xlsx', '.xls', '.csv')):
                full = f'/root/{f}'
                existing_files.append({
                    'name': f,
                    'path': full,
                    'size_mb': round(os.path.getsize(full) / 1024 / 1024, 1),
                    'mtime': os.path.getmtime(full)
                })
        existing_files.sort(key=lambda x: x['mtime'], reverse=True)
    except Exception:
        pass

    return render_template('process_server_file.html', existing_files=existing_files)


@app.route('/admin/excel-tools', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_excel_tools():
    result = None

    if request.method == 'POST':
        action = request.form.get('action', 'upload')

        if action == 'reset':
            session.pop('excel_tool_pending', None)
            flash('Excel tool form has been reset.', 'info')
            return redirect(url_for('admin_excel_tools'))

        if action == 'upload':
            if 'file' not in request.files or request.files['file'].filename == '':
                flash('No file selected', 'danger')
                return redirect(request.url)

            file = request.files['file']
            if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
                flash('Please upload an Excel (.xlsx/.xls) or CSV (.csv) file', 'danger')
                return redirect(request.url)

            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            stored_filename = f"tool_{timestamp}_{filename}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], stored_filename)
            file.save(file_path)

            try:
                df, _ = read_excel_file(file_path)
                detected = detect_columns_by_content(df)
                preview = []
                for _, row in df.head(6).iterrows():
                    preview.append({col: clean_data_value(row[col]) for col in df.columns})

                session['excel_tool_pending'] = {
                    'file_path': file_path,
                    'filename': filename,
                    'available_columns': list(df.columns),
                    'detected_mapping': detected,
                    'preview_data': preview
                }
                flash('File uploaded. Review the mapping and choose how to handle duplicates.', 'info')
                return redirect(url_for('admin_excel_tools'))
            except Exception as e:
                flash(f'Error reading file: {str(e)}', 'danger')
                return redirect(url_for('admin_excel_tools'))

        if action == 'process':
            pending = session.get('excel_tool_pending')
            if not pending:
                flash('Please upload a file first.', 'danger')
                return redirect(url_for('admin_excel_tools'))

            manual_mapping = {
                'name': request.form.get('name_column'),
                'contact_number': request.form.get('contact_column'),
                'ic_number': request.form.get('ic_column'),
                'address': request.form.get('address_column'),
                'email': request.form.get('email_column')
            }
            manual_mapping = {k: v for k, v in manual_mapping.items() if v}

            mapped_cols = set(manual_mapping.values())
            extra_columns = {}
            for col in pending.get('available_columns', []):
                if col in mapped_cols:
                    continue
                col_key = col.replace(' ', '_').replace('.', '_')
                if request.form.get(f'extra_include_{col_key}'):
                    label = request.form.get(f'extra_label_{col_key}', '').strip()
                    extra_columns[col] = label if label else col

            duplicate_fields = request.form.getlist('duplicate_keys')
            if not duplicate_fields:
                duplicate_fields = ['ic_number', 'contact_number']

            duplicate_strategy = request.form.get('duplicate_strategy', 'keep_first')
            result = process_excel_alignment_tool(
                pending['file_path'],
                manual_mapping,
                extra_columns,
                duplicate_fields,
                duplicate_strategy
            )

            if result.get('success'):
                flash('Aligned files were generated successfully.', 'success')
            else:
                flash(f"Processing failed: {result.get('error', 'Unknown error')}", 'danger')

    pending = session.get('excel_tool_pending')
    if pending:
        detected = pending.get('detected_mapping', {})
        mapped_cols = set(detected.values())
        extra_cols = [col for col in pending['available_columns'] if col not in mapped_cols]
    else:
        detected = {}
        extra_cols = []

    return render_template(
        'admin_excel_tools.html',
        pending=pending,
        detected=detected,
        extra_cols=extra_cols,
        result=result
    )

@app.route('/admin/excel-tools/download/<path:filename>')
@login_required
@admin_required
def download_excel_tool_file(filename):
    safe_filename = secure_filename(filename)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
    if not os.path.exists(file_path):
        flash('Requested file was not found.', 'danger')
        return redirect(url_for('admin_excel_tools'))
    return send_file(file_path, as_attachment=True, download_name=safe_filename)

@app.route('/admin/sync', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_sync():
    if request.method == 'POST':
        if 'file' not in request.files or request.files['file'].filename == '':
            flash('No file selected', 'danger')
            return redirect(request.url)

        file = request.files['file']
        if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
            flash('Please upload an Excel (.xlsx/.xls) or CSV (.csv) file', 'danger')
            return redirect(request.url)

        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"sync_{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        try:
            df, _ = read_excel_file(file_path)
            detected = detect_columns_by_content(df)
            preview = []
            for _, row in df.head(5).iterrows():
                preview.append({col: str(row[col]) if not pd.isna(row[col]) else '' for col in df.columns})
            session['pending_file'] = {
                'file_path': file_path,
                'filename': filename,
                'available_columns': list(df.columns),
                'detected_mapping': detected,
                'preview_data': preview,
                'is_sync': True
            }
            return redirect(url_for('manual_column_mapping'))
        except Exception as e:
            flash(f"Error reading file: {str(e)}", 'danger')
            return redirect(url_for('admin_sync'))

    recent_syncs = Upload.query.filter_by(status='synced').order_by(Upload.upload_date.desc()).limit(10).all()
    return render_template('admin_sync.html', recent_syncs=recent_syncs)

@app.route('/admin/manual-mapping', methods=['GET', 'POST'])
@login_required
@admin_required
def manual_column_mapping():
    if 'pending_file' not in session:
        flash('No file pending for mapping', 'danger')
        return redirect(url_for('admin_upload'))
    
    pending = session['pending_file']
    
    if request.method == 'POST':
        manual_mapping = {
            'name': request.form.get('name_column'),
            'contact_number': request.form.get('contact_column'),
            'ic_number': request.form.get('ic_column'),
            'address': request.form.get('address_column'),
            'email': request.form.get('email_column')
        }
        manual_mapping = {k: v for k, v in manual_mapping.items() if v}

        mapped_cols = set(manual_mapping.values())
        all_cols = pending.get('available_columns', [])
        extra_columns = {}
        for col in all_cols:
            if col in mapped_cols:
                continue
            col_key = col.replace(' ', '_').replace('.', '_')
            if request.form.get(f'extra_include_{col_key}'):
                label = request.form.get(f'extra_label_{col_key}', '').strip()
                extra_columns[col] = label if label else col

        is_sync = pending.get('is_sync', False)
        session.pop('pending_file', None)

        if is_sync:
            result = sync_file_with_mapping(pending['file_path'], current_user.id, manual_mapping, extra_columns)
            if result.get('success'):
                flash(f"Sync complete — Updated: {result['updated']}, New: {result['inserted']}, Unchanged: {result['unchanged']}", 'success')
                if result.get('errors'):
                    flash(f"Warnings: {', '.join(result['errors'][:5])}", 'warning')
            else:
                flash(f"Sync failed: {result.get('error', 'Unknown error')}", 'danger')
            return redirect(url_for('admin_dashboard'))

        # Run processing in background thread so the browser doesn't time out
        job_id = str(uuid.uuid4())
        _set_job(job_id, status='running', message='Starting...', progress=0, result=None)
        file_path = pending['file_path']
        admin_id  = current_user.id

        def _bg_process():
            with app.app_context():
                try:
                    # Step 1: convert xlsx → CSV inside background thread
                    actual_file = file_path
                    if file_path.lower().endswith(('.xlsx', '.xls')):
                        import csv as _csv_pre
                        from python_calamine import CalamineWorkbook
                        csv_pre = file_path + '_pre.csv'
                        _set_job(job_id, status='running', message='Converting Excel to CSV...')
                        _wb_pre = CalamineWorkbook.from_path(file_path)
                        _ws_pre = _wb_pre.get_sheet_by_index(0)
                        _n = 0
                        with open(csv_pre, 'w', newline='', encoding='utf-8') as _f:
                            _w = _csv_pre.writer(_f, quoting=_csv_pre.QUOTE_ALL)
                            for _r in _ws_pre.iter_rows():
                                _w.writerow(['' if v is None else str(v).strip() for v in _r])
                                _n += 1
                                if _n % 100000 == 0:
                                    _set_job(job_id, message=f'Converting... {_n:,} rows done')
                        actual_file = csv_pre

                    # Step 2: process CSV
                    _set_job(job_id, status='running', message='Importing records into database...')
                    result = process_uploaded_file_with_mapping(actual_file, admin_id, manual_mapping, extra_columns, job_id=job_id)
                    _set_job(job_id, status='done', result=result,
                             message=result.get('message', 'Done'))
                    # Cleanup pre-converted CSV
                    try:
                        if actual_file != file_path and os.path.exists(actual_file):
                            os.remove(actual_file)
                    except Exception:
                        pass
                except Exception as e:
                    app.logger.error(f'Background upload error: {e}')
                    _set_job(job_id, status='error', message=str(e), result={'error': str(e)})

        t = threading.Thread(target=_bg_process, daemon=True)
        t.start()
        return redirect(url_for('upload_progress', job_id=job_id))
    
    detected = pending.get('detected_mapping', {})
    mapped_cols = set(detected.values())
    extra_cols = [col for col in pending['available_columns'] if col not in mapped_cols]
    return render_template('manual_mapping.html',
                         filename=pending['filename'],
                         columns=pending['available_columns'],
                         detected=detected,
                         extra_cols=extra_cols,
                         preview=pending['preview_data'])

@app.route('/admin/upload-progress/<job_id>')
@login_required
@admin_required
def upload_progress(job_id):
    return render_template('upload_progress.html', job_id=job_id)

@app.route('/admin/upload-progress/<job_id>/status')
@login_required
@admin_required
def upload_progress_status(job_id):
    job = _get_job(job_id)
    if not job:
        return jsonify({'status': 'error', 'message': 'Job not found'})
    return jsonify(job)

@app.route('/admin/duplicates/<int:upload_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def resolve_duplicates(upload_id):
    upload = Upload.query.get_or_404(upload_id)
    pending_total = DuplicateRecordLog.query.filter_by(upload_id=upload_id, action_taken='pending').count()

    if request.method == 'POST':
        bulk_action = request.form.get('bulk_action')
        now = datetime.utcnow()

        if bulk_action == 'skip_all':
            DuplicateRecordLog.query.filter_by(upload_id=upload_id, action_taken='pending').update(
                {'action_taken': 'skipped', 'resolution_time': now}
            )
            db.session.commit()
            upload.status = 'processed'
            db.session.commit()
            flash(f'All {pending_total} duplicates marked as skipped (kept existing records).', 'success')
            return redirect(url_for('admin_dashboard'))

        elif bulk_action == 'accept_all':
            # Batch-load all pending dups and their existing customers
            CHUNK = 500
            offset = 0
            while True:
                dups = DuplicateRecordLog.query.filter_by(upload_id=upload_id, action_taken='pending').offset(offset).limit(CHUNK).all()
                if not dups:
                    break
                customer_ids = [d.existing_customer_id for d in dups]
                customers = {c.id: c for c in CustomerData.query.filter(CustomerData.id.in_(customer_ids)).all()}
                for dup in dups:
                    customer = customers.get(dup.existing_customer_id)
                    if customer:
                        new_data = json.loads(dup.duplicate_data)
                        for field in ['name', 'contact_number', 'ic_number', 'address', 'email']:
                            val = new_data.get(field, '')
                            if val and val not in ('nan', 'None', ''):
                                setattr(customer, field, val)
                        customer.updated_at = now
                    dup.action_taken = 'updated'
                    dup.resolution_time = now
                db.session.flush()
                offset += CHUNK
            db.session.commit()
            upload.status = 'processed'
            db.session.commit()
            flash(f'All {pending_total} duplicates accepted — existing records updated with new data.', 'success')
            return redirect(url_for('admin_dashboard'))

        elif bulk_action == 'keep_both_all':
            # Insert all new records as separate entries tagged [PHONE-DUP] for later review
            CHUNK = 500
            offset = 0
            inserted = 0
            while True:
                dups = DuplicateRecordLog.query.filter_by(upload_id=upload_id, action_taken='pending').offset(offset).limit(CHUNK).all()
                if not dups:
                    break
                insert_batch = []
                for dup in dups:
                    new_data = json.loads(dup.duplicate_data)
                    # Tag the name so it's easy to find and review later
                    new_data['name'] = ('[PHONE-DUP] ' + (new_data.get('name') or '')).strip()
                    new_data['upload_id'] = upload_id
                    insert_batch.append(new_data)
                    dup.action_taken = 'kept_both'
                    dup.resolution_time = now
                db.session.bulk_insert_mappings(CustomerData, insert_batch)
                inserted += len(insert_batch)
                db.session.flush()
                offset += CHUNK
            db.session.commit()
            upload.status = 'processed'
            db.session.commit()
            flash(f'Kept both — {inserted} new records inserted and tagged [PHONE-DUP] for later review.', 'success')
            return redirect(url_for('admin_dashboard'))

        else:
            # Individual page form submit — resolve only what's on this page
            page = request.form.get('page', 1, type=int)
            dups_on_page = DuplicateRecordLog.query.filter_by(upload_id=upload_id, action_taken='pending').paginate(page=page, per_page=50, error_out=False).items
            customer_ids = [d.existing_customer_id for d in dups_on_page]
            customers = {c.id: c for c in CustomerData.query.filter(CustomerData.id.in_(customer_ids)).all()}
            now = datetime.utcnow()
            for dup in dups_on_page:
                action = request.form.get(f'duplicate_{dup.id}')
                customer = customers.get(dup.existing_customer_id)
                if action == 'update' and customer:
                    new_data = json.loads(dup.duplicate_data)
                    _merge_payload_into_customer(
                        customer,
                        new_data,
                        source_label=f'duplicate_log:{dup.id}:update',
                        merge_reason='duplicate_log_update'
                    )
                    dup.action_taken = 'updated'
                    dup.resolution_time = now
                elif action == 'merge' and customer:
                    new_data = json.loads(dup.duplicate_data)
                    _merge_payload_into_customer(
                        customer,
                        new_data,
                        source_label=f'duplicate_log:{dup.id}:merge',
                        merge_reason='duplicate_log_merge'
                    )
                    dup.action_taken = 'merged'
                    dup.resolution_time = now
                elif action == 'skip':
                    dup.action_taken = 'skipped'
                    dup.resolution_time = now
            db.session.commit()
            remaining = DuplicateRecordLog.query.filter_by(upload_id=upload_id, action_taken='pending').count()
            if remaining == 0:
                upload.status = 'processed'
                db.session.commit()
                flash('All duplicates resolved.', 'success')
                return redirect(url_for('admin_dashboard'))
            flash(f'Page resolved. {remaining} duplicates still pending.', 'info')
            return redirect(url_for('resolve_duplicates', upload_id=upload_id))

    # GET — paginated load, batch fetch customers
    page = request.args.get('page', 1, type=int)
    pagination = DuplicateRecordLog.query.filter_by(upload_id=upload_id, action_taken='pending').paginate(page=page, per_page=50, error_out=False)
    dups = pagination.items
    customer_ids = [d.existing_customer_id for d in dups]
    customers = {c.id: c for c in CustomerData.query.filter(CustomerData.id.in_(customer_ids)).all()}
    for dup in dups:
        dup.customer = customers.get(dup.existing_customer_id)
        dup.duplicate_data_parsed = json.loads(dup.duplicate_data)

    return render_template('resolve_duplicates.html', duplicates=dups, upload=upload,
                           pagination=pagination, pending_total=pending_total)

@app.route('/admin/agents')
@login_required
@admin_required
def admin_agents():
    agents = User.query.filter_by(role='agent').all()
    global_device_limit = _max_agent_devices()
    return render_template('admin_agents.html', agents=agents, global_device_limit=global_device_limit)

@app.route('/admin/agent/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_agent():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        full_name = request.form.get('full_name')
        phone = request.form.get('phone')
        
        existing = User.query.filter(or_(User.username == username, User.email == email)).first()
        if existing:
            flash('Username or email already exists', 'danger')
            return redirect(request.url)
        
        user = User(
            username=username,
            email=email,
            role='agent',
            full_name=full_name,
            phone=phone,
            is_active=True
        )
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        flash(f'Agent {full_name} added successfully', 'success')
        return redirect(url_for('admin_agents'))
    
    return render_template('add_agent.html')

@app.route('/admin/agents/bulk-upload', methods=['GET', 'POST'])
@login_required
@admin_required
def bulk_upload_agents():
    if request.method == 'POST':
        if 'file' not in request.files or request.files['file'].filename == '':
            flash('No file selected', 'danger')
            return redirect(request.url)

        file = request.files['file']
        if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
            flash('Please upload an Excel (.xlsx/.xls) or CSV (.csv) file', 'danger')
            return redirect(request.url)

        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"agents_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
        file.save(file_path)

        try:
            df = pd.read_excel(file_path, engine='openpyxl')
            df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]

            added = 0
            skipped = 0
            errors = []

            for idx, row in df.iterrows():
                try:
                    username  = clean_data_value(row.get('username', ''))
                    email     = clean_data_value(row.get('email', ''))
                    password  = clean_data_value(row.get('password', ''))
                    full_name = clean_data_value(row.get('full_name', row.get('name', '')))
                    phone     = clean_data_value(row.get('phone', row.get('phone_number', '')))

                    if not username or not email or not password:
                        errors.append(f"Row {idx+2}: Missing username, email or password — skipped")
                        skipped += 1
                        continue

                    if User.query.filter(or_(User.username == username, User.email == email)).first():
                        errors.append(f"Row {idx+2}: {username} / {email} already exists — skipped")
                        skipped += 1
                        continue

                    user = User(username=username, email=email, role='agent',
                                full_name=full_name, phone=phone, is_active=True)
                    user.set_password(password)
                    db.session.add(user)
                    added += 1

                except Exception as e:
                    errors.append(f"Row {idx+2}: {str(e)}")
                    skipped += 1

            db.session.commit()
            flash(f'Bulk upload done — Added: {added}, Skipped: {skipped}', 'success' if added else 'warning')
            if errors:
                flash('Issues: ' + ' | '.join(errors[:5]), 'warning')

        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'danger')

        return redirect(url_for('admin_agents'))

    return render_template('bulk_upload_agents.html')

@app.route('/admin/agent/<int:agent_id>/change-password', methods=['POST'])
@login_required
@admin_required
def change_agent_password(agent_id):
    agent = User.query.get_or_404(agent_id)
    new_password = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()

    if not new_password:
        flash('Password cannot be empty', 'danger')
    elif len(new_password) < 6:
        flash('Password must be at least 6 characters', 'danger')
    elif new_password != confirm_password:
        flash('Passwords do not match', 'danger')
    else:
        agent.set_password(new_password)
        db.session.commit()
        flash(f'Password updated for {agent.full_name}', 'success')

    return redirect(url_for('admin_agents'))

@app.route('/admin/agent/<int:agent_id>/device-limit', methods=['GET', 'POST'])
@login_required
@admin_required
def agent_device_limit(agent_id):
    agent = User.query.get_or_404(agent_id)
    if request.method == 'POST':
        val = request.form.get('max_devices', '').strip()
        if val == '':
            agent.max_devices = None
        else:
            try:
                agent.max_devices = max(1, int(val))
            except (ValueError, TypeError):
                return jsonify({'ok': False, 'msg': 'Invalid value'})
        db.session.commit()
        return jsonify({'ok': True})
    return jsonify({'max_devices': agent.max_devices})

@app.route('/admin/agent/<int:agent_id>/resign', methods=['POST'])
@login_required
@admin_required
def resign_agent(agent_id):
    agent = User.query.get_or_404(agent_id)
    if agent.is_resigned:
        agent.is_resigned = False
        agent.resigned_at = None
    else:
        agent.is_resigned = True
        agent.resigned_at = datetime.utcnow()
        agent.is_active = False
    db.session.commit()
    return jsonify({'ok': True, 'is_resigned': agent.is_resigned,
                    'resigned_at': agent.resigned_at.strftime('%d %b %Y') if agent.resigned_at else None})

@app.route('/admin/agent/<int:agent_id>/toggle')
@login_required
@admin_required
def toggle_agent(agent_id):
    agent = User.query.get_or_404(agent_id)
    agent.is_active = not agent.is_active
    db.session.commit()
    
    status = 'activated' if agent.is_active else 'deactivated'
    flash(f'Agent {agent.username} {status}', 'success')
    return redirect(url_for('admin_agents'))

@app.route('/admin/credits/add', methods=['POST'])
@login_required
@admin_required
def admin_add_credits():
    action = request.form.get('credit_action')
    amount = int(request.form.get('amount', 0))

    if amount <= 0:
        flash('Amount must be greater than 0.', 'danger')
        return redirect(url_for('admin_agents'))

    if action == 'all':
        agents = User.query.filter_by(role='agent').all()
        for agent in agents:
            agent.credit_balance += amount
            db.session.add(CreditLog(
                user_id=agent.id,
                amount=amount,
                balance_after=agent.credit_balance,
                reason=f'Admin bulk top-up',
                admin_id=current_user.id
            ))
        db.session.commit()
        flash(f'Added {amount} credit(s) to all {len(agents)} agents.', 'success')

    elif action == 'specific':
        raw = request.form.get('agent_ids', '')
        ids = [i.strip() for i in raw.replace(',', ' ').split() if i.strip().isdigit()]
        if not ids:
            flash('No valid agent IDs entered.', 'danger')
            return redirect(url_for('admin_agents'))
        count = 0
        for aid in ids:
            agent = User.query.filter_by(id=int(aid), role='agent').first()
            if agent:
                agent.credit_balance += amount
                db.session.add(CreditLog(
                    user_id=agent.id,
                    amount=amount,
                    balance_after=agent.credit_balance,
                    reason=f'Admin top-up',
                    admin_id=current_user.id
                ))
                count += 1
        db.session.commit()
        flash(f'Added {amount} credit(s) to {count} agent(s).', 'success')

    return redirect(url_for('admin_agents'))

@app.route('/admin/customers')
@login_required
@admin_required
def admin_customers():
    flash('All Customers page is coming soon.', 'info')
    return redirect(url_for('admin_dashboard'))
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    if per_page not in (25, 50, 100, 200):
        per_page = 50

    name_filter    = request.args.get('name', '').strip()
    phone_filter   = request.args.get('phone', '').strip()
    ic_filter      = request.args.get('ic', '').strip()
    source_filter  = request.args.get('source', '').strip()
    email_filter   = request.args.get('email', '').strip()

    query = CustomerData.query.options(load_only(
        CustomerData.id,
        CustomerData.name,
        CustomerData.contact_number,
        CustomerData.ic_number,
        CustomerData.email,
        CustomerData.address,
        CustomerData.data_source,
        CustomerData.created_at,
    ))

    if name_filter:
        query = query.filter(CustomerData.name.ilike(f'%{name_filter}%'))
    if phone_filter:
        digits = re.sub(r'\D', '', phone_filter)
        query = query.filter(CustomerData.contact_number.ilike(f'%{digits}%'))
    if ic_filter:
        clean = re.sub(r'[\s\-]', '', ic_filter)
        query = query.filter(CustomerData.ic_number.ilike(f'%{clean}%'))
    if source_filter:
        query = query.filter(CustomerData.data_source.ilike(f'%{source_filter}%'))
    if email_filter:
        query = query.filter(CustomerData.email.ilike(f'%{email_filter}%'))

    pagination = query.order_by(CustomerData.id.desc()).paginate(page=page, per_page=per_page, error_out=False)

    sources = [r[0] for r in db.session.execute(
        text("SELECT DISTINCT data_source FROM customer_data WHERE data_source IS NOT NULL AND data_source != '' ORDER BY data_source")
    ).fetchall()]

    return render_template(
        'admin_customers.html',
        pagination=pagination,
        customers=pagination.items,
        per_page=per_page,
        name_filter=name_filter,
        phone_filter=phone_filter,
        ic_filter=ic_filter,
        source_filter=source_filter,
        email_filter=email_filter,
        sources=sources,
    )

@app.route('/admin/search', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_search():
    results = []
    search_term = ''
    search_type = ''
    search_meta = {}

    if request.method == 'POST':
        search_term = request.form.get('search_term', '').strip()
        search_type = request.form.get('search_type', '')
        results, search_meta = perform_customer_search(search_type, search_term)
        flash(search_meta.get('message') or f'Found {len(results)} result(s)', 'info' if search_meta.get('ok') else 'warning')

    return render_template('admin_search.html', results=results, search_term=search_term,
                           search_type=search_type, search_meta=search_meta)

@app.route('/admin/customer/<int:customer_id>/request-delete', methods=['POST'])
@login_required
@admin_required
def admin_request_customer_delete(customer_id):
    customer = CustomerData.query.get_or_404(customer_id)

    pending_request = CustomerDeleteApproval.query.filter_by(
        customer_id=customer.id,
        status='pending'
    ).first()
    if pending_request:
        return jsonify({
            'ok': False,
            'message': 'A delete request for this customer is already pending boss approval.'
        }), 409

    approval = create_customer_delete_approval(customer, current_user)
    sent = send_customer_delete_approval_request(approval)

    if not sent:
        approval.status = 'failed'
        approval.processed_at = datetime.utcnow()
        db.session.commit()
        return jsonify({
            'ok': False,
            'message': 'Could not send the delete approval message to boss. Please check Telegram settings.'
        }), 500

    return jsonify({
        'ok': True,
        'message': 'Delete request sent to boss for approval.'
    })

@app.route('/admin/customer/<int:customer_id>/enrichment', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_customer_enrichment(customer_id):
    customer = CustomerData.query.get_or_404(customer_id)
    crawl_summary = None
    connectivity_test = None

    if request.method == 'POST':
        action = request.form.get('action', 'save')

        if action == 'test':
            connectivity_test = run_enrichment_connectivity_test()

        if action == 'crawl':
            job, created = enqueue_enrichment_job(customer.id, current_user.id)
            if created:
                flash(f'Background crawl job #{job.id} started.', 'success')
            else:
                flash(f'Background crawl job #{job.id} is already {job.status}.', 'info')
            return redirect(url_for('admin_customer_enrichment', customer_id=customer.id))

        if action == 'save':
            enrichment = CustomerEnrichment(
                customer_id=customer.id,
                platform=clean_data_value(request.form.get('platform')),
                source_url=clean_data_value(request.form.get('source_url')),
                source_type='public_web',
                matched_name=clean_data_value(request.form.get('matched_name')),
                matched_location=clean_data_value(request.form.get('matched_location')),
                matched_company=clean_data_value(request.form.get('matched_company')),
                matched_title=clean_data_value(request.form.get('matched_title')),
                matched_phone=clean_data_value(request.form.get('matched_phone')),
                matched_email=clean_data_value(request.form.get('matched_email')),
                confidence_score=max(0, min(100, int(request.form.get('confidence_score') or 50))),
                review_status=clean_data_value(request.form.get('review_status')) or 'pending',
                notes=clean_data_value(request.form.get('notes')),
                raw_data=json.dumps({
                    'search_keywords': request.form.get('search_keywords', ''),
                    'captured_from': request.form.get('captured_from', '')
                }),
                created_by=current_user.id
            )
            db.session.add(enrichment)
            db.session.commit()
            flash('Public enrichment record saved.', 'success')
            return redirect(url_for('admin_customer_enrichment', customer_id=customer.id))

        if action == 'apply':
            enrichment = CustomerEnrichment.query.filter_by(
                id=request.form.get('enrichment_id'),
                customer_id=customer.id
            ).first_or_404()

            additional = json.loads(customer.additional_data) if customer.additional_data else {}
            if enrichment.matched_company:
                additional['enriched_company'] = enrichment.matched_company
            if enrichment.matched_title:
                additional['enriched_title'] = enrichment.matched_title
            if enrichment.source_url:
                additional['enrichment_source_url'] = enrichment.source_url
            if enrichment.platform:
                additional['enrichment_platform'] = enrichment.platform
            if enrichment.notes:
                additional['enrichment_notes'] = enrichment.notes

            _merge_payload_into_customer(
                customer,
                {
                    'contact_number': enrichment.matched_phone,
                    'email': enrichment.matched_email,
                    'address': enrichment.matched_location,
                    'additional_data': json.dumps(additional),
                },
                source_label=f'enrichment:{enrichment.id}',
                merge_reason='approved_enrichment'
            )
            enrichment.review_status = 'approved'
            db.session.commit()
            flash('Selected enrichment was applied to the customer record.', 'success')
            return redirect(url_for('admin_customer_enrichment', customer_id=customer.id))

    if request.method == 'GET' and request.args.get('auto', '0') == '1':
        has_auto_records = CustomerEnrichment.query.filter_by(
            customer_id=customer.id,
            source_type='public_web_auto'
        ).count() > 0
        if not has_auto_records or request.args.get('refresh') == '1':
            crawl_summary = auto_crawl_customer_enrichment(customer, current_user.id)
            if crawl_summary['created'] > 0:
                flash(f"Auto-crawl saved {crawl_summary['created']} candidate result(s).", 'success')
            elif crawl_summary['errors']:
                flash('Auto-crawl could not fetch public results right now.', 'warning')

    enrichments = CustomerEnrichment.query.filter_by(customer_id=customer.id).order_by(CustomerEnrichment.created_at.desc()).all()
    latest_job = get_latest_enrichment_job(customer.id)
    search_links = build_customer_search_queries(customer)
    return render_template(
        'admin_customer_enrichment.html',
        customer=customer,
        enrichments=enrichments,
        search_links=search_links,
        crawl_summary=crawl_summary,
        connectivity_test=connectivity_test,
        enrichment_mode=app.config.get('ENRICHMENT_CRAWL_MODE', 'auto'),
        latest_job=latest_job
    )

@app.route('/admin/customer/<int:customer_id>/enrichment/job-status')
@login_required
@admin_required
def admin_customer_enrichment_job_status(customer_id):
    customer = CustomerData.query.get_or_404(customer_id)
    job = get_latest_enrichment_job(customer.id)
    if not job:
        return jsonify({'found': False})

    return jsonify({
        'found': True,
        'id': job.id,
        'status': job.status,
        'message': job.message or '',
        'created_count': job.created_count or 0,
        'checked_count': job.checked_count or 0,
        'providers_used': json.loads(job.providers_used or '[]'),
        'error_log': json.loads(job.error_log or '[]'),
        'debug_samples': json.loads(job.debug_samples or '[]'),
        'created_at': job.created_at.isoformat() if job.created_at else '',
        'started_at': job.started_at.isoformat() if job.started_at else '',
        'completed_at': job.completed_at.isoformat() if job.completed_at else ''
    })

@app.route('/admin/waiting-approval')
def waiting_approval():
    token = session.get('pending_approval_token')
    if not token:
        return redirect(url_for('login'))
    approval = AdminLoginApproval.query.filter_by(approval_token=token).first()
    if not approval:
        return redirect(url_for('login'))
    return render_template('waiting_approval.html', token=token,
                           expires_at=approval.expires_at.isoformat())

@app.route('/admin/check-approval/<token>')
def check_approval(token):
    approval = AdminLoginApproval.query.filter_by(approval_token=token).first()
    if not approval:
        return jsonify({'status': 'not_found'})

    if approval.status == 'pending' and datetime.utcnow() > approval.expires_at:
        approval.status = 'expired'
        SystemSettings.set('system_locked', '1')
        SystemSettings.set('system_locked_reason',
            f'Admin login approval expired at {datetime.utcnow().strftime("%Y-%m-%d %H:%M")} '
            f'for {approval.user.full_name} from {approval.ip_address}')
        db.session.commit()
        return jsonify({'status': 'expired'})

    if approval.status == 'approved':
        user = User.query.get(approval.user_id)
        if user and user.is_active:
            session_id = str(uuid.uuid4())
            login_log = LoginLog(
                user_id=user.id,
                ip_address=approval.ip_address,
                location=get_location_from_ip(approval.ip_address),
                device_info='Telegram-approved login',
                connection_type='wifi',
                session_id=session_id
            )
            db.session.add(login_log)
            device = AdminDevice.query.filter_by(
                user_id=user.id, fingerprint=approval.fingerprint).first()
            if device:
                device.status = 'trusted'
            user.last_login = datetime.utcnow()
            db.session.commit()
            login_user(user)
            session.permanent = True
            session['login_log_id'] = login_log.id
            session['session_id'] = session_id
            session.pop('pending_approval_token', None)
        return jsonify({'status': 'approved'})

    return jsonify({'status': approval.status})

def _handle_panic_command(chat_id, activate):
    """Activate or deactivate panic mode via Telegram command."""
    token = SystemSettings.get('telegram_bot_token')
    if not token:
        return
    now = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
    if activate:
        dp.activate_panic()
        text = (
            f"\U0001f6a8 *PANIC MODE ACTIVATED*\n\n"
            f"Customer database is now replaced with decoy records.\n"
            f"Anyone downloading data will receive fake records only.\n\n"
            f"⏰ Activated at: {now}\n\n"
            f"Send /unpanic to deactivate when safe."
        )
    else:
        dp.deactivate_panic()
        text = (
            f"✅ *Panic mode deactivated*\n\n"
            f"System restored to normal. Real customer data is served again.\n\n"
            f"⏰ Deactivated at: {now}"
        )
    try:
        requests.post(
            f'https://api.telegram.org/bot{token}/sendMessage',
            json={'chat_id': chat_id, 'text': text, 'parse_mode': 'Markdown'},
            timeout=5
        )
    except Exception as e:
        app.logger.error(f"Telegram panic response error: {e}")


@app.route('/telegram/webhook', methods=['POST'])
def telegram_webhook():
    data = request.json
    if not data:
        return jsonify({'ok': True})

    # Handle text message commands from boss only
    if 'message' in data:
        msg = data['message']
        text = msg.get('text', '').strip()
        cmd = text.split('@')[0].lower()  # strip bot username suffix if present
        from_id = str(msg.get('from', {}).get('id', ''))
        boss_id = str(SystemSettings.get('telegram_boss_chat_id', ''))
        if from_id and boss_id and from_id == boss_id:
            if cmd == '/panic':
                _handle_panic_command(from_id, activate=True)
            elif cmd in ('/unpanic', '/calm'):
                _handle_panic_command(from_id, activate=False)
        return jsonify({'ok': True})

    if 'callback_query' not in data:
        return jsonify({'ok': True})

    cb = data['callback_query']
    cb_data = cb.get('data', '')
    cb_id = cb['id']
    msg = cb.get('message', {})
    message_id = msg.get('message_id')
    chat_id = msg.get('chat', {}).get('id')

    if '_' not in cb_data:
        return jsonify({'ok': True})

    parts = cb_data.split('_', 1)
    if len(parts) != 2:
        return jsonify({'ok': True})
    action, token = parts

    # Handle brute force alert callbacks
    if action == 'unlocksys':
        stored_token = SystemSettings.get('unlock_reminder_token')
        if stored_token == token:
            SystemSettings.set('system_locked', '0')
            SystemSettings.set('system_locked_reason', '')
            SystemSettings.set('system_locked_since', '')
            SystemSettings.set('unlock_reminder_token', '')
            tg_answer_callback(cb_id, '🔓 System unlocked!')
            tg_edit_message(chat_id, message_id,
                f'🔓 *System Unlocked*\n\n'
                f'Unlocked by boss at {datetime.utcnow().strftime("%Y-%m-%d %H:%M")} UTC')
        else:
            tg_answer_callback(cb_id, 'Invalid or expired token.')
        return jsonify({'ok': True})

    if action == 'keeplocked':
        SystemSettings.set('unlock_reminder_token', '')
        SystemSettings.set('system_locked_since', datetime.utcnow().isoformat())
        tg_answer_callback(cb_id, '🔒 System remains locked. Reminder in 30 minutes.')
        tg_edit_message(chat_id, message_id, '🔒 *System remains locked.*\nYou will receive another reminder in 30 minutes.')
        return jsonify({'ok': True})

    if action == 'lockbf':
        alert = BruteForceAlert.query.filter_by(alert_token=token).first()
        if alert and alert.status == 'pending':
            alert.status = 'locked'
            SystemSettings.set('system_locked', '1')
            SystemSettings.set('system_locked_since', datetime.utcnow().isoformat())
            SystemSettings.set('system_locked_reason',
                f'System locked by boss via Telegram — suspicious login activity detected '
                f'for {alert.user.full_name} from {alert.ip_address}')
            db.session.commit()
            tg_answer_callback(cb_id, '🔒 System locked!')
            tg_edit_message(chat_id, message_id,
                f'🔒 *System Locked*\n\n'
                f'Locked due to suspicious activity for *{alert.user.full_name}*\n'
                f'IP: `{alert.ip_address}`\n'
                f'Locked at: {datetime.utcnow().strftime("%Y-%m-%d %H:%M")} UTC')
        else:
            tg_answer_callback(cb_id, 'Already handled.')
        return jsonify({'ok': True})

    if action == 'ignorebf':
        alert = BruteForceAlert.query.filter_by(alert_token=token).first()
        if alert and alert.status == 'pending':
            alert.status = 'ignored'
            db.session.commit()
            tg_answer_callback(cb_id, '✅ Alert ignored.')
            tg_edit_message(chat_id, message_id,
                f'✅ *Alert Ignored*\n\n'
                f'Suspicious activity alert for *{alert.user.full_name}* was dismissed.')
        else:
            tg_answer_callback(cb_id, 'Already handled.')
        return jsonify({'ok': True})

    if action in ('custdelapprove', 'custdeldecline'):
        process_customer_delete_callback(
            action,
            token,
            callback_id=cb_id,
            chat_id=chat_id,
            message_id=message_id
        )
        return jsonify({'ok': True})

    approval = AdminLoginApproval.query.filter_by(approval_token=token).first()

    if not approval or approval.status != 'pending':
        tg_answer_callback(cb_id, 'This request has already been handled.')
        return jsonify({'ok': True})

    if datetime.utcnow() > approval.expires_at:
        approval.status = 'expired'
        db.session.commit()
        tg_answer_callback(cb_id, 'This request has expired.')
        tg_edit_message(chat_id, message_id, '⏰ *Login request expired.*')
        return jsonify({'ok': True})

    if action == 'approve':
        approval.status = 'approved'
        db.session.commit()
        tg_answer_callback(cb_id, f'✅ Login approved for {approval.user.full_name}')
        tg_edit_message(chat_id, message_id,
            f'✅ *Login Approved*\n\n'
            f'Admin: *{approval.user.full_name}*\n'
            f'IP: `{approval.ip_address}`\n'
            f'Approved at: {datetime.utcnow().strftime("%Y-%m-%d %H:%M")} UTC')

    elif action == 'decline':
        approval.status = 'declined'
        db.session.commit()
        tg_answer_callback(cb_id, f'❌ Login declined for {approval.user.full_name}')
        tg_edit_message(chat_id, message_id,
            f'❌ *Login Declined*\n\n'
            f'Admin: *{approval.user.full_name}*\n'
            f'IP: `{approval.ip_address}`\n'
            f'Declined at: {datetime.utcnow().strftime("%Y-%m-%d %H:%M")} UTC')

    return jsonify({'ok': True})

@app.route('/admin/unlock-system')
@login_required
@admin_required
def unlock_system():
    SystemSettings.set('system_locked', '0')
    SystemSettings.set('system_locked_reason', '')
    flash('System unlocked. All users can now log in.', 'success')
    return redirect(url_for('admin_security'))

@app.route('/system-locked')
def system_locked_page():
    reason = session.get('lock_page_reason') or SystemSettings.get('system_locked_reason', 'Unauthorised login attempt detected.')
    visitor_ip = get_client_ip()
    user_agent = request.headers.get('User-Agent', '')
    telemetry = {
        'ip_address': visitor_ip,
        'browser': detect_browser(user_agent),
        'operating_system': detect_operating_system(user_agent),
        'device_category': detect_device_category(user_agent),
        'connection_type': detect_connection_type(user_agent),
        'server_location': get_location_from_ip(visitor_ip),
        'user_agent': user_agent
    }
    return render_template('system_locked.html', reason=reason, visitor_ip=visitor_ip, telemetry=telemetry), 423

@app.route('/security/client-telemetry', methods=['POST'])
def security_client_telemetry():
    payload = request.get_json(silent=True) or {}
    ip_address = get_client_ip()
    details = {
        'browser': payload.get('browser'),
        'operating_system': payload.get('operating_system'),
        'device_category': payload.get('device_category'),
        'language': payload.get('language'),
        'timezone': payload.get('timezone'),
        'screen': payload.get('screen'),
        'fingerprint': payload.get('fingerprint'),
        'platform': payload.get('platform'),
        'location_permission': payload.get('location_permission')
    }
    log_security_event(
        'client_telemetry',
        'observed',
        ip_address,
        details=json.dumps(details)
    )
    return jsonify({'ok': True})

@app.route('/admin/block-ip', methods=['POST'])
@login_required
@admin_required
def admin_block_ip():
    ip_address = request.form.get('ip_address', '').strip()
    if not ip_address:
        flash('No IP address provided.', 'danger')
        return redirect(url_for('admin_logs'))

    created = block_ip_address(ip_address)
    log_security_event('ip_block', 'blocked', ip_address, user=current_user,
                       details='IP address blocked by administrator action.')

    if created:
        flash(f'IP address blocked: {ip_address}', 'warning')
    else:
        flash(f'IP address already blocked: {ip_address}', 'info')

    next_url = request.form.get('next') or url_for('admin_logs')
    return redirect(next_url)

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def force_change_password():
    if request.method == 'POST':
        new_pw = request.form.get('new_password', '').strip()
        confirm_pw = request.form.get('confirm_password', '').strip()
        if len(new_pw) < 6:
            flash('Password must be at least 6 characters.', 'danger')
        elif new_pw != confirm_pw:
            flash('Passwords do not match.', 'danger')
        elif current_user.check_password(new_pw):
            flash('New password cannot be the same as your current password.', 'danger')
        else:
            current_user.set_password(new_pw)
            db.session.commit()
            session.pop('force_pw_change', None)
            flash('Password updated successfully!', 'success')
            return redirect(url_for('admin_dashboard') if current_user.role == 'admin' else url_for('agent_dashboard'))
    return render_template('force_change_password.html')

@app.route('/admin/change-password', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_change_password():
    if request.method == 'POST':
        current_pw = request.form.get('current_password', '').strip()
        new_pw = request.form.get('new_password', '').strip()
        confirm_pw = request.form.get('confirm_password', '').strip()
        if not current_user.check_password(current_pw):
            flash('Current password is incorrect.', 'danger')
        elif len(new_pw) < 6:
            flash('New password must be at least 6 characters.', 'danger')
        elif new_pw != confirm_pw:
            flash('Passwords do not match.', 'danger')
        elif current_user.check_password(new_pw):
            flash('New password cannot be the same as current password.', 'danger')
        else:
            current_user.set_password(new_pw)
            db.session.commit()
            flash('Password updated successfully!', 'success')
            return redirect(url_for('admin_dashboard'))
    return render_template('admin_change_password.html')

@app.route('/admin/security', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_security():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'password_policy':
            expiry = request.form.get('password_expiry_days', '0')
            SystemSettings.set('password_expiry_days', expiry)
            if int(expiry) > 0:
                cutoff = datetime.utcnow() - timedelta(days=int(expiry))
                agents = User.query.filter_by(role='agent').all()
                forced = sum(1 for a in agents
                             if not a.password_changed_at or a.password_changed_at < cutoff)
                for a in agents:
                    if not a.password_changed_at or a.password_changed_at < cutoff:
                        a.must_change_password = True
                db.session.commit()
                flash(f'Policy saved — {forced} agent(s) flagged for password reset.', 'success')
            else:
                flash('Password expiry disabled.', 'info')

        elif action == 'pw_reset_limit':
            try:
                max_r = max(1, int(request.form.get('pw_reset_max', '3')))
            except (ValueError, TypeError):
                max_r = 3
            try:
                window_d = max(1, int(request.form.get('pw_reset_window_days', '7')))
            except (ValueError, TypeError):
                window_d = 7
            SystemSettings.set('pw_reset_max', str(max_r))
            SystemSettings.set('pw_reset_window_days', str(window_d))
            flash(f'Password reset limit set to {max_r} times per {window_d} days.', 'success')

        elif action == 'session_timeout':
            mins = request.form.get('session_timeout_minutes', '30').strip()
            try:
                mins = max(5, int(mins))
            except (ValueError, TypeError):
                mins = 30
            SystemSettings.set('session_timeout_minutes', str(mins))
            app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=mins)
            flash(f'Session timeout set to {mins} minutes.', 'success')

        elif action == 'device_limit':
            limit = request.form.get('max_agent_devices', '3').strip()
            try:
                limit = max(1, int(limit))
            except (ValueError, TypeError):
                limit = 3
            SystemSettings.set('max_agent_devices', str(limit))
            flash(f'Device limit set to {limit} per agent.', 'success')

        elif action == 'telegram_settings':
            SystemSettings.set('telegram_approval_enabled',
                               '1' if request.form.get('telegram_enabled') else '0')
            SystemSettings.set('agent_ip_approval_enabled',
                               '1' if request.form.get('agent_ip_approval_enabled') else '0')
            SystemSettings.set('telegram_bot_token',
                               request.form.get('telegram_bot_token', '').strip())
            SystemSettings.set('telegram_boss_chat_id',
                               request.form.get('telegram_boss_chat_id', '').strip())
            SystemSettings.set('telegram_bot_username',
                               request.form.get('telegram_bot_username', '').strip().lstrip('@'))
            flash('Telegram settings saved.', 'success')

        elif action == 'phase2_toggle':
            SystemSettings.set('phase2_enabled',
                               '1' if request.form.get('phase2_enabled') else '0')
            flash('Phase 2 setting updated.', 'success')

        elif action == 'create_admin':
            new_username = request.form.get('new_username', '').strip()
            new_email = request.form.get('new_email', '').strip()
            new_full_name = request.form.get('new_full_name', '').strip()
            new_password = request.form.get('new_password', '').strip()
            if not all([new_username, new_email, new_full_name, new_password]):
                flash('All fields are required to create an admin account.', 'danger')
            elif len(new_password) < 8:
                flash('Password must be at least 8 characters.', 'danger')
            elif User.query.filter_by(username=new_username).first():
                flash(f'Username "{new_username}" is already taken.', 'danger')
            elif User.query.filter_by(email=new_email).first():
                flash(f'Email "{new_email}" is already registered.', 'danger')
            else:
                new_admin = User(
                    username=new_username,
                    email=new_email,
                    full_name=new_full_name,
                    role='admin',
                    is_active=True
                )
                new_admin.set_password(new_password)
                db.session.add(new_admin)
                db.session.commit()
                ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)
                log_security_event('admin_created', 'success', ip_address, user=current_user,
                                   details=f'Created admin account: {new_username}')
                flash(f'Admin account "{new_username}" created successfully.', 'success')

        return redirect(url_for('admin_security'))

    current_expiry = SystemSettings.get('password_expiry_days', '0')
    admin_device_list = AdminDevice.query.order_by(AdminDevice.first_seen.desc()).all()
    system_locked = SystemSettings.get('system_locked', '0') == '1'
    lock_reason = SystemSettings.get('system_locked_reason', '')
    admin_users = User.query.filter_by(role='admin').order_by(User.created_at.asc()).all()
    return render_template('admin_security.html',
                           current_expiry=current_expiry,
                           admin_devices=admin_device_list,
                           system_locked=system_locked,
                           lock_reason=lock_reason,
                           admin_users=admin_users,
                           dp_enabled=dp.status(),
                           panic_enabled=dp.is_panic(),
                           panic_token=SystemSettings.get('panic_secret_token', ''),
                           phase2_enabled=SystemSettings.get('phase2_enabled', '0') == '1',
                           max_agent_devices=SystemSettings.get('max_agent_devices', '3'),
                           telegram_enabled=SystemSettings.get('telegram_approval_enabled', '0'),
                           agent_ip_approval_enabled=SystemSettings.get('agent_ip_approval_enabled', '0') == '1',
                           telegram_bot_token=SystemSettings.get('telegram_bot_token', ''),
                           telegram_boss_chat_id=SystemSettings.get('telegram_boss_chat_id', ''),
                           telegram_bot_username=SystemSettings.get('telegram_bot_username', ''),
                           current_session_timeout=SystemSettings.get('session_timeout_minutes', '30'),
                           pw_reset_max=SystemSettings.get('pw_reset_max', '3'),
                           pw_reset_window_days=SystemSettings.get('pw_reset_window_days', '7'))

@app.route('/admin/security/delete-admin/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_admin_account(user_id):
    target = User.query.get_or_404(user_id)
    if target.id == current_user.id:
        flash('You cannot delete your own admin account.', 'danger')
        return redirect(url_for('admin_security'))
    if target.role != 'admin':
        flash('Target user is not an admin.', 'danger')
        return redirect(url_for('admin_security'))
    ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)
    log_security_event('admin_deleted', 'success', ip_address, user=current_user,
                       details=f'Deleted admin account: {target.username}')
    # Delete related records before deleting user to avoid NOT NULL FK constraint errors
    AdminDevice.query.filter_by(user_id=target.id).delete()
    LoginLog.query.filter_by(user_id=target.id).delete()
    db.session.delete(target)
    db.session.commit()
    flash(f'Admin account "{target.username}" has been deleted.', 'success')
    return redirect(url_for('admin_security'))


@app.route('/admin/security/toggle-data-protection', methods=['POST'])
@login_required
@admin_required
def toggle_data_protection():
    if dp.status():
        dp.disable()
        flash('Data protection disabled. Agents can now see full customer data.', 'info')
    else:
        dp.enable()
        flash('Data protection enabled. Agent views and downloads are now masked.', 'success')
    return redirect(url_for('admin_security'))


@app.route('/x/<token>/on')
def panic_activate(token):
    stored = SystemSettings.get('panic_secret_token', '')
    if not stored or token != stored:
        return '', 404
    dp.activate_panic()
    return '''<!DOCTYPE html>
<html><head><meta name="viewport" content="width=device-width,initial-scale=1">
<style>body{font-family:sans-serif;display:flex;align-items:center;justify-content:center;
min-height:100vh;margin:0;background:#fee2e2}
.box{background:#fff;border-radius:12px;padding:32px;text-align:center;max-width:320px;
border:2px solid #dc2626;box-shadow:0 4px 20px rgba(220,38,38,0.2)}
h2{color:#dc2626;margin:0 0 12px}p{color:#374151;font-size:15px;line-height:1.6}
</style></head><body>
<div class="box">
<h2>PANIC MODE ON</h2>
<p>System is now serving decoy data.<br>All customer searches return fake records.</p>
<p style="font-size:13px;color:#6b7280;margin-top:16px">Tap your OFF bookmark when safe.</p>
</div></body></html>''', 200


@app.route('/x/<token>/off')
def panic_deactivate(token):
    stored = SystemSettings.get('panic_secret_token', '')
    if not stored or token != stored:
        return '', 404
    dp.deactivate_panic()
    return '''<!DOCTYPE html>
<html><head><meta name="viewport" content="width=device-width,initial-scale=1">
<style>body{font-family:sans-serif;display:flex;align-items:center;justify-content:center;
min-height:100vh;margin:0;background:#dcfce7}
.box{background:#fff;border-radius:12px;padding:32px;text-align:center;max-width:320px;
border:2px solid #16a34a;box-shadow:0 4px 20px rgba(22,163,74,0.15)}
h2{color:#16a34a;margin:0 0 12px}p{color:#374151;font-size:15px;line-height:1.6}
</style></head><body>
<div class="box">
<h2>PANIC MODE OFF</h2>
<p>System restored to normal.<br>Real customer data is now served.</p>
</div></body></html>''', 200


@app.route('/admin/phase2')
@login_required
@admin_required
def phase2_overview():
    if SystemSettings.get('phase2_enabled', '0') != '1':
        flash('Phase 2 is not enabled.', 'warning')
        return redirect(url_for('admin_dashboard'))
    return render_template('phase2_overview.html')

@app.route('/admin/admin-device/<int:device_id>/trust')
@login_required
@admin_required
def trust_admin_device(device_id):
    device = AdminDevice.query.get_or_404(device_id)
    device.status = 'trusted'
    db.session.commit()
    flash(f'Device trusted: {device.label}', 'success')
    return redirect(url_for('admin_security'))

@app.route('/admin/admin-device/<int:device_id>/block')
@login_required
@admin_required
def block_admin_device(device_id):
    device = AdminDevice.query.get_or_404(device_id)
    device.status = 'blocked'
    db.session.commit()
    flash(f'Device blocked: {device.label}', 'warning')
    return redirect(url_for('admin_security'))

@app.route('/admin/admin-device/<int:device_id>/delete')
@login_required
@admin_required
def delete_admin_device(device_id):
    device = AdminDevice.query.get_or_404(device_id)
    db.session.delete(device)
    db.session.commit()
    flash('Device removed.', 'info')
    return redirect(url_for('admin_security'))

@app.route('/admin/devices')
@login_required
@admin_required
def admin_devices():
    pending = AgentDevice.query.filter_by(status='pending').order_by(AgentDevice.first_seen.desc()).all()
    all_devices = AgentDevice.query.order_by(AgentDevice.first_seen.desc()).all()
    return render_template('admin_devices.html', pending=pending, all_devices=all_devices)

@app.route('/admin/device/<int:device_id>/approve')
@login_required
@admin_required
def approve_device(device_id):
    device = AgentDevice.query.get_or_404(device_id)
    device.status = 'approved'
    device.approved_by = current_user.id
    device.approved_at = datetime.utcnow()
    if device.pending_ip:
        device.ip_address = device.pending_ip
        device.pending_ip = None
    db.session.commit()
    flash(f'Device approved for {device.user.full_name}', 'success')
    return redirect(url_for('admin_devices'))

@app.route('/admin/device/<int:device_id>/block')
@login_required
@admin_required
def block_device(device_id):
    device = AgentDevice.query.get_or_404(device_id)
    device.status = 'blocked'
    db.session.commit()
    flash(f'Device blocked for {device.user.full_name}', 'warning')
    return redirect(url_for('admin_devices'))

@app.route('/admin/device/<int:device_id>/delete')
@login_required
@admin_required
def delete_device(device_id):
    device = AgentDevice.query.get_or_404(device_id)
    db.session.delete(device)
    db.session.commit()
    flash('Device removed', 'info')
    return redirect(url_for('admin_devices'))

@app.route('/admin/device/<int:device_id>/label', methods=['POST'])
@login_required
@admin_required
def label_device(device_id):
    device = AgentDevice.query.get_or_404(device_id)
    device.label = request.form.get('label', device.label)
    db.session.commit()
    return redirect(url_for('admin_devices'))

@app.route('/admin/logs')
@login_required
@admin_required
def admin_logs():
    login_logs = LoginLog.query.order_by(LoginLog.login_time.desc()).limit(100).all()
    search_logs = SearchLog.query.order_by(SearchLog.timestamp.desc()).limit(100).all()
    security_logs = SecurityEventLog.query.order_by(SecurityEventLog.created_at.desc()).limit(200).all()

    return render_template('admin_logs.html',
                           login_logs=login_logs,
                           search_logs=search_logs,
                           security_logs=security_logs)


@app.route('/admin/work-log', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_work_log():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'delete':
            entry = WorkLog.query.get(request.form.get('entry_id'))
            if entry:
                db.session.delete(entry)
                db.session.commit()
                flash('Entry deleted.', 'info')
        else:
            d = request.form.get('date', '')
            try:
                entry_date = datetime.strptime(d, '%Y-%m-%d').date()
            except Exception:
                entry_date = datetime.utcnow().date()
            entry = WorkLog(
                date=entry_date,
                title=request.form.get('title', '').strip(),
                description=request.form.get('description', '').strip(),
                category=request.form.get('category', 'Feature'),
                hours=float(request.form.get('hours') or 0),
            )
            db.session.add(entry)
            db.session.commit()
            flash('Entry added.', 'success')
        return redirect(url_for('admin_work_log'))

    if WorkLog.query.count() == 0:
        _seed_work_log()

    entries = WorkLog.query.order_by(WorkLog.date.desc(), WorkLog.id.desc()).all()
    total_hours = sum(e.hours or 0 for e in entries)
    return render_template('admin_work_log.html', entries=entries, total_hours=total_hours,
                           now=datetime.utcnow())


@app.route('/admin/data/merge', methods=['GET', 'POST'])
@login_required
@admin_required
def merge_data():
    if request.method == 'POST':
        old_upload_id = request.form.get('old_upload_id')
        new_upload_id = request.form.get('new_upload_id')
        
        old_upload = Upload.query.get(old_upload_id)
        new_upload = Upload.query.get(new_upload_id)
        
        if not old_upload or not new_upload:
            flash('Invalid upload selected', 'danger')
            return redirect(request.url)
        
        old_customers = CustomerData.query.filter_by(upload_id=old_upload_id).all()
        new_customers = CustomerData.query.filter_by(upload_id=new_upload_id).all()
        
        updated_count = 0
        new_count = 0
        
        for new_customer in new_customers:
            existing = find_customer_by_normalized_ic(new_customer.ic_number)
            if not existing and new_customer.contact_number:
                existing = CustomerData.query.filter(
                    CustomerData.contact_number == new_customer.contact_number
                ).first()
            
            if existing:
                _merge_payload_into_customer(
                    existing,
                    _customer_payload(new_customer),
                    source_label=f'upload_merge:{new_upload_id}',
                    customer_id=new_customer.id,
                    merge_reason='upload_to_upload_merge'
                )
                updated_count += 1
            else:
                db.session.add(new_customer)
                new_count += 1
        
        db.session.commit()
        
        flash(f'Data merged successfully! Updated: {updated_count}, New: {new_count}', 'success')
        return redirect(url_for('admin_dashboard'))
    
    uploads = Upload.query.filter_by(status='processed').order_by(Upload.upload_date.desc()).all()
    return render_template('merge_data.html', uploads=uploads)

# ==================== AGENT ROUTES ====================

@app.route('/agent/device-waiting')
@login_required
def agent_device_waiting():
    if current_user.role != 'agent':
        return redirect(url_for('admin_dashboard'))
    fp = session.get('pending_device_fp')
    if not fp:
        return redirect(url_for('agent_dashboard'))
    device = AgentDevice.query.filter_by(user_id=current_user.id, fingerprint=fp).first()
    if not device or device.status == 'approved':
        session.pop('pending_device_fp', None)
        return redirect(url_for('agent_dashboard'))
    bot_username = SystemSettings.get('telegram_bot_username', '')
    return render_template('agent_device_waiting.html', bot_username=bot_username)

@app.route('/agent/device-resend-telegram', methods=['POST'])
@login_required
def agent_device_resend_telegram():
    if current_user.role != 'agent':
        return jsonify({'ok': False, 'msg': 'Unauthorized'})
    fp = session.get('pending_device_fp')
    if not fp:
        return jsonify({'ok': False, 'msg': 'No pending device'})
    device = AgentDevice.query.filter_by(user_id=current_user.id, fingerprint=fp, status='pending').first()
    if not device:
        return jsonify({'ok': False, 'msg': 'Device not found or already handled'})
    ok = send_agent_device_confirmation(current_user, device)
    return jsonify({'ok': ok})

@app.route('/agent/device-pending')
@login_required
def agent_device_pending():
    if current_user.role != 'agent':
        return redirect(url_for('admin_dashboard'))
    fp = session.get('pending_device_fp', '')
    device = AgentDevice.query.filter_by(user_id=current_user.id, fingerprint=fp).first() if fp else None
    if not fp or not device:
        session.pop('pending_device_fp', None)
        return redirect(url_for('agent_dashboard'))
    status = device.status
    if status == 'approved':
        session.pop('pending_device_fp', None)
        return redirect(url_for('agent_dashboard'))
    return render_template('agent_device_pending.html', device_status=status)

@app.route('/agent/device-confirm-status')
@login_required
def agent_device_confirm_status():
    if current_user.role != 'agent':
        return jsonify({'status': 'error'})
    fp = session.get('pending_device_fp')
    if not fp:
        return jsonify({'status': 'no_pending'})
    device = AgentDevice.query.filter_by(user_id=current_user.id, fingerprint=fp).first()
    if not device:
        return jsonify({'status': 'not_found'})
    if device.status == 'approved':
        session.pop('pending_device_fp', None)
        session['agent_device_status'] = 'approved'
    elif device.status == 'blocked':
        session.pop('pending_device_fp', None)
        session['agent_device_status'] = 'blocked'
    return jsonify({'status': device.status})

@app.route('/admin/agent/<int:agent_id>/gen-tg-token', methods=['POST'])
@login_required
@admin_required
def admin_gen_tg_token(agent_id):
    agent = User.query.filter_by(id=agent_id, role='agent').first_or_404()
    agent.tg_link_token = secrets.token_hex(4).upper()
    db.session.commit()
    return jsonify({'token': agent.tg_link_token, 'name': agent.full_name})

@app.route('/admin/agent/<int:agent_id>/set-telegram', methods=['POST'])
@login_required
@admin_required
def admin_set_agent_telegram(agent_id):
    agent = User.query.filter_by(id=agent_id, role='agent').first_or_404()
    chat_id = request.form.get('telegram_chat_id', '').strip()
    agent.telegram_chat_id = chat_id or None
    db.session.commit()
    flash(f'Telegram chat ID updated for {agent.full_name}', 'success')
    return redirect(url_for('admin_agents'))

@app.route('/agent')
@login_required
def agent_dashboard():
    if current_user.role != 'agent':
        return redirect(url_for('admin_dashboard'))

    recent_logins = (LoginLog.query
                     .filter_by(user_id=current_user.id)
                     .order_by(LoginLog.login_time.desc())
                     .limit(5).all())
    current_ip = get_client_ip()
    return render_template('agent_dashboard.html',
                           recent_logins=recent_logins,
                           current_ip=current_ip)

@app.route('/agent/settings', methods=['GET', 'POST'])
@login_required
def agent_settings():
    if current_user.role != 'agent':
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'gen_token':
            current_user.tg_link_token = secrets.token_hex(4).upper()
            db.session.commit()
            flash(f'New code generated: {current_user.tg_link_token}', 'success')
        elif action == 'unlink':
            current_user.telegram_chat_id = None
            current_user.tg_link_token = None
            db.session.commit()
            flash('Telegram unlinked.', 'info')
        return redirect(url_for('agent_settings'))

    bot_username = SystemSettings.get('telegram_bot_username', '')
    return render_template('agent_settings.html',
                           bot_username=bot_username,
                           token=current_user.tg_link_token,
                           linked=bool(current_user.telegram_chat_id))


@app.route('/agent/search', methods=['GET', 'POST'])
@login_required
def agent_search():
    if current_user.role != 'agent':
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        return redirect(url_for('agent_search'))

    ip_address  = get_client_ip()
    fingerprint = request.cookies.get('fp', '')
    if fingerprint:
        device_status = check_device_trusted(current_user.id, fingerprint, ip_address)
    else:
        device_status = 'unknown'

    blocked = device_status in ('pending', 'new', 'blocked')
    return render_template('agent_search.html', blocked=blocked, device_status=device_status)

@app.route('/agent/search-ajax', methods=['POST'])
@login_required
def agent_search_ajax():
    if current_user.role != 'agent':
        return jsonify({'ok': False, 'message': 'Unauthorized'}), 403

    ip_address  = get_client_ip()
    fingerprint = request.form.get('fp') or request.cookies.get('fp', '')
    search_term = request.form.get('search_term', '').strip()
    search_type = request.form.get('search_type', 'keyword')

    if fingerprint:
        device_status = check_device_trusted(current_user.id, fingerprint, ip_address)
    else:
        device_status = 'unknown'

    if device_status in ('pending', 'new', 'blocked'):
        return jsonify({'ok': False, 'message': 'This device is not approved to view customer data yet.'})

    results, meta = perform_customer_search(search_type, search_term)

    search_log = SearchLog(
        user_id=current_user.id,
        search_term=search_term,
        search_type=search_type,
        results_count=len(results),
        ip_address=ip_address,
        user_agent=request.headers.get('User-Agent', ''),
        screenshot_taken=False,
        data_downloaded=False
    )
    db.session.add(search_log)
    db.session.commit()
    session['last_search_log_id'] = search_log.id

    results = dp.mask_customer_list(results, current_user)
    serialized = [{
        'id':             r.id,
        'name':           r.name or '',
        'contact_number': r.contact_number or '',
        'ic_number':      r.ic_number or '',
        'address':        (r.address or '')[:120],
        'data_source':    r.data_source or '',
        'email':          r.email or '',
    } for r in results]

    return jsonify({
        'ok':       meta.get('ok', True),
        'message':  meta.get('message', ''),
        'results':  serialized,
        'truncated':meta.get('truncated', False),
        'limit':    meta.get('limit', 500),
    })


@app.route('/agent/download/<int:search_log_id>')
@login_required
def download_results(search_log_id):
    if current_user.role != 'agent':
        return redirect(url_for('admin_dashboard'))

    search_log = SearchLog.query.get_or_404(search_log_id)

    if search_log.user_id != current_user.id:
        flash('Unauthorized access', 'danger')
        return redirect(url_for('agent_search'))

    results, search_meta = perform_customer_search(search_log.search_type, search_log.search_term, requested_limit='download')
    if not search_meta.get('ok'):
        flash(search_meta.get('message', 'Could not prepare download results.'), 'warning')
        return redirect(url_for('agent_search'))

    cost = calc_download_cost(len(results))
    if current_user.credit_balance < cost:
        flash(f'Insufficient credits. This download requires {cost} credit(s) but you only have {current_user.credit_balance}. Please contact your admin.', 'danger')
        return redirect(url_for('agent_search'))

    data = []
    for customer in results:
        additional = json.loads(customer.additional_data) if customer.additional_data else {}
        row = {
            'Name': customer.name,
            'Contact Number': customer.contact_number,
            'IC Number': customer.ic_number,
            'Address': customer.address,
            'Email': customer.email,
        }
        row = dp.mask_row(row, current_user)
        row.update(additional)
        data.append(row)

    df = pd.DataFrame(data)

    filename = f"search_results_{search_log.id}.xlsx"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    df.to_excel(filepath, index=False)

    current_user.credit_balance -= cost
    credit_log = CreditLog(
        user_id=current_user.id,
        amount=-cost,
        balance_after=current_user.credit_balance,
        reason=f'Download: {len(results)} row(s) — "{search_log.search_term}"'
    )
    db.session.add(credit_log)
    
    download_log = DataDownloadLog(
        user_id=current_user.id,
        search_log_id=search_log_id,
        download_format='excel',
        ip_address=get_client_ip()
    )
    db.session.add(download_log)
    
    search_log.data_downloaded = True
    db.session.commit()
    
    return send_file(filepath, as_attachment=True, download_name=filename)

@app.route('/customer/<int:customer_id>/detail')
@login_required
def customer_detail(customer_id):
    customer = CustomerData.query.get_or_404(customer_id)
    identity_summary, additional = _identity_summary_from_customer(customer)
    additional_display = _displayable_additional_data(additional)
    enrichments = []
    pending_delete_approval = None
    if current_user.role == 'admin':
        enrichments = CustomerEnrichment.query.filter_by(customer_id=customer.id).order_by(CustomerEnrichment.created_at.desc()).all()
        pending_delete_approval = CustomerDeleteApproval.query.filter_by(
            customer_id=customer.id,
            status='pending'
        ).order_by(CustomerDeleteApproval.created_at.desc()).first()
    household_lookup_skipped = current_user.role == 'admin' and not _household_lookup_enabled()
    household_candidates = _same_household_candidates(customer, identity_summary, limit=8) if current_user.role == 'admin' else []
    data = {
        'id': customer.id,
        'name': customer.name,
        'contact_number': customer.contact_number,
        'ic_number': customer.ic_number,
        'address': customer.address,
        'email': customer.email,
        'pending_delete_approval': bool(pending_delete_approval),
        'additional': additional_display,
        'identity_summary': identity_summary,
        'same_household': household_candidates,
        'same_household_skipped': household_lookup_skipped,
        'enrichments': [
            {
                'platform': enrichment.platform,
                'source_url': enrichment.source_url,
                'matched_company': enrichment.matched_company,
                'matched_title': enrichment.matched_title,
                'matched_location': enrichment.matched_location,
                'confidence_score': enrichment.confidence_score,
                'review_status': enrichment.review_status
            }
            for enrichment in enrichments
        ]
    }
    return jsonify(data)

@app.route('/customer/<int:customer_id>/view')
@login_required
def customer_view(customer_id):
    customer = CustomerData.query.get_or_404(customer_id)
    identity_summary, extra = _identity_summary_from_customer(customer)
    additional_display = _displayable_additional_data(extra)
    enrichments = []
    if current_user.role == 'admin':
        enrichments = CustomerEnrichment.query.filter_by(customer_id=customer.id).order_by(CustomerEnrichment.created_at.desc()).all()
    return render_template('customer_view.html',
                           customer=customer,
                           identity_summary=identity_summary,
                           additional=additional_display,
                           enrichments=enrichments)


@app.route('/agent/screenshot/<int:search_log_id>', methods=['POST'])
@login_required
def log_screenshot(search_log_id):
    if current_user.role != 'agent':
        return jsonify({'error': 'Unauthorized'}), 403
    
    search_log = SearchLog.query.get_or_404(search_log_id)
    
    if search_log.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    screenshot_log = ScreenshotLog(
        user_id=current_user.id,
        search_log_id=search_log_id,
        screenshot_path=f"screenshot_{search_log_id}_{datetime.now().timestamp()}.png",
        ip_address=get_client_ip()
    )
    db.session.add(screenshot_log)
    
    search_log.screenshot_taken = True
    db.session.commit()
    
    return jsonify({'success': True})

# ==================== HEALTH CHECK ENDPOINT ====================
@app.route('/health')
def health_check():
    """Health check endpoint for Render"""
    try:
        # Test database connection
        db.session.execute('SELECT 1')
        db_status = 'connected'
    except Exception as e:
        db_status = f'error: {str(e)}'
    
    return jsonify({
        'status': 'healthy',
        'database': db_status,
        'environment': 'production' if IS_RENDER else 'development',
        'upload_folder': app.config['UPLOAD_FOLDER']
    }), 200

# ==================== ERROR HANDLERS ====================
@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    app.logger.error(f'Server Error: {error}')
    return render_template('500.html'), 500

# ==================== INITIALIZATION ====================
def ensure_sqlite_column(table_name, column_name, column_definition):
    """Add a missing column to an existing SQLite table."""
    if not app.config['SQLALCHEMY_DATABASE_URI'].startswith('sqlite:'):
        return

    result = db.session.execute(text(f"PRAGMA table_info({table_name})"))
    existing_columns = {row[1] for row in result}

    if column_name in existing_columns:
        return

    db.session.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_definition}"))
    db.session.commit()
    app.logger.info(f"Added missing column {column_name} to {table_name}")
    print(f"[MIGRATION] Added missing column {column_name} to {table_name}")


def run_startup_migrations():
    """Patch older local SQLite databases to match the current models."""
    ensure_sqlite_column('users', 'failed_login_attempts', 'failed_login_attempts INTEGER DEFAULT 0')
    ensure_sqlite_column('users', 'locked_until', 'locked_until DATETIME')
    ensure_sqlite_column('login_logs', 'fingerprint', 'fingerprint VARCHAR(64)')
    ensure_sqlite_column('login_logs', 'device_category', 'device_category VARCHAR(50)')
    ensure_sqlite_column('customer_enrichment_jobs', 'debug_samples', "debug_samples TEXT DEFAULT '[]'")


def init_database():
    """Initialize database tables and create admin user"""
    with app.app_context():
        try:
            # Create all tables
            db.create_all()
            run_startup_migrations()
            app.logger.info("Database tables created successfully")
            print("[INIT] Database tables created successfully")
            
            # Create admin user if not exists
            admin = User.query.filter_by(username='admin').first()
            if not admin:
                admin = User(
                    username='admin',
                    email='admin@example.com',
                    role='admin',
                    full_name='System Administrator',
                    is_active=True
                )
                admin.set_password('admin123')
                db.session.add(admin)
                db.session.commit()
                app.logger.info("Created admin user")
                print("[INIT] Admin user created: username=admin, password=admin123")
            else:
                app.logger.info("Admin user already exists")
                print("[INIT] Admin user already exists")
                
        except Exception as e:
            app.logger.error(f"Database initialization error: {e}")
            print(f"[INIT ERROR] Database initialization error: {e}")

# Run database initialization
init_database()

# Pre-warm data_source cache in background so first search is instant
def _prewarm_ds_cache():
    try:
        with app.app_context():
            _get_data_sources()
            print(f'[CACHE] data_source cache warmed: {len(_ds_cache["values"])} unique sources')
    except Exception as e:
        print(f'[CACHE] prewarm failed: {e}')

threading.Thread(target=_prewarm_ds_cache, daemon=True).start()

# ==================== TELEGRAM POLLING ====================
def telegram_poll_loop():
    """Background thread: poll Telegram getUpdates and process approve/decline callbacks."""
    last_update_id = 0
    while True:
        try:
            with app.app_context():
                bot_token = SystemSettings.get('telegram_bot_token')
                if not bot_token:
                    time.sleep(5)
                    continue
                resp = http_req.get(
                    f'https://api.telegram.org/bot{bot_token}/getUpdates',
                    params={'offset': last_update_id + 1, 'timeout': 20, 'allowed_updates': ['callback_query', 'message']},
                    timeout=25
                )
                if not resp.ok:
                    time.sleep(5)
                    continue
                updates = resp.json().get('result', [])
                for update in updates:
                    last_update_id = update['update_id']

                    # Handle /register messages from agents
                    msg = update.get('message')
                    if msg and msg.get('text', '').startswith('/register'):
                        parts = msg['text'].strip().split()
                        tg_chat_id = str(msg['chat']['id'])
                        if len(parts) >= 2:
                            token_provided = parts[1].strip().upper()
                            agent = User.query.filter_by(tg_link_token=token_provided, role='agent').first()
                            if agent:
                                agent.telegram_chat_id = tg_chat_id
                                agent.tg_link_token = None
                                db.session.commit()
                                reply = f"✅ Your Telegram is now linked to account *{agent.full_name}*. New device logins will be sent here for your confirmation."
                            else:
                                reply = "❌ Invalid or expired registration code. Ask your admin for a new one."
                        else:
                            reply = "Usage: `/register YOUR_CODE`\nAsk your admin for your registration code."
                        try:
                            http_req.post(
                                f"https://api.telegram.org/bot{bot_token}/sendMessage",
                                json={"chat_id": tg_chat_id, "text": reply, "parse_mode": "Markdown"},
                                timeout=10
                            )
                        except Exception:
                            pass
                        continue

                    # Handle /panic and /unpanic commands from boss
                    if msg:
                        raw_cmd = msg.get('text', '').strip()
                        cmd = raw_cmd.split('@')[0].lower()
                        from_id = str(msg.get('from', {}).get('id', ''))
                        boss_id = str(SystemSettings.get('telegram_boss_chat_id', ''))
                        if from_id and boss_id and from_id == boss_id:
                            if cmd == '/panic':
                                _handle_panic_command(from_id, activate=True)
                                print(f"[TELEGRAM] Panic mode ACTIVATED by boss (chat_id={from_id})")
                                continue
                            elif cmd in ('/unpanic', '/calm'):
                                _handle_panic_command(from_id, activate=False)
                                print(f"[TELEGRAM] Panic mode DEACTIVATED by boss (chat_id={from_id})")
                                continue

                    cb = update.get('callback_query')
                    if not cb:
                        continue
                    cb_data = cb.get('data', '')
                    cb_id = cb['id']
                    if cb_data.startswith('custdelapprove_'):
                        token = cb_data[len('custdelapprove_'):]
                        process_customer_delete_callback(
                            'custdelapprove',
                            token,
                            callback_id=cb_id,
                            chat_id=cb['message']['chat']['id'],
                            message_id=cb['message']['message_id']
                        )
                    elif cb_data.startswith('custdeldecline_'):
                        token = cb_data[len('custdeldecline_'):]
                        process_customer_delete_callback(
                            'custdeldecline',
                            token,
                            callback_id=cb_id,
                            chat_id=cb['message']['chat']['id'],
                            message_id=cb['message']['message_id']
                        )
                    elif cb_data.startswith('approve_'):
                        token = cb_data[len('approve_'):]
                        approval = AdminLoginApproval.query.filter_by(approval_token=token, status='pending').first()
                        if approval:
                            approval.status = 'approved'
                            db.session.commit()
                            tg_answer_callback(cb_id, '✅ Login approved!')
                            tg_edit_message(cb['message']['chat']['id'],
                                            cb['message']['message_id'],
                                            f"✅ *Login Approved*\nAccess granted for {approval.user.full_name}.")
                            print(f"[TELEGRAM] Approved login for user {approval.user_id}")
                        else:
                            tg_answer_callback(cb_id, 'Request not found or already handled.')
                    elif cb_data.startswith('decline_'):
                        token = cb_data[len('decline_'):]
                        approval = AdminLoginApproval.query.filter_by(approval_token=token, status='pending').first()
                        if approval:
                            approval.status = 'declined'
                            db.session.commit()
                            tg_answer_callback(cb_id, '❌ Login declined.')
                            tg_edit_message(cb['message']['chat']['id'],
                                            cb['message']['message_id'],
                                            f"❌ *Login Declined*\nAccess denied for {approval.user.full_name}.")
                            print(f"[TELEGRAM] Declined login for user {approval.user_id}")
                        else:
                            tg_answer_callback(cb_id, 'Request not found or already handled.')
                    elif cb_data.startswith('lockbf_'):
                        token = cb_data[len('lockbf_'):]
                        alert = BruteForceAlert.query.filter_by(alert_token=token, status='pending').first()
                        if alert:
                            alert.status = 'locked'
                            SystemSettings.set('system_locked', '1')
                            SystemSettings.set('system_locked_reason',
                                f'System locked by boss via Telegram — suspicious login activity '
                                f'for {alert.user.full_name} from {alert.ip_address}')
                            db.session.commit()
                            tg_answer_callback(cb_id, '🔒 System locked!')
                            tg_edit_message(cb['message']['chat']['id'],
                                            cb['message']['message_id'],
                                            f"🔒 *System Locked*\nSuspicious activity for {alert.user.full_name}.")
                        else:
                            tg_answer_callback(cb_id, 'Already handled.')
                    elif cb_data.startswith('ignorebf_'):
                        token = cb_data[len('ignorebf_'):]
                        alert = BruteForceAlert.query.filter_by(alert_token=token, status='pending').first()
                        if alert:
                            alert.status = 'ignored'
                            db.session.commit()
                            tg_answer_callback(cb_id, '✅ Alert ignored.')
                            tg_edit_message(cb['message']['chat']['id'],
                                            cb['message']['message_id'],
                                            f"✅ *Alert Ignored* for {alert.user.full_name}.")
                        else:
                            tg_answer_callback(cb_id, 'Already handled.')
                    elif cb_data.startswith('unlocksys_'):
                        token = cb_data[len('unlocksys_'):]
                        stored_token = SystemSettings.get('unlock_reminder_token')
                        if stored_token == token:
                            SystemSettings.set('system_locked', '0')
                            SystemSettings.set('system_locked_reason', '')
                            SystemSettings.set('system_locked_since', '')
                            SystemSettings.set('unlock_reminder_token', '')
                            tg_answer_callback(cb_id, '🔓 System unlocked!')
                            tg_edit_message(cb['message']['chat']['id'],
                                            cb['message']['message_id'],
                                            f"🔓 *System Unlocked* at {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC")
                            print("[TELEGRAM] System unlocked by boss via Telegram")
                        else:
                            tg_answer_callback(cb_id, 'Invalid or expired token.')
                    elif cb_data.startswith('agentconfirm_'):
                        approval_token = cb_data[len('agentconfirm_'):]
                        device = AgentDevice.query.filter_by(approval_token=approval_token, status='pending').first()
                        if device:
                            device.status = 'approved'
                            device.approved_at = datetime.utcnow()
                            if device.pending_ip:
                                device.ip_address = device.pending_ip
                                device.pending_ip = None
                            db.session.commit()
                            tg_answer_callback(cb_id, '✅ Approved! Agent can now continue logging in.')
                            tg_edit_message(cb['message']['chat']['id'],
                                            cb['message']['message_id'],
                                            f"✅ *Approved*\nLogin approved for {device.user.full_name}.\nThe login will continue automatically.")
                            print(f"[TELEGRAM] Agent device/IP approved for user {device.user_id}")
                        else:
                            tg_answer_callback(cb_id, 'Request not found or already handled.')
                    elif cb_data.startswith('agentdeny_'):
                        approval_token = cb_data[len('agentdeny_'):]
                        device = AgentDevice.query.filter_by(approval_token=approval_token, status='pending').first()
                        if device:
                            device.status = 'blocked'
                            db.session.commit()
                            tg_answer_callback(cb_id, '❌ Device denied and blocked.')
                            tg_edit_message(cb['message']['chat']['id'],
                                            cb['message']['message_id'],
                                            f"❌ *Device Denied*\nLogin attempt blocked for {device.user.full_name}.\nThe device has been blocked.")
                            print(f"[TELEGRAM] Agent device denied for user {device.user_id}")
                        else:
                            tg_answer_callback(cb_id, 'Request not found or already handled.')
                    elif cb_data.startswith('keeplocked_'):
                        SystemSettings.set('unlock_reminder_token', '')
                        SystemSettings.set('system_locked_since', datetime.utcnow().isoformat())
                        tg_answer_callback(cb_id, '🔒 System remains locked. You will be reminded again in 30 minutes.')
                        tg_edit_message(cb['message']['chat']['id'],
                                        cb['message']['message_id'],
                                        '🔒 *System remains locked.*\nYou will receive another reminder in 30 minutes.')
        except Exception as e:
            print(f"[TELEGRAM POLL] Error: {e}")
            time.sleep(5)

def tg_send_unlock_reminder():
    """Send Telegram message to boss after system has been locked for 30 minutes."""
    token = SystemSettings.get('telegram_bot_token')
    chat_id = SystemSettings.get('telegram_boss_chat_id')
    if not token or not chat_id:
        return
    try:
        unlock_token = secrets.token_hex(16)
        SystemSettings.set('unlock_reminder_token', unlock_token)
        reason = SystemSettings.get('system_locked_reason', 'Security lockout')
        text = (
            f"⏰ *System Still Locked*\n\n"
            f"The system has been locked for *30 minutes*.\n\n"
            f"📋 Reason: {reason}\n\n"
            f"Do you want to unlock the system now?"
        )
        keyboard = {"inline_keyboard": [[
            {"text": "🔓 Unlock System", "callback_data": f"unlocksys_{unlock_token}"},
            {"text": "🔒 Keep Locked", "callback_data": f"keeplocked_{unlock_token}"}
        ]]}
        http_req.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": chat_id, "text": text,
                  "parse_mode": "Markdown", "reply_markup": keyboard},
            timeout=10
        )
        print("[TELEGRAM] Unlock reminder sent to boss")
    except Exception as e:
        app.logger.error(f"Unlock reminder error: {e}")

def auto_unlock_checker():
    """Background thread: after 30 mins of lockout, send unlock reminder to boss."""
    reminder_sent_at = None
    while True:
        time.sleep(60)
        try:
            with app.app_context():
                if SystemSettings.get('system_locked', '0') != '1':
                    reminder_sent_at = None
                    continue
                locked_since_str = SystemSettings.get('system_locked_since')
                if not locked_since_str:
                    SystemSettings.set('system_locked_since', datetime.utcnow().isoformat())
                    continue
                locked_since = datetime.fromisoformat(locked_since_str)
                elapsed = (datetime.utcnow() - locked_since).total_seconds()
                if elapsed >= 1800 and reminder_sent_at is None:
                    tg_send_unlock_reminder()
                    reminder_sent_at = datetime.utcnow()
        except Exception as e:
            print(f"[AUTO UNLOCK] Error: {e}")

def enrichment_job_worker():
    """Background worker for enrichment crawl jobs."""
    while True:
        time.sleep(3)
        try:
            with app.app_context():
                job = CustomerEnrichmentJob.query.filter_by(status='queued').order_by(CustomerEnrichmentJob.created_at.asc()).first()
                if not job:
                    continue

                job.status = 'running'
                job.started_at = datetime.utcnow()
                job.message = 'Crawl is running.'
                db.session.commit()

                customer = db.session.get(CustomerData, job.customer_id)
                if not customer:
                    job.status = 'failed'
                    job.completed_at = datetime.utcnow()
                    job.message = 'Customer record not found.'
                    job.error_log = json.dumps(['Customer record not found.'])
                    db.session.commit()
                    continue

                result = auto_crawl_customer_enrichment(customer, job.requested_by or 0)
                job.created_count = result.get('created', 0)
                job.checked_count = result.get('checked', 0)
                job.providers_used = json.dumps(result.get('providers_used', []))
                job.error_log = json.dumps(result.get('errors', []))
                job.debug_samples = json.dumps(result.get('debug_samples', []))
                job.completed_at = datetime.utcnow()

                if result.get('errors') and result.get('created', 0) == 0:
                    job.status = 'failed'
                    job.message = 'Crawl finished with errors and no saved matches.'
                else:
                    job.status = 'completed'
                    if result.get('created', 0) > 0:
                        job.message = f"Saved {result.get('created', 0)} candidate result(s)."
                    else:
                        job.message = 'Crawl finished with no new candidate matches.'

                db.session.commit()
        except Exception as e:
            print(f"[ENRICHMENT WORKER] Error: {e}")

# Start Telegram polling only in production or main process
if not app.debug or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
    poll_thread = threading.Thread(target=telegram_poll_loop, daemon=True)
    poll_thread.start()
    print("[TELEGRAM] Polling thread started")
    unlock_thread = threading.Thread(target=auto_unlock_checker, daemon=True)
    unlock_thread.start()
    print("[AUTO UNLOCK] Checker thread started")
    enrichment_thread = threading.Thread(target=enrichment_job_worker, daemon=True)
    enrichment_thread.start()
    print("[ENRICHMENT WORKER] Started")

# ==================== DB DUPLICATE SCAN ====================

def _normalize_scan_ic(v):
    return re.sub(r'[^0-9]', '', str(v).strip()) if v else ''

def _normalize_scan_phone(v):
    raw = re.sub(r'[^0-9]', '', str(v).strip()) if v else ''
    # Strip leading country code 60 → treat 601x and 1x as the same
    if raw.startswith('60') and len(raw) > 9:
        raw = raw[2:]
    return raw

def _normalize_scan_name(v):
    return re.sub(r'\s+', ' ', str(v).strip()).casefold() if v else ''


def _normalize_scan_address(v):
    value = clean_data_value(v)
    if not value:
        return ''
    value = value.upper()
    replacements = {
        r'\bJLN\b': 'JALAN',
        r'\bJAL\b': 'JALAN',
        r'\bTMN\b': 'TAMAN',
        r'\bKG\b': 'KAMPUNG',
    }
    for pattern, repl in replacements.items():
        value = re.sub(pattern, repl, value)
    value = re.sub(r'[^A-Z0-9]+', ' ', value)
    return re.sub(r'\s+', ' ', value).strip()


def _name_tokens_for_compare(value):
    stop_words = {'BIN', 'BINTI', 'A/L', 'A/P', 'BT', 'BTE', 'MOHD', 'MUHAMMAD', 'MD'}
    tokens = [tok for tok in re.split(r'[^A-Z0-9]+', clean_data_value(value).upper()) if len(tok) >= 2]
    return {tok for tok in tokens if tok not in stop_words}


def _token_overlap_score(left, right):
    if not left or not right:
        return 0.0
    intersection = len(left & right)
    union = len(left | right)
    return intersection / union if union else 0.0


def _build_duplicate_group_signal(group_type, customers):
    signal = {
        'is_suspect': False,
        'severity': 'normal',
        'reasons': [],
        'summary': '',
        'unique_names': 0,
        'unique_phones': 0,
        'unique_addresses': 0,
    }
    if group_type != 'ic' or not customers or len(customers) < 2:
        return signal

    normalized_names = []
    normalized_name_set = []
    unique_phones = sorted({_normalize_scan_phone(c.contact_number) for c in customers if _normalize_scan_phone(c.contact_number)})
    unique_addresses = sorted({_normalize_scan_address(c.address) for c in customers if _normalize_scan_address(c.address)})

    for customer in customers:
        tokens = _name_tokens_for_compare(customer.name)
        if tokens and tokens not in normalized_name_set:
            normalized_name_set.append(tokens)
            normalized_names.append(clean_data_value(customer.name))

    signal['unique_names'] = len(normalized_name_set)
    signal['unique_phones'] = len(unique_phones)
    signal['unique_addresses'] = len(unique_addresses)

    lowest_overlap = 1.0
    for i in range(len(normalized_name_set)):
        for j in range(i + 1, len(normalized_name_set)):
            lowest_overlap = min(lowest_overlap, _token_overlap_score(normalized_name_set[i], normalized_name_set[j]))

    if signal['unique_names'] >= 2 and lowest_overlap <= 0.20:
        signal['reasons'].append('Names are very different for the same IC.')
    elif signal['unique_names'] >= 2 and lowest_overlap <= 0.40:
        signal['reasons'].append('Names only partially match for the same IC.')

    if signal['unique_phones'] >= 2:
        signal['reasons'].append('Phone numbers do not match.')

    if signal['unique_addresses'] >= 2:
        signal['reasons'].append('Addresses do not match.')

    if signal['unique_names'] >= 2 and (signal['unique_phones'] >= 2 or signal['unique_addresses'] >= 2):
        signal['is_suspect'] = True

    if signal['is_suspect']:
        signal['severity'] = 'high' if len(signal['reasons']) >= 3 or lowest_overlap <= 0.20 else 'medium'
        signal['summary'] = 'Possible fake or misused IC: the same IC appears with conflicting identity details.'
    elif signal['reasons']:
        signal['severity'] = 'watch'
        signal['summary'] = 'This IC group has mismatched details and may need manual review.'

    return signal


def _append_suspect_ic_flag(job_id, group_type, group_key, customers, signal, flagged_by):
    os.makedirs(os.path.dirname(SUSPECT_IC_FLAGS_FILE), exist_ok=True)
    entry = {
        'flagged_at': datetime.utcnow().isoformat(),
        'job_id': job_id,
        'group_type': group_type,
        'group_key': group_key,
        'flagged_by': {
            'id': getattr(flagged_by, 'id', None),
            'username': getattr(flagged_by, 'username', ''),
            'full_name': getattr(flagged_by, 'full_name', ''),
        },
        'signal': signal,
        'customers': [
            {
                'id': c.id,
                'name': c.name,
                'contact_number': c.contact_number,
                'ic_number': c.ic_number,
                'address': c.address,
                'email': c.email,
                'upload_id': c.upload_id,
                'updated_at': c.updated_at.isoformat() if getattr(c, 'updated_at', None) else None,
            }
            for c in customers
        ]
    }
    with open(SUSPECT_IC_FLAGS_FILE, 'a', encoding='utf-8') as handle:
        handle.write(json.dumps(entry, ensure_ascii=False) + '\n')


def _count_suspect_ic_flags():
    if not os.path.exists(SUSPECT_IC_FLAGS_FILE):
        return 0
    try:
        with open(SUSPECT_IC_FLAGS_FILE, 'r', encoding='utf-8') as handle:
            return sum(1 for line in handle if line.strip())
    except Exception:
        return 0


def _get_scan_groups_page(job_id, filter_type=None, page=1, per_page=15):
    """Load scan groups from the dedicated groups file, filtered + paginated."""
    groups_path = os.path.join(_JOB_DIR, f'{job_id}_groups.json')
    resolved_keys = _get_resolved_scan_keys(job_id)
    try:
        with open(groups_path, 'r') as f:
            groups = json.load(f)
    except Exception:
        return [], 0
    if resolved_keys:
        groups = [g for g in groups if f"{g['type']}::{g['key']}" not in resolved_keys]
    if filter_type:
        groups = [g for g in groups if g['type'] == filter_type]
    total = len(groups)
    start = (page - 1) * per_page
    return groups[start:start + per_page], total


def _resolved_scan_groups_file(job_id):
    return os.path.join(_JOB_DIR, f'{job_id}_resolved_groups.jsonl')


def _get_resolved_scan_keys(job_id):
    path = _resolved_scan_groups_file(job_id)
    if not os.path.exists(path):
        return set()
    try:
        with open(path, 'r', encoding='utf-8') as handle:
            return {json.loads(line)['token'] for line in handle if line.strip()}
    except Exception as e:
        app.logger.error(f'[DB SCAN] _get_resolved_scan_keys error: {e}')
        return set()


def _mark_scan_group_resolved(job_id, group_type, group_key, record_count=0):
    token = f'{group_type}::{group_key}'
    path = _resolved_scan_groups_file(job_id)
    resolved = _get_resolved_scan_keys(job_id)
    if token not in resolved:
        with open(path, 'a', encoding='utf-8') as handle:
            handle.write(json.dumps({
                'token': token,
                'group_type': group_type,
                'group_key': group_key,
                'resolved_at': datetime.utcnow().isoformat(),
            }) + '\n')

    job = _get_job(job_id)
    summary = dict(job.get('summary') or {})
    if not summary:
        return

    group_field = f'{group_type}_groups'
    record_field = f'{group_type}_records'
    if group_field in summary:
        summary[group_field] = max(0, int(summary.get(group_field, 0)) - 1)
    if record_field in summary and record_count:
        summary[record_field] = max(0, int(summary.get(record_field, 0)) - int(record_count))
    _set_job(job_id, summary=summary)


def run_db_duplicate_scan(job_id):
    """
    SQL-based duplicate scanner. Uses GROUP BY … HAVING COUNT(*) > 1 so the
    database finds duplicates without loading all rows into Python memory.
    Safe for databases of any size on low-RAM servers.
    """
    with app.app_context():
        try:
            _set_job(job_id, status='running', message='Counting records…', progress=0)
            total = CustomerData.query.count()

            if total == 0:
                _set_job(job_id, status='done', progress=100,
                         message='Database is empty — nothing to scan.',
                         summary={'total_records': 0,
                                  'ic_groups': 0, 'ic_records': 0,
                                  'phone_groups': 0, 'phone_records': 0,
                                  'name_groups': 0, 'name_records': 0})
                return

            _set_job(job_id, total=total,
                     message=f'Scanning {total:,} records via SQL GROUP BY…')

            is_pg = app.config['SQLALCHEMY_DATABASE_URI'].startswith('postgresql')
            groups = []

            # ── IC duplicates (strip dashes/spaces, min 6 digits) ──────────────
            _set_job(job_id, progress=10, message='Finding IC number duplicates…')
            if is_pg:
                ic_rows = db.session.execute(text("""
                    SELECT regexp_replace(ic_number, '[^0-9]', '', 'g') AS norm_key,
                           array_agg(id ORDER BY id)                    AS ids
                    FROM   customer_data
                    WHERE  ic_number IS NOT NULL AND ic_number <> ''
                    GROUP  BY norm_key
                    HAVING COUNT(*) > 1
                       AND length(regexp_replace(ic_number, '[^0-9]', '', 'g')) >= 6
                    ORDER  BY COUNT(*) DESC
                """)).fetchall()
                for row in ic_rows:
                    groups.append({'type': 'ic', 'key': row.norm_key, 'ids': list(row.ids)})
            else:
                ic_rows = db.session.execute(text("""
                    SELECT replace(replace(ic_number, '-', ''), ' ', '') AS norm_key,
                           group_concat(id)                               AS ids,
                           count(*)                                       AS cnt
                    FROM   customer_data
                    WHERE  ic_number IS NOT NULL AND ic_number <> ''
                    GROUP  BY norm_key
                    HAVING cnt > 1 AND length(norm_key) >= 6
                    ORDER  BY cnt DESC
                """)).fetchall()
                for row in ic_rows:
                    ids = [int(i) for i in row.ids.split(',') if i.strip().isdigit()]
                    groups.append({'type': 'ic', 'key': row.norm_key, 'ids': ids})

            ic_covered   = {cid for g in groups if g['type'] == 'ic'    for cid in g['ids']}

            # ── Phone duplicates (strip non-digits, normalise leading 60) ──────
            _set_job(job_id, progress=40, message='Finding phone number duplicates…')
            if is_pg:
                ph_rows = db.session.execute(text("""
                    WITH stripped AS (
                        SELECT id,
                               CASE WHEN regexp_replace(contact_number,'[^0-9]','','g') LIKE '60%'
                                     AND length(regexp_replace(contact_number,'[^0-9]','','g')) > 9
                                    THEN substring(regexp_replace(contact_number,'[^0-9]','','g') FROM 3)
                                    ELSE regexp_replace(contact_number,'[^0-9]','','g')
                               END AS norm_key
                        FROM   customer_data
                        WHERE  contact_number IS NOT NULL AND contact_number <> ''
                    )
                    SELECT norm_key, array_agg(id ORDER BY id) AS ids
                    FROM   stripped
                    WHERE  length(norm_key) >= 7
                    GROUP  BY norm_key
                    HAVING COUNT(*) > 1
                    ORDER  BY COUNT(*) DESC
                """)).fetchall()
                for row in ph_rows:
                    ids = list(row.ids)
                    if not all(cid in ic_covered for cid in ids):
                        groups.append({'type': 'phone', 'key': row.norm_key, 'ids': ids})
            else:
                ph_rows = db.session.execute(text("""
                    WITH stripped AS (
                        SELECT id,
                               CASE WHEN SUBSTR(REPLACE(REPLACE(contact_number,'-',''),' ',''),1,2) = '60'
                                     AND LENGTH(REPLACE(REPLACE(contact_number,'-',''),' ','')) > 9
                                    THEN SUBSTR(REPLACE(REPLACE(contact_number,'-',''),' ',''),3)
                                    ELSE REPLACE(REPLACE(contact_number,'-',''),' ','')
                               END AS norm_key
                        FROM   customer_data
                        WHERE  contact_number IS NOT NULL AND contact_number <> ''
                    )
                    SELECT norm_key, group_concat(id) AS ids, count(*) AS cnt
                    FROM   stripped
                    WHERE  length(norm_key) >= 7
                    GROUP  BY norm_key
                    HAVING cnt > 1
                    ORDER  BY cnt DESC
                """)).fetchall()
                for row in ph_rows:
                    ids = [int(i) for i in row.ids.split(',') if i.strip().isdigit()]
                    if not all(cid in ic_covered for cid in ids):
                        groups.append({'type': 'phone', 'key': row.norm_key, 'ids': ids})

            phone_covered = {cid for g in groups if g['type'] == 'phone' for cid in g['ids']}

            # ── Name duplicates (case-insensitive, skip IC/phone-already-covered) ─
            _set_job(job_id, progress=70, message='Finding name duplicates…')
            if is_pg:
                nm_rows = db.session.execute(text("""
                    SELECT lower(trim(name)) AS norm_key,
                           array_agg(id ORDER BY id) AS ids
                    FROM   customer_data
                    WHERE  name IS NOT NULL AND name <> ''
                    GROUP  BY norm_key
                    HAVING COUNT(*) > 1 AND length(lower(trim(name))) >= 3
                    ORDER  BY COUNT(*) DESC
                """)).fetchall()
                for row in nm_rows:
                    ids = list(row.ids)
                    uncovered = [cid for cid in ids
                                 if cid not in ic_covered and cid not in phone_covered]
                    if len(uncovered) > 1:
                        groups.append({'type': 'name', 'key': row.norm_key, 'ids': ids})
            else:
                nm_rows = db.session.execute(text("""
                    SELECT LOWER(TRIM(name)) AS norm_key,
                           group_concat(id)  AS ids,
                           count(*)          AS cnt
                    FROM   customer_data
                    WHERE  name IS NOT NULL AND name <> ''
                    GROUP  BY norm_key
                    HAVING cnt > 1 AND length(norm_key) >= 3
                    ORDER  BY cnt DESC
                """)).fetchall()
                for row in nm_rows:
                    ids = [int(i) for i in row.ids.split(',') if i.strip().isdigit()]
                    uncovered = [cid for cid in ids
                                 if cid not in ic_covered and cid not in phone_covered]
                    if len(uncovered) > 1:
                        groups.append({'type': 'name', 'key': row.norm_key, 'ids': ids})

            _set_job(job_id, progress=95, message='Saving results…')

            groups_path = os.path.join(_JOB_DIR, f'{job_id}_groups.json')
            with open(groups_path, 'w') as gf:
                json.dump(groups, gf)

            ic_groups    = [g for g in groups if g['type'] == 'ic']
            phone_groups = [g for g in groups if g['type'] == 'phone']
            name_groups  = [g for g in groups if g['type'] == 'name']

            summary = {
                'total_records': total,
                'ic_groups':     len(ic_groups),
                'ic_records':    sum(len(g['ids']) for g in ic_groups),
                'phone_groups':  len(phone_groups),
                'phone_records': sum(len(g['ids']) for g in phone_groups),
                'name_groups':   len(name_groups),
                'name_records':  sum(len(g['ids']) for g in name_groups),
            }

            total_dup_groups = len(ic_groups) + len(phone_groups) + len(name_groups)
            _set_job(job_id, status='done', progress=100,
                     message=(f'Scan complete. '
                               f'{len(ic_groups)} IC groups, '
                               f'{len(phone_groups)} phone groups, '
                               f'{len(name_groups)} name groups — '
                               f'{total_dup_groups} total.'),
                     summary=summary)

        except Exception as e:
            _set_job(job_id, status='error', message=str(e))
            app.logger.error(f'[DB SCAN] Error: {e}')


# ==================== DB SCAN ROUTES ====================

@app.route('/admin/db-scan', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_db_scan():
    if request.method == 'POST':
        job_id = str(uuid.uuid4())
        _set_job(job_id, status='running', message='Initializing…', progress=0)
        t = threading.Thread(target=run_db_duplicate_scan, args=(job_id,), daemon=True)
        t.start()
        return redirect(url_for('admin_db_scan_progress', job_id=job_id))
    return render_template('db_scan.html')


@app.route('/admin/db-scan/<job_id>/progress')
@login_required
@admin_required
def admin_db_scan_progress(job_id):
    return render_template('db_scan_progress.html', job_id=job_id)


@app.route('/admin/db-scan/<job_id>/status')
@login_required
@admin_required
def admin_db_scan_status(job_id):
    job = _get_job(job_id)
    if not job:
        return jsonify({'status': 'not_found'})
    # Strip bulky 'groups' key — it lives in its own file now
    return jsonify({k: v for k, v in job.items() if k != 'groups'})


@app.route('/admin/db-scan/<job_id>/report')
@login_required
@admin_required
def admin_db_scan_report(job_id):
    job = _get_job(job_id)
    if not job or job.get('status') != 'done':
        flash('Scan not complete or not found.', 'warning')
        return redirect(url_for('admin_db_scan'))

    filter_type = request.args.get('type', 'ic')
    page        = request.args.get('page', 1, type=int)
    per_page    = 15

    page_groups, total_groups = _get_scan_groups_page(
        job_id, filter_type=filter_type, page=page, per_page=per_page)

    all_ids   = [cid for g in page_groups for cid in g['ids']]
    customers = {c.id: c for c in
                 CustomerData.query.filter(CustomerData.id.in_(all_ids)).all()}

    enriched_groups = []
    for g in page_groups:
        group_customers = [customers[cid] for cid in g['ids'] if cid in customers]
        enriched_groups.append({
            'type': g['type'],
            'key':  g['key'],
            'ids':  g['ids'],
            'customers': group_customers,
            'signal': _build_duplicate_group_signal(g['type'], group_customers),
        })

    total_pages = max(1, (total_groups + per_page - 1) // per_page)

    return render_template('db_scan_report.html',
                           job_id=job_id,
                           summary=job.get('summary', {}),
                           filter_type=filter_type,
                           groups=enriched_groups,
                           total_groups=total_groups,
                           page=page,
                           per_page=per_page,
                           total_pages=total_pages,
                           suspect_flag_count=_count_suspect_ic_flags(),
                           suspect_flag_file=os.path.basename(SUSPECT_IC_FLAGS_FILE))


@app.route('/admin/db-scan/flagged-cases/download')
@login_required
@admin_required
def admin_db_scan_flagged_cases_download():
    if not os.path.exists(SUSPECT_IC_FLAGS_FILE):
        flash('No suspect IC review file exists yet.', 'warning')
        return redirect(url_for('admin_db_scan'))
    return send_file(
        SUSPECT_IC_FLAGS_FILE,
        as_attachment=True,
        download_name=os.path.basename(SUSPECT_IC_FLAGS_FILE),
        mimetype='application/x-ndjson'
    )


@app.route('/admin/db-scan/<job_id>/resolve', methods=['POST'])
@login_required
@admin_required
def admin_db_scan_resolve(job_id):
    data       = request.get_json(silent=True) or {}
    action     = data.get('action')      # keep_first | keep_last | keep_selected | merge | flag_suspect | skip
    group_type = data.get('group_type')
    group_key  = data.get('group_key')
    keep_id    = data.get('keep_id')     # used by keep_selected
    posted_ids = data.get('ids') or []

    if not all([action, group_type, group_key is not None]):
        return jsonify({'ok': False, 'message': 'Missing parameters'})

    ids = [int(i) for i in posted_ids if str(i).isdigit()]
    if not ids:
        groups_path = os.path.join(_JOB_DIR, f'{job_id}_groups.json')
        try:
            with open(groups_path, 'r') as f:
                all_groups = json.load(f)
        except Exception:
            return jsonify({'ok': False, 'message': 'Groups file not found'})

        target = next((g for g in all_groups
                       if g['type'] == group_type and g['key'] == group_key), None)
        if not target:
            return jsonify({'ok': False, 'message': 'Group not found (may already be resolved)'})
        ids = target['ids']
    customers = (CustomerData.query
                 .filter(CustomerData.id.in_(ids))
                 .order_by(CustomerData.id.asc()).all())

    if not customers:
        _mark_scan_group_resolved(job_id, group_type, group_key, record_count=len(ids))
        return jsonify({'ok': True, 'message': 'Records already deleted.'})

    try:
        if action == 'skip':
            _mark_scan_group_resolved(job_id, group_type, group_key, record_count=len(customers))
            return jsonify({'ok': True, 'message': 'Group skipped — no changes made.'})

        elif action == 'flag_suspect':
            signal = _build_duplicate_group_signal(group_type, customers)
            if not signal.get('summary'):
                signal['summary'] = 'Flagged manually by admin for later investigation.'
            if not signal.get('reasons'):
                signal['reasons'] = ['Flagged manually by admin.']
            if not signal.get('severity') or signal['severity'] == 'normal':
                signal['severity'] = 'manual'
            _append_suspect_ic_flag(job_id, group_type, group_key, customers, signal, current_user)
            _mark_scan_group_resolved(job_id, group_type, group_key, record_count=len(customers))
            return jsonify({
                'ok': True,
                'message': f'Suspect case saved to {os.path.basename(SUSPECT_IC_FLAGS_FILE)} for admin review.'
            })

        elif action == 'keep_first':
            keep      = customers[0]
            to_delete = customers[1:]
            merge_customer_records(keep, to_delete, merge_reason=f'db_scan_{group_type}_keep_first')
            for c in to_delete:
                delete_customer_with_related_records(c)
            db.session.commit()
            msg = f'Kept ID {keep.id} ({keep.name or "-"}), merged facts from {len(to_delete)} record(s).'

        elif action == 'keep_last':
            keep      = customers[-1]
            to_delete = customers[:-1]
            merge_customer_records(keep, to_delete, merge_reason=f'db_scan_{group_type}_keep_last')
            for c in to_delete:
                delete_customer_with_related_records(c)
            db.session.commit()
            msg = f'Kept ID {keep.id} ({keep.name or "-"}), merged facts from {len(to_delete)} record(s).'

        elif action == 'keep_selected':
            if not keep_id:
                return jsonify({'ok': False, 'message': 'No record selected to keep'})
            keep = next((c for c in customers if c.id == int(keep_id)), None)
            if not keep:
                return jsonify({'ok': False, 'message': 'Selected record not found'})
            to_delete = [c for c in customers if c.id != keep.id]
            merge_customer_records(keep, to_delete, merge_reason=f'db_scan_{group_type}_keep_selected')
            for c in to_delete:
                delete_customer_with_related_records(c)
            db.session.commit()
            msg = f'Kept ID {keep.id} ({keep.name or "-"}), merged facts from {len(to_delete)} record(s).'

        elif action == 'merge':
            primary = _choose_primary_customer(customers)
            others  = [c for c in customers if c.id != primary.id]
            merge_customer_records(primary, others, merge_reason=f'db_scan_{group_type}_merge')
            for c in others:
                delete_customer_with_related_records(c)
            db.session.commit()
            msg = f'Merged {len(customers)} records into ID {primary.id} ({primary.name or "-"}) and preserved unique names, phones, emails, and addresses.'

        else:
            return jsonify({'ok': False, 'message': f'Unknown action: {action}'})

        _mark_scan_group_resolved(job_id, group_type, group_key, record_count=len(customers))
        return jsonify({'ok': True, 'message': msg})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f'[DB SCAN RESOLVE] {e}')
        return jsonify({'ok': False, 'message': str(e)})


# ==================== CUSTOMER MAP ====================

@app.route('/admin/customer-map')
@login_required
@admin_required
def admin_customer_map():
    return render_template('admin_customer_map.html')


@app.route('/admin/customer-map/search', methods=['POST'])
@login_required
@admin_required
def admin_customer_map_search():
    data = request.get_json(silent=True) or {}
    search_type = data.get('search_type', 'address')
    search_term = (data.get('search_term') or '').strip()

    if not search_term:
        return jsonify({'ok': False, 'message': 'Please enter a search term.', 'customers': []})

    results, meta = perform_customer_search(search_type, search_term, requested_limit=200)

    customers = []
    for c in results:
        customers.append({
            'id': c.id,
            'name': c.name or '',
            'phone': c.contact_number or '',
            'ic': c.ic_number or '',
            'address': c.address or '',
            'source': c.data_source or '',
        })

    return jsonify({
        'ok': meta.get('ok', True),
        'message': meta.get('message', ''),
        'customers': customers,
        'total': len(customers),
    })


# ==================== JPPH DATABASE SEARCH ====================

def _jpph_db_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'jpph.db')


def _jpph_fts_query(term, column=None):
    words = [w for w in re.sub(r'[^\w]', ' ', term).split() if len(w) >= 2]
    if not words:
        return None
    quoted = ' '.join(f'"{w}"' for w in words)
    if column:
        return f'{{{column}}} : {quoted}'
    return quoted


def search_jpph(search_type, term, limit=500):
    db_path = _jpph_db_path()
    if not os.path.exists(db_path):
        return [], {'ok': False, 'message': 'JPPH database not yet imported. Run import_jpph.py on the server first.'}

    term = (term or '').strip()
    if len(term) < 2:
        return [], {'ok': False, 'message': 'Enter at least 2 characters.'}

    col_map = {
        'address':   'address',
        'scheme':    'scheme',
        'area':      None,
        'mukim':     'mukim_town',
        'daerah':    'district',
        'vendor':    'vendor',
        'purchaser': 'purchaser',
    }

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        if search_type == 'area':
            words = [w for w in re.sub(r'[^\w]', ' ', term).split() if len(w) >= 2]
            if not words:
                return [], {'ok': False, 'message': 'Invalid search term.'}
            quoted = ' '.join(f'"{w}"' for w in words)
            fts_q = f'{{mukim_town district}} : {quoted}'
        else:
            col = col_map.get(search_type)
            fts_q = _jpph_fts_query(term, column=col)

        if not fts_q:
            return [], {'ok': False, 'message': 'Invalid search term.'}

        rows = conn.execute(
            "SELECT * FROM jpph_fts WHERE jpph_fts MATCH ? ORDER BY rank LIMIT ?",
            (fts_q, limit)
        ).fetchall()

        results = [dict(r) for r in rows]
        truncated = len(results) == limit
        return results, {
            'ok': True,
            'count': len(results),
            'truncated': truncated,
            'limit': limit,
            'message': f'Found {len(results)} transaction(s)' + (f' (showing first {limit})' if truncated else ''),
        }
    except Exception as e:
        return [], {'ok': False, 'message': f'Search error: {str(e)}'}
    finally:
        conn.close()


@app.route('/admin/transacted-data', methods=['GET'])
@login_required
@admin_required
def admin_transacted_data():
    return render_template('admin_transacted_data.html')


@app.route('/admin/transacted-data/search', methods=['POST'])
@login_required
@admin_required
def admin_transacted_data_search():
    search_type = request.form.get('search_type', 'keyword')
    term = request.form.get('term', '').strip()
    results, meta = search_jpph(search_type, term)
    return jsonify({'results': results, 'meta': meta})


@app.route('/agent/transacted-data', methods=['GET'])
@login_required
def agent_transacted_data():
    if current_user.role != 'agent':
        return redirect(url_for('admin_dashboard'))
    return render_template('agent_transacted_data.html')


@app.route('/agent/transacted-data/search', methods=['POST'])
@login_required
def agent_transacted_data_search():
    if current_user.role != 'agent':
        return jsonify({'ok': False, 'message': 'Unauthorized'}), 403
    search_type = request.form.get('search_type', 'keyword')
    term = request.form.get('term', '').strip()
    results, meta = search_jpph(search_type, term)
    return jsonify({'results': results, 'meta': meta})


# ==================== MYS DATABASE SEARCH ====================

def _mys_db_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'mys.db')


# 120-second in-memory cache for MYS searches — MYS data is static so TTL can be longer
_mys_cache = {}

def _mys_cache_get(search_type, term):
    key = (search_type, term)
    e = _mys_cache.get(key)
    return e[1] if (e and time.time() - e[0] < 120) else None

def _mys_cache_set(search_type, term, results, meta):
    _mys_cache[(search_type, term)] = (time.time(), (results, meta))
    if len(_mys_cache) > 300:
        cutoff = time.time() - 120
        for k in list(_mys_cache):
            if _mys_cache.get(k, (0,))[0] < cutoff:
                _mys_cache.pop(k, None)

def _mys_connect():
    db_path = _mys_db_path()
    conn = sqlite3.connect(db_path, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    # mmap_size maps up to 2GB of the DB file into virtual memory —
    # avoids repeated read() syscalls for hot index pages
    conn.execute('PRAGMA mmap_size=2147483648')
    conn.execute('PRAGMA cache_size=-131072')   # 128 MB page cache
    conn.execute('PRAGMA temp_store=MEMORY')
    conn.execute('PRAGMA journal_mode=WAL')
    return conn


def search_mys(search_type, term, limit=200):
    db_path = _mys_db_path()
    if not os.path.exists(db_path):
        return [], {'ok': False, 'message': 'MYS database not yet imported. Run import_mys.py on the server first.'}

    term = (term or '').strip()
    if len(term) < 2:
        return [], {'ok': False, 'message': 'Enter at least 2 characters.'}

    # Return cached result if available
    cached = _mys_cache_get(search_type, term)
    if cached:
        return cached

    conn = _mys_connect()
    try:
        if search_type == 'ic':
            rows = conn.execute(
                'SELECT * FROM mys_data WHERE ic >= ? AND ic < ? LIMIT ?',
                (term, term + '\xff', limit)
            ).fetchall()

        elif search_type == 'phone':
            digits = re.sub(r'\D', '', term)
            if not digits:
                return [], {'ok': False, 'message': 'Enter a valid phone number.'}
            rows = conn.execute(
                'SELECT * FROM mys_data WHERE tel >= ? AND tel < ? LIMIT ?',
                (digits, digits + '\xff', limit)
            ).fetchall()

        elif search_type == 'name':
            words = [w for w in re.sub(r'[^\w]', ' ', term.upper()).split() if len(w) >= 2]
            if not words:
                return [], {'ok': False, 'message': 'Invalid search term.'}
            fts_q = ' '.join(f'"{w}"*' for w in words)
            rows = conn.execute(
                'SELECT mys_data.* FROM mys_nama_fts JOIN mys_data ON mys_data.id = mys_nama_fts.rowid'
                ' WHERE mys_nama_fts MATCH ? ORDER BY rank LIMIT ?',
                (fts_q, limit)
            ).fetchall()

        elif search_type == 'state':
            # Prefix match on indexed negeri column — avoids full table scan
            rows = conn.execute(
                'SELECT * FROM mys_data WHERE negeri >= ? AND negeri < ? LIMIT ?',
                (term.upper(), term.upper() + '\xff', limit)
            ).fetchall()

        else:  # keyword
            if re.sub(r'\D', '', term) == term and len(term) >= 6:
                rows = conn.execute(
                    'SELECT * FROM mys_data WHERE ic >= ? AND ic < ? LIMIT ?',
                    (term, term + '\xff', limit)
                ).fetchall()
            else:
                words = [w for w in re.sub(r'[^\w]', ' ', term.upper()).split() if len(w) >= 2]
                if not words:
                    return [], {'ok': False, 'message': 'Invalid search term.'}
                fts_q = ' '.join(f'"{w}"*' for w in words)
                rows = conn.execute(
                    'SELECT mys_data.* FROM mys_nama_fts JOIN mys_data ON mys_data.id = mys_nama_fts.rowid'
                    ' WHERE mys_nama_fts MATCH ? ORDER BY rank LIMIT ?',
                    (fts_q, limit)
                ).fetchall()

        results   = [dict(r) for r in rows]
        truncated = len(results) == limit
        meta = {
            'ok': True,
            'count': len(results),
            'truncated': truncated,
            'limit': limit,
            'message': f'Found {len(results)} record(s)' + (f' (showing first {limit})' if truncated else ''),
        }
        _mys_cache_set(search_type, term, results, meta)
        return results, meta
    except Exception as e:
        return [], {'ok': False, 'message': f'Search error: {str(e)}'}
    finally:
        conn.close()


def _download_alert(user, source, download_type, record_count, search_type, search_term):
    """Send Telegram alert if agent hits download thresholds (covers MYS + JPPH)."""
    token   = SystemSettings.get('telegram_bot_token')
    chat_id = SystemSettings.get('telegram_boss_chat_id')
    if not token or not chat_id:
        return
    cutoff_1h = datetime.utcnow() - timedelta(hours=1)
    stats = db.session.query(
        db.func.count(MysDownloadLog.id).label('cnt'),
        db.func.sum(MysDownloadLog.record_count).label('total')
    ).filter(
        MysDownloadLog.user_id == user.id,
        MysDownloadLog.timestamp >= cutoff_1h
    ).first()
    dl_count = stats.cnt   or 0
    dl_total = stats.total or 0

    # Thresholds: > 5 downloads OR > 500 records in last 1 hour
    if dl_count > 5 or dl_total > 500:
        source_label = 'MYS' if source == 'mys' else 'Transacted Data (JPPH)'
        try:
            text = (
                f"⚠️ *Download Alert — {source_label}*\n\n"
                f"Agent *{user.full_name}* (`{user.username}`) is downloading heavily.\n\n"
                f"📥 This download: *{record_count} record(s)* ({download_type})\n"
                f"🔍 Search: `{search_type}` → `{search_term}`\n\n"
                f"📊 Combined last 1 hour (MYS + JPPH):\n"
                f"  • Downloads: *{dl_count}* times\n"
                f"  • Total records: *{dl_total}*\n\n"
                f"Review at /admin/mys-downloads"
            )
            http_req.post(
                f"https://api.telegram.org/bot{token}/sendMessage",
                json={"chat_id": chat_id, "text": text, "parse_mode": "Markdown"},
                timeout=10
            )
        except Exception as e:
            app.logger.error(f"[TELEGRAM] Download alert error: {e}")


def _log_agent_download(source, download_type, search_type, search_term, record_count):
    """Shared helper — log a download and fire alert if thresholds exceeded."""
    ip_address = get_client_ip()
    log = MysDownloadLog(
        user_id       = current_user.id,
        download_type = download_type,
        search_type   = search_type,
        search_term   = search_term,
        record_count  = record_count,
        ip_address    = ip_address,
        source        = source
    )
    db.session.add(log)
    db.session.commit()
    _download_alert(current_user, source, download_type, record_count, search_type, search_term)


@app.route('/agent/mys/log-download', methods=['POST'])
@login_required
def agent_mys_log_download():
    if current_user.role != 'agent':
        return jsonify({'ok': False}), 403
    data = request.get_json(silent=True) or {}
    _log_agent_download(
        source        = 'mys',
        download_type = data.get('download_type', 'bulk'),
        search_type   = data.get('search_type', '')[:20],
        search_term   = data.get('search_term', '')[:200],
        record_count  = int(data.get('record_count', 1))
    )
    return jsonify({'ok': True})


@app.route('/agent/transacted-data/log-download', methods=['POST'])
@login_required
def agent_jpph_log_download():
    if current_user.role != 'agent':
        return jsonify({'ok': False}), 403
    data = request.get_json(silent=True) or {}
    _log_agent_download(
        source        = 'jpph',
        download_type = data.get('download_type', 'bulk'),
        search_type   = data.get('search_type', '')[:20],
        search_term   = data.get('search_term', '')[:200],
        record_count  = int(data.get('record_count', 1))
    )
    return jsonify({'ok': True})


@app.route('/admin/mys-downloads')
@login_required
@admin_required
def admin_mys_downloads():
    logs = (MysDownloadLog.query
            .order_by(MysDownloadLog.timestamp.desc())
            .limit(200).all())
    return render_template('admin_mys_downloads.html', logs=logs)


@app.route('/admin/mys')
@login_required
@admin_required
def admin_mys():
    return render_template('admin_mys.html')


@app.route('/admin/mys/search', methods=['POST'])
@login_required
@admin_required
def admin_mys_search():
    search_type = request.form.get('search_type', 'keyword')
    term = request.form.get('term', '').strip()
    results, meta = search_mys(search_type, term)
    return jsonify({'results': results, 'meta': meta})


@app.route('/agent/mys')
@login_required
def agent_mys():
    if current_user.role != 'agent':
        return redirect(url_for('admin_dashboard'))
    return render_template('agent_mys.html')


@app.route('/agent/mys/search', methods=['POST'])
@login_required
def agent_mys_search():
    if current_user.role != 'agent':
        return jsonify({'ok': False, 'message': 'Unauthorized'}), 403
    search_type = request.form.get('search_type', 'keyword')
    term = request.form.get('term', '').strip()
    results, meta = search_mys(search_type, term)
    return jsonify({'results': results, 'meta': meta})


@app.route('/admin/director')
@login_required
@admin_required
def admin_director():
    total_count = CustomerData.query.filter(CustomerData.data_source == 'Director').count()
    return render_template('admin_director.html', total_count=total_count)


@app.route('/admin/director-search', methods=['POST'])
@login_required
@admin_required
def admin_director_search():
    search_type = request.form.get('search_type', 'name')
    term = request.form.get('search_term', '').strip()
    if not term or len(term) < 2:
        return jsonify({'ok': False, 'message': 'Please enter at least 2 characters.'})

    limit = 500
    query = CustomerData.query.filter(CustomerData.data_source == 'Director')

    if search_type == 'name':
        query = query.filter(CustomerData.name.ilike(f'%{term}%'))
    elif search_type == 'phone':
        query = query.filter(CustomerData.contact_number.ilike(f'%{term}%'))
    elif search_type == 'address':
        query = query.filter(or_(
            CustomerData.address.ilike(f'%{term}%'),
            CustomerData.additional_data.ilike(f'%{term}%'),
        ))

    rows = query.order_by(CustomerData.id.desc()).limit(limit + 1).all()
    truncated = len(rows) > limit
    rows = rows[:limit]

    import json as _json
    results = []
    for r in rows:
        try:
            add = _json.loads(r.additional_data) if r.additional_data else {}
        except Exception:
            add = {}
        results.append({
            'id': r.id,
            'name': r.name or '',
            'contact_number': r.contact_number or '',
            'address': r.address or '',
            'email': r.email or '',
            'additional_data': add,
        })

    return jsonify({'ok': True, 'results': results, 'truncated': truncated})


@app.route('/agent/director')
@login_required
def agent_director():
    if current_user.role != 'agent':
        return redirect(url_for('admin_dashboard'))
    return render_template('agent_director.html')


@app.route('/agent/director-search', methods=['POST'])
@login_required
def agent_director_search():
    if current_user.role != 'agent':
        return jsonify({'ok': False, 'message': 'Unauthorized'}), 403
    search_type = request.form.get('search_type', 'name')
    term = request.form.get('search_term', '').strip()
    if not term or len(term) < 2:
        return jsonify({'ok': False, 'message': 'Please enter at least 2 characters.'})

    limit = 500
    query = CustomerData.query.filter(CustomerData.data_source == 'Director')

    if search_type == 'name':
        query = query.filter(CustomerData.name.ilike(f'%{term}%'))
    elif search_type == 'phone':
        query = query.filter(CustomerData.contact_number.ilike(f'%{term}%'))
    elif search_type == 'address':
        query = query.filter(or_(
            CustomerData.address.ilike(f'%{term}%'),
            CustomerData.additional_data.ilike(f'%{term}%'),
        ))

    rows = query.order_by(CustomerData.id.desc()).limit(limit + 1).all()
    truncated = len(rows) > limit
    rows = rows[:limit]

    import json as _json
    results = []
    for r in rows:
        try:
            add = _json.loads(r.additional_data) if r.additional_data else {}
        except Exception:
            add = {}
        results.append({
            'id': r.id,
            'name': r.name or '',
            'contact_number': r.contact_number or '',
            'address': r.address or '',
            'email': r.email or '',
            'additional_data': add,
        })

    return jsonify({'ok': True, 'results': results, 'truncated': truncated})


# ==================== UNIFIED SEARCH ====================

def _unified_row_customer(r):
    import json as _j
    try:
        add = r.additional_data
        if isinstance(add, str):
            add = _j.loads(add)
    except Exception:
        add = {}
    return {
        'name': getattr(r, 'name', '') or '',
        'contact_number': getattr(r, 'contact_number', '') or '',
        'address': getattr(r, 'address', '') or '',
        'email': getattr(r, 'email', '') or '',
        'additional_data': add if isinstance(add, dict) else {},
        'source': 'Customer',
    }

def _unified_search_customer(term):
    try:
        seen_ids = set()
        out = []
        for stype in ('name', 'address', 'keyword'):
            rows, meta = perform_customer_search(stype, term, requested_limit=50)
            for r in rows:
                rid = getattr(r, 'id', None) or id(r)
                if rid in seen_ids:
                    continue
                seen_ids.add(rid)
                out.append(_unified_row_customer(r))
                if len(out) >= 50:
                    break
            if len(out) >= 50:
                break
        return out
    except Exception as e:
        app.logger.error(f'[UNIFIED] customer error: {e}')
        return []

def _unified_search_director(term):
    try:
        import json as _j
        rows = (CustomerData.query
                .filter(CustomerData.data_source == 'Director')
                .filter(or_(
                    CustomerData.name.ilike(f'%{term}%'),
                    CustomerData.address.ilike(f'%{term}%'),
                    CustomerData.additional_data.ilike(f'%{term}%'),
                ))
                .order_by(CustomerData.id.desc())
                .limit(50).all())
        out = []
        for r in rows:
            try:
                add = _j.loads(r.additional_data) if r.additional_data else {}
            except Exception:
                add = {}
            out.append({
                'name': r.name or '',
                'contact_number': r.contact_number or '',
                'address': r.address or '',
                'email': r.email or '',
                'additional_data': add,
                'source': 'Director',
            })
        return out
    except Exception as e:
        app.logger.error(f'[UNIFIED] director error: {e}')
        return []

def _unified_search_jpph(term):
    try:
        seen = set()
        out = []
        for stype, lim in [('area', 30), ('vendor', 15), ('purchaser', 15)]:
            rows, _ = search_jpph(stype, term, limit=lim)
            for r in rows:
                key = (r.get('vendor',''), r.get('purchaser',''), r.get('address',''))
                if key in seen:
                    continue
                seen.add(key)
                out.append({
                    'address': r.get('address', '') or '',
                    'vendor': r.get('vendor', '') or '',
                    'purchaser': r.get('purchaser', '') or '',
                    'scheme': r.get('scheme', '') or '',
                    'mukim_town': r.get('mukim_town', '') or '',
                    'district': r.get('district', '') or '',
                    'source': 'JPPH',
                })
                if len(out) >= 50:
                    break
        return out
    except Exception as e:
        app.logger.error(f'[UNIFIED] jpph error: {e}')
        return []

def _unified_search_mys(term):
    try:
        seen_ic = set()
        out = []

        # Name FTS search
        name_results, meta = search_mys('name', term, limit=30)
        if meta.get('ok'):
            for r in name_results:
                ic = r.get('ic', '')
                seen_ic.add(ic)
                out.append({
                    'name': r.get('nama', '') or '',
                    'ic': ic or '',
                    'tel': r.get('tel', '') or '',
                    'lokaliti': r.get('lokaliti', '') or '',
                    'alamat': r.get('alamat', '') or '',
                    'negeri': r.get('negeri', '') or '',
                    'source': 'MYS',
                })

        # Location search — lokaliti and alamat LIKE %term%
        db_path = _mys_db_path()
        if os.path.exists(db_path):
            conn = _mys_connect()
            try:
                loc_rows = conn.execute(
                    'SELECT * FROM mys_data WHERE lokaliti LIKE ? OR alamat LIKE ? LIMIT 30',
                    (f'%{term}%', f'%{term}%')
                ).fetchall()
                for r in loc_rows:
                    r = dict(r)
                    ic = r.get('ic', '')
                    if ic in seen_ic:
                        continue
                    seen_ic.add(ic)
                    out.append({
                        'name': r.get('nama', '') or '',
                        'ic': ic or '',
                        'tel': r.get('tel', '') or '',
                        'lokaliti': r.get('lokaliti', '') or '',
                        'alamat': r.get('alamat', '') or '',
                        'negeri': r.get('negeri', '') or '',
                        'source': 'MYS',
                    })
            finally:
                conn.close()

        return out[:50]
    except Exception as e:
        app.logger.error(f'[UNIFIED] mys error: {e}')
        return []

@app.route('/admin/unified-search', methods=['POST'])
@app.route('/agent/unified-search', methods=['POST'])
@login_required
def unified_search():
    from concurrent.futures import ThreadPoolExecutor, as_completed
    data = request.get_json(silent=True) or {}
    term = _normalize_search_term(data.get('name', ''))
    if len(term) < 2:
        return jsonify({'ok': False, 'message': 'Enter at least 2 characters.'})

    tasks = {
        'customer': _unified_search_customer,
        'director': _unified_search_director,
        'jpph':     _unified_search_jpph,
        'mys':      _unified_search_mys,
    }

    results = {}
    with ThreadPoolExecutor(max_workers=4) as pool:
        futures = {pool.submit(fn, term): key for key, fn in tasks.items()}
        for future in as_completed(futures):
            key = futures[future]
            try:
                results[key] = future.result()
            except Exception as e:
                app.logger.error(f'[UNIFIED] {key} future error: {e}')
                results[key] = []

    total = sum(len(v) for v in results.values())
    return jsonify({
        'ok': True,
        'query': term,
        'total_hits': total,
        'results': results,
    })

@app.route('/api/search-feedback', methods=['POST'])
@login_required
def save_search_feedback():
    import json as _j
    data = request.get_json(silent=True) or {}
    fb = SearchFeedback(
        user_id=current_user.id,
        query=data.get('query', '')[:500],
        result_counts=_j.dumps(data.get('counts', {})),
        rating=data.get('rating', ''),
        comment=(data.get('comment', '') or '')[:1000],
    )
    db.session.add(fb)
    db.session.commit()
    return jsonify({'ok': True})

# ==================== MAIN ENTRY POINT ====================
if __name__ == '__main__':
    # Use port from environment variable for Render
    port = int(os.environ.get('PORT', 5000))
    
    # Run with debug=False in production
    if IS_RENDER:
        app.run(host='0.0.0.0', port=port, debug=False)
        print(f"[STARTUP] Running in production mode on port {port}")
    else:
        app.run(debug=True, host='0.0.0.0', port=port)
        print(f"[STARTUP] Running in development mode on port {port}")
